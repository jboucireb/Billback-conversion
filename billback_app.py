#!/usr/bin/env python3
"""Monin Billback Processor v2 — exact Tellus Sheet1 format"""

import os, re, io, json, threading, webbrowser, email, uuid, shutil, tempfile, traceback
from http.server import HTTPServer, BaseHTTPRequestHandler
from datetime import datetime
import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font
import warnings
warnings.filterwarnings('ignore')

TODAY = datetime.now().strftime('%Y%m%d')

# ─── MONIN ITEM LOOKUP TABLE (code → description) ───────────────────────────
ITEM_LOOKUP = {
    'COUPON1': 'Promotional Web Coupon',
    'KK-FR079F': 'Krispy Kreme Original Glazed Doughnut 4pk-1L',
    'LBLNA00': '1 Pump 1 Packet 750/1L Neck',
    'M-AD009B': 'Natural Zero Caramel 6pk-750ml',
    'M-AD009U': 'Natural Zero Caramel 750ml',
    'M-AD023B': 'Natural Zero Hazelnut 6pk-750mL',
    'M-AD023U': 'Natural Zero Hazelnut 750ml',
    'M-AD040B': 'Natural Zero Raspberry 6pk-750ml',
    'M-AD040U': 'Natural Zero Raspberry 750ml',
    'M-AD045B': 'Natural Zero Vanilla 6pk-750mL',
    'M-AD045U': 'Natural Zero Vanilla 750ml',
    'M-AD050B': 'Natural Zero Peppermint 6pk-750ml',
    'M-AD050U': 'Natural Zero Peppermint 750ml',
    'M-AD083B': 'Natural Zero Sweetener 6pk-750mL',
    'M-AD083U': 'Natural Zero Sweetener 750ml',
    'M-AD105B': 'Natural Zero Pumpkin Spice 6pk-750ml',
    'M-AD105U': 'Natural Zero Pumpkin Spice 750ml',
    'M-AD196B': 'Natural Zero Chocolate 6pk-750ml',
    'M-AD196U': 'Natural Zero Chocolate 750ml',
    'M-AO009B': 'Organic Caramel 6pk-750mL',
    'M-AO009U': 'Organic Caramel 750mL',
    'M-AO023B': 'Organic Hazelnut 6pk-750mL',
    'M-AO023U': 'Organic Hazelnut 750mL',
    'M-AO040B': 'Organic Raspberry 6pk-750mL',
    'M-AO040U': 'Organic Raspberry750mL',
    'M-AO045B': 'Organic Vanilla 6pk-750mL',
    'M-AO045U': 'Organic Vanilla 750mL',
    'M-AO062B': 'Organic Chocolate 6pk-750mL',
    'M-AO062U': 'Organic Chocolate 750mL',
    'M-AO157B': 'Organic Agave Nectar 6pk-750mL',
    'M-AO157U': 'Organic Agave 750mL',
    'M-AR000A': 'Pure Cane Syrup 12pk-750mL',
    'M-AR000U': 'Pure Cane Syrup 750 mL',
    'M-AR001A': 'Almond 12pk-750mL',
    'M-AR001U': 'Almond 750 mL',
    'M-AR003A': 'Apple 12pk-750mL',
    'M-AR003U': 'Apple 750 mL',
    'M-AR005A': 'Blackcurrant 12pk-750mL GL',
    'M-AR005U': 'Blackcurrant 750 mL',
    'M-AR006A': 'Blackberry 12pk-750mL',
    'M-AR006U': 'Blackberry 750 mL',
    'M-AR007A': 'Blue Curacao 12pk-750mL',
    'M-AR007U': 'Blue Curacao 750 mL',
    'M-AR008A': 'Blueberry 12pk-750mL',
    'M-AR008U': 'Blueberry 750 mL',
    'M-AR009A': 'Caramel 12pk-750mL',
    'M-AR009U': 'Caramel 750 mL',
    'M-AR010A': 'Cherry 12pk-750mL',
    'M-AR010U': 'Cherry 750 mL',
    'M-AR012A': 'Cinnamon 12pk-750mL',
    'M-AR012U': 'Cinnamon 750 mL',
    'M-AR013A': 'Coconut 12pk-750mL',
    'M-AR013U': 'Coconut 750 mL',
    'M-AR014A': 'Espresso 12pk-750mL',
    'M-AR014U': 'Espresso 750 mL',
    'M-AR015A': 'Cranberry 12pk-750mL',
    'M-AR015U': 'Cranberry 750 mL',
    'M-AR016A': 'Frosted Mint 12pk-750mL',
    'M-AR016U': 'Frosted Mint 750mL',
    'M-AR018A': 'Ginger 12pk-750mL',
    'M-AR018U': 'Ginger 750 mL',
    'M-AR021A': 'Green Mint 12pk-750mL',
    'M-AR021U': 'Green Mint 750 mL',
    'M-AR022A': 'Grenadine 12pk-750mL',
    'M-AR022U': 'Grenadine 750 mL',
    'M-AR023A': 'Hazelnut 12pk-750mL',
    'M-AR023U': 'Hazelnut 750 mL',
    'M-AR025A': 'Irish Cream 12pk-750mL',
    'M-AR025U': 'Irish Cream 750 mL',
    'M-AR027A': 'Kiwi 12pk-750mL',
    'M-AR027U': 'Kiwi 750 mL',
    'M-AR028A': 'Lemon 12pk-750mL',
    'M-AR028U': 'Lemon 750 mL',
    'M-AR029A': 'Lime 12pk-750mL',
    'M-AR029U': 'Lime 750 mL',
    'M-AR031A': 'Mandarin 12pk-750mL',
    'M-AR031U': 'Mandarin 750 mL',
    'M-AR032A': 'Mango 12pk-750mL',
    'M-AR032U': 'Mango 750 mL',
    'M-AR034A': 'Orange 12pk-750mL',
    'M-AR034U': 'Orange 750 mL',
    'M-AR035A': 'Passion Fruit 12pk-750mL',
    'M-AR035U': 'Passion Fruit 750 mL',
    'M-AR036A': 'Peach 12pk-750mL',
    'M-AR036U': 'Peach 750 mL',
    'M-AR037A': 'Pear 12pk-750mL',
    'M-AR037U': 'Pear 750 mL',
    'M-AR038A': 'Pineapple 12pk-750mL',
    'M-AR038U': 'Pineapple 750 mL',
    'M-AR039A': 'Pistachio 12 Pk-750 mL',
    'M-AR039U': 'Pistachio 750 mL',
    'M-AR040A': 'Raspberry 12pk-750mL',
    'M-AR040U': 'Raspberry 750 mL',
    'M-AR042A': 'Strawberry 12pk-750mL',
    'M-AR042U': 'Strawberry 750 mL',
    'M-AR043A': 'Swiss Chocolate 12pk-750mL',
    'M-AR043U': 'Swiss Chocolate 750ml',
    'M-AR045A': 'Vanilla 12pk-750mL',
    'M-AR045U': 'Vanilla 750 mL',
    'M-AR046A': 'Banana 12pk-750mL',
    'M-AR046U': 'Banana 750 mL',
    'M-AR047A': 'Amaretto 12pk-750mL',
    'M-AR047U': 'Amaretto 750 mL',
    'M-AR048A': 'Macadamia Nut 12pk-750mL',
    'M-AR048U': 'Macadamia Nut 750 mL',
    'M-AR049A': 'Granny Smith Apple 12pk-750mL',
    'M-AR049U': 'Granny Smith Apple 750 mL',
    'M-AR050A': 'Peppermint 12pk-750mL',
    'M-AR050U': 'Peppermint 750 mL',
    'M-AR051A': 'Toffee Nut 12pk-750mL',
    'M-AR051U': 'Toffee Nut Syrup 750 mL',
    'M-AR053A': 'Mojito Mix 12pk-750mL',
    'M-AR053U': 'Mojito Mix 750 mL',
    'M-AR055A': 'Violet 12pk-750mL',
    'M-AR055U': 'Violet 750 mL',
    'M-AR056A': 'Rose 12pk-750mL',
    'M-AR056U': 'Rose 750 mL',
    'M-AR058A': 'Toasted Almond Mocha 12pk-750mL',
    'M-AR058U': 'Toasted Almond Mocha 750 mL',
    'M-AR060A': 'Gingerbread 12pk-750mL',
    'M-AR060U': 'Gingerbread 750 mL',
    'M-AR061A': 'Lavender 12pk-750mL',
    'M-AR061U': 'Lavender 750 mL',
    'M-AR062A': 'Dark Chocolate 12pk-750mL',
    'M-AR062U': 'Dark Chocolate 750 mL',
    'M-AR063A': 'White Chocolate 12pk-750mL',
    'M-AR063U': 'White Chocolate 750 mL',
    'M-AR064A': 'Praline 12pk-750mL',
    'M-AR064U': 'Praline 750 mL',
    'M-AR066A': 'Guava 12pk-750mL',
    'M-AR066U': 'Guava 750 mL',
    'M-AR075A': 'Pomegranate 12pk-750mL',
    'M-AR075U': 'Pomegranate 750 mL',
    'M-AR084A': 'Honey Syrup 12pk-750ml',
    'M-AR084U': 'Honey Syrup 750ml',
    'M-AR087A': 'Candied Orange 12pk-750mL',
    'M-AR087U': 'Candied Orange 750 mL',
    'M-AR101A': 'White Sangria Mix 12pk-750mL',
    'M-AR101U': 'White Sangria Mix 750 mL',
    'M-AR105A': 'Pumpkin Spice 12pk-750mL',
    'M-AR105U': 'Pumpkin Spice 750 mL',
    'M-AR111A': 'Sugarcane Cola 12pk-750mL',
    'M-AR111U': 'Sugarcane Cola 750ml',
    'M-AR113A': 'Maple Spice 12pk-750mL',
    'M-AR113U': 'Maple Spice 750 mL',
    'M-AR133A': 'Huckleberry 12pk-750mL',
    'M-AR133U': 'Huckleberry 750ml',
    'M-AR145A': 'Toasted Marshmallow 12pk-750mL',
    'M-AR145U': 'Toasted Marshmallow 750 mL',
    'M-AR147A': 'Elderflower 12pk-750mL',
    'M-AR147U': 'Elderflower 750 mL',
    'M-AR148A': 'White Peach 12pk-750mL',
    'M-AR148U': 'White Peach Syrup 750 mL',
    'M-AR150A': 'French Raspberry 12pk-750mL',
    'M-AR150U': 'French Raspberry 750ml',
    'M-AR190A': 'French Vanilla 12pk-750mL',
    'M-AR190U': 'French Vanilla 750 mL',
    'M-AR193A': 'Roasted Hazelnut 12pk-750mL',
    'M-AR193U': 'Roasted Hazelnut 750 mL',
    'M-AR197A': 'Spiced Brown Sugar 12pk-750mL',
    'M-AR197U': 'Spiced Brown Sugar 750ml',
    'M-AR210A': 'Salted Caramel 12pk-750mL',
    'M-AR210U': 'Salted Caramel 750ml',
    'M-AR212A': 'Cinnamon Bun 12pk-750mL',
    'M-AR212U': 'Cinnamon Bun 750ml',
    'M-AR226A': 'Old Fashioned Root Beer 12pk-750mL',
    'M-AR228A': 'Chocolate Fudge 12pk-750mL',
    'M-AR228U': 'Chocolate Fudge 750ml',
    'M-AR238A': 'Stone Fruit 12pk-750ml',
    'M-AR238U': 'Stone Fruit 750ml',
    'M-AR241A': 'Pumpkin Pie 12pk-750mL',
    'M-AR241U': 'Pumpkin Pie 750ml',
    'M-AR246A': 'Hawaiian Island 12pk-750ml',
    'M-AR246U': 'Hawaiian Island 750ml',
    'M-AR247A': 'Cookie Butter 12pk-750ml',
    'M-AR247U': 'Cookie Butter 750ml',
    'M-AR255A': 'Classic Watermelon 12pk-750ml',
    'M-AR255U': 'Classic Watermelon 750ml',
    'M-AR256A': 'Tiramisu 12pk-750mL',
    'M-AR256U': 'Tiramisu 750ml',
    'M-AR258A': 'Butter Pecan 12pk-750mL',
    'M-AR258U': 'Butter Pecan 750ml',
    'M-AR261A': 'Peanut Butter 12pk-750mL',
    'M-AR261U': 'Peanut Butter 750ml',
    'M-AR268A': 'Caramel Apple Butter 12pk-750mL',
    'M-AR268U': 'Caramel Apple Butter 750ml',
    'M-AR274A': 'Brown Butter 12pk-750mL',
    'M-AR274U': 'Brown Butter 750ml',
    'M-AR275A': 'Brown Butter Toffee 12pk-750mL',
    'M-AR275U': 'Brown Butter Toffee 750ml',
    'M-AR276A': 'Maple Pancake 12pk-750ml',
    'M-AR276U': 'Maple Pancake 750ml',
    'M-AR278A': 'Vanilla Spice 12pk-750ml',
    'M-AR290A': 'Golden Turmeric 12pk-750ml',
    'M-AR290U': 'Golden Turmeric 750ml',
    'M-AR318A': 'Strawberry Rose 12pk-750ml',
    'M-AR318U': 'Strawberry Rose 750ml',
    'M-AR319A': 'Honey Jasmine 12pk-750ml',
    'M-AR319U': 'Honey Jasmine 750ml',
    'M-AR321A': 'Lavender Lemon 12pk-750ml',
    'M-AR321U': 'Lavender Lemon 750ml',
    'M-AR361A': 'Ube 12pk-750ml',
    'M-AR361U': 'Ube 750ml',
    'M-AR400A': 'Toasted Coconut 12pk-750ml',
    'M-AR400U': 'Toasted Coconut 750ml',
    'M-AS009A': 'Sugar Free Caramel 12pk-750mL',
    'M-AS009U': 'Sugar Free Caramel 750 mL',
    'M-AS013A': 'Sugar Free Coconut 12pk-750ml',
    'M-AS013U': 'Sugar Free Coconut 750ml',
    'M-AS023A': 'Sugar Free Hazelnut 12pk-750mL',
    'M-AS023U': 'Sugar Free Hazelnut 750 mL',
    'M-AS025A': 'Sugar Free Irish Cream 12pk-750mL',
    'M-AS025U': 'Sugar Free Irish Cream 750 mL',
    'M-AS036A': 'Sugar Free Peach 12pk-750mL',
    'M-AS036U': 'Sugar Free Peach 750 mL',
    'M-AS040A': 'Sugar Free Raspberry 12pk-750mL',
    'M-AS040U': 'Sugar Free Raspberry 750 mL',
    'M-AS042A': 'Sugar Free Strawberry 12pk-750mL',
    'M-AS042U': 'Sugar Free Strawberry 750 mL',
    'M-AS045A': 'Sugar Free Vanilla 12pk-750mL',
    'M-AS045U': 'Sugar Free Vanilla 750 mL',
    'M-AS061A': 'Sugar Free Lavender 12pk-750ml',
    'M-AS061U': 'Sugar Free Lavender 750ml',
    'M-AS062A': 'Sugar Free Chocolate 12pk-750ml',
    'M-AS062U': 'Sugar Free Chocolate 750mL',
    'M-AS063A': 'Sugar Free White Chocolate 12pk-750mL',
    'M-AS063U': 'Sugar Free White Chocolate 750 mL',
    'M-AT028A': 'Lemon Tea 12pk-750mL',
    'M-AT028U': 'Lemon Tea 750 mL',
    'M-AT032A': 'Mango Tea 12pk-750mL',
    'M-AT032U': 'Mango Tea 750 mL',
    'M-AT036A': 'Peach Tea 12pk-750mL',
    'M-AT036U': 'Peach Tea 750 mL',
    'M-AT040A': 'Raspberry Tea 12pk-750mL',
    'M-AT040U': 'Raspberry Tea 750 mL',
    'M-AT080A': 'Chai Tea 12pk-750mL',
    'M-AT080U': 'Chai Tea Concentrate 750 mL',
    'M-AT366A': 'Matcha Green Tea 12pk-750ml',
    'M-AT366U': 'Matcha Green Tea 750ml',
    'M-AX253B': 'HomeCrafted Margarita Mix 6pk-750ml',
    'M-AX253U': 'HomeCrafted Margarita Mix 750ml',
    'M-AX298B': 'HomeCrafted Dragon Fruit Cosmo Mix 6pk-750ml',
    'M-AX298U': 'HomeCrafted Dragon Fruit Cosmo Mix 750ml',
    'M-AX312B': 'HomeCrafted Strawberry Ginger Lmnde Mix 6pk-750ml',
    'M-AX312U': 'HomeCrafted Strawberry Ginger Lmnde Mix 750ml',
    'M-AX317B': 'HomeCrafted Mai Tai Mix 6pk-750ml',
    'M-AX317U': 'HomeCrafted Mai Tai Mix 750ml',
    'M-AX331B': 'HomeCrafted Cherry Smash Mix 6pk-750ml',
    'M-AX331U': 'HomeCrafted Cherry Smash Mix 750ml',
    'M-AX343B': 'HomeCrafted Blackberry Mint Lemonade Mix 6pk-750ml',
    'M-AX343U': 'HomeCrafted Blackberry Mint Lemonade Mix 750ml',
    'M-AX344B': 'HomeCrafted Spicy Watermln Margarita Mix 6pk-750ml',
    'M-AX344U': 'HomeCrafted Spicy Watermelon Margarita Mix 750ml',
    'M-AX351B': 'HomeCrafted Prickly Pear Margarita 6pk-750ml',
    'M-AX351U': 'HomeCrafted Prickly Pear Margarita 750ml',
    'M-DJ059U': 'Watermelon Conc Flavor 55 Gal. Drum',
    'M-DJ059U2': 'PD Watermelon Conc Flvr 55 Gal. Drum-Orig Color',
    'M-DJ188U': 'Mint Conc Flavor 55 Gal. Drum',
    'M-DJ292U': 'Lime Concentrated Flavor 55 Gal. Drum',
    'M-DR000U': 'Pure Cane 55 Gal. Drum',
    'M-DR006U': 'Blackberry 55Gal Drum',
    'M-DR013U': 'Coconut 55 Gal. Drum',
    'M-DR019U': 'Ruby Red Grapefruit 55 Gal. Drum',
    'M-DR023U': 'Hazelnut 55 Gal Drum',
    'M-DR025U': 'Irish Cream 55 Gal. Drum',
    'M-DR032U': 'Mango 55 Gal. Drum',
    'M-DR036U': 'Peach 55 Gal. Drum',
    'M-DR038U': 'Pineapple 55 Gal. Drum',
    'M-DR039U2': 'Pistachio TTB',
    'M-DR040U': 'Raspberry 55 Gal. Drum',
    'M-DR042U': 'Strawberry 55 Gal. Drum',
    'M-DR045U': 'Vanilla 55 Gal. Drum',
    'M-DR049U': 'Granny Smith Apple 55 Gal. Dru',
    'M-DR050U': 'Peppermint 55 Gal. Drum',
    'M-DR059U': 'Watermelon 55 Gal. Drum',
    'M-DR061U': 'Lavender 55 Gal. Drum',
    'M-DR063U': 'White Chocolate 55 Gal Drum',
    'M-DR072U': 'Wild Strawberry 55 Gal. Drum',
    'M-DR075U': 'Pomegranate 55 Gal. Drum',
    'M-DR095U': 'Cucumber 55 Gal. Drum',
    'M-DR105U': 'Pumpkin Spice 55 Gal Drum',
    'M-DR112U': 'Butterscotch 55 Gal. Drum',
    'M-DR136U': 'Wild Raspberry 55 Gal. Drum',
    'M-DR144U': 'Wild Grape 55 Gal. Drum',
    'M-DR180U': 'Spicy Red Cinnamon 55 Gal Drum',
    'M-DR190U': 'French Vanilla 55 Gal. Drum',
    'M-DR195U': 'Blue Raspberry 55 Gal. Drum',
    'M-DR210U': 'Salted Caramel 55Gal Drum',
    'M-DR212U': 'Cinnamon Bun 55 Gal. Drum',
    'M-DR224U': 'Vanilla Creme 55 Gal. Drum',
    'M-DR228U': 'Chocolate Fudge 55 Gal. Drum',
    'M-DR268U': 'Caramel Apple Butter 55 Gal. Drum',
    'M-DR298U': 'Dragon Fruit 55 Gal. Drum',
    'M-DR318U': 'Strawberry Rose 55 Gal. Drum',
    'M-DT036U': 'Peach Tea 55 Gal. Drum',
    'M-DX326U': 'Energy Boost 55 Gal. Drum',
    'M-EG032B': 'Mango Smoothie 6pk-46oz',
    'M-EG032U': 'Mango Smoothie 46oz',
    'M-EG036B': 'Peach Smoothie 6pk-46oz',
    'M-EG036U': 'Peach Smoothie 46oz',
    'M-EG042B': 'Strawberry Smoothie 6pk-46oz',
    'M-EG042U': 'Strawberry Smoothie 46oz',
    'M-EG114B': 'Wildberry Smoothie 6pk-46oz',
    'M-EG183B': 'Pina Colada Smoothie 6pk-46oz',
    'M-EG183U': 'Pina Colada Smoothie 46oz',
    'M-EG207B': 'Strawberry Banana Smoothie 6pk-46oz',
    'M-EG207U': 'Strawberry Banana Smoothie 46oz',
    'M-FD009F': 'Natural Zero Caramel 4pk-1L',
    'M-FD009U': 'Natural Zero Caramel 1L',
    'M-FD023F': 'Natural Zero Hazelnut 4pk-1L',
    'M-FD023U': 'Natural Zero Hazelnut 1L',
    'M-FD036F': 'Natural Zero Peach 4pk-1L',
    'M-FD036U': 'Natural Zero Peach 1L',
    'M-FD040F': 'Natural Zero Raspberry 4pk-1L',
    'M-FD040U': 'Natural Zero Raspberry 1L',
    'M-FD045F': 'Natural Zero Vanilla 4pk-1L',
    'M-FD045U': 'Natural Zero Vanilla 1L',
    'M-FD083F': 'Natural Zero Sweetener 4pk-1L',
    'M-FD083U': 'Natural Zero Sweetener 1L',
    'M-FD210F': 'Natural Zero Salted Caramel 4pk-1L',
    'M-FD388F': 'Natural Zero Salted Watermelon 4pk-1L',
    'M-FD389F': 'Natural Zero Passion Fruit Yuzu 4pk-1L',
    'M-FD389U': 'Natural Zero Passion Fruit Yuzu 1L',
    'M-FDF036F': 'Peach Flavoring Concentrate 4pk-1L',
    'M-FDF040F': 'Raspberry Flavoring Concentrate 4pk-1L',
    'M-FDF042F': 'Strawberry Flavoring Concentrate 4pk-1L',
    'M-FDF042U': 'Strawberry Flavoring Concentrate 1L',
    'M-FJ008FP': 'Blueberry Concentrated Flavor 4pk-1L',
    'M-FJ013FP': 'Coconut Concentrated Flavor 4pk-1L',
    'M-FJ098F': 'Energy Concentrate 4pk-1L',
    'M-FJ354FP': 'Cherry Berry Concentrated Flavor 4pk-1L',
    'M-FJ354U': 'Cherry Berry Concentrated Flavor 1L',
    'M-FJ355FP': 'Mango Passion Concentrated Flavor 4pk-1L',
    'M-FJ376FP': 'Key Lime Concentrated Flavor 4pk-1L V2',
    'M-FJ376U': 'Key Lime Concentrated Flavor 1L V2',
    'M-FJ387FP': 'Pineapple Dragonfruit Conc Flavor 4pk-1L',
    'M-FJ387U': 'Pineapple Dragonfruit Conc Flavor 1L',
    'M-FJ408FP': 'Strawberry Hibiscus Concentrated Flavor 4pk-1L',
    'M-FL084F': 'Organic Honey Sweetener 4pk-1L',
    'M-FL084U': 'Organic Honey Sweetener 1 L',
    'M-FL157A': 'Organic Agave Nectar 12pk-1L',
    'M-FL157A2': 'Organic Agave Nectar 12pk-1L',
    'M-FL157F': 'Organic Agave Nectar 4pk-1L',
    'M-FL157U': 'Agave Organic Nectar 1L',
    'M-FN008F': 'TruFlavour Blueberry 4pk-1L',
    'M-FN008U': 'TruFlavour Blueberry 1L',
    'M-FN013F': 'TruFlavour Coconut 4pk-1L',
    'M-FN013U': 'TruFlavour Coconut 1L',
    'M-FN032F': 'TruFlavour Mango 4pk-1L',
    'M-FN032U': 'TruFlavour Mango 1L',
    'M-FN035F': 'TruFlavour Passion Fruit 4pk-1L',
    'M-FN035U': 'TruFlavour Passion Fruit 1L',
    'M-FN036F': 'TruFlavour Peach 4pk-1L',
    'M-FN036U': 'TruFlavour Peach 1L',
    'M-FN042F': 'TruFlavour Strawberry 4pk-1L',
    'M-FN042U': 'TruFlavour Strawberry 1L',
    'M-FN059F': 'TruFlavour Watermelon 4pk-1L',
    'M-FN059U': 'TruFlavour Watermelon 1L',
    'M-FN095F': 'TruFlavour Cucumber 4pk-1L',
    'M-FN095U': 'TruFlavour Cucumber 1L',
    'M-FN177F': 'TruFlavour Hibiscus 4pk-1L',
    'M-FN177U': 'TruFlavour Hibiscus 1L',
    'M-FN188F': 'TruFlavour Mint 4pk-1L',
    'M-FN188U': 'TruFlavour Mint 1L',
    'M-FN235F': 'TruFlavour Basil 4pk-1L',
    'M-FN235U': 'TruFlavour Basil 1L',
    'M-FN260F': 'TruFlavour Jalapeno 4pk-1L',
    'M-FN260U': 'TruFlavour Jalapeno 1L',
    'M-FN262F': 'TruFlavour Oak Barrel 4pk-1L',
    'M-FN262U': 'TruFlavour Oak Barrel 1L',
    'M-FN397F': 'TruFlavour Pickle 4pk-1L',
    'M-FN397U': 'TruFlavour Pickle 1L',
    'M-FN398F': 'TruFlavour Mushroom 4pk-1L',
    'M-FN398U': 'TruFlavour Mushroom 1L',
    'M-FR000F': 'Pure Cane Syrup 4pk-1L',
    'M-FR000FP': 'Pure Cane Syrup 4pk-1L w/Pmp',
    'M-FR000FP2': 'Pure Cane Syrup 4pk-1L w/ Pump',
    'M-FR000U': 'Pure Cane Syrup 1 L',
    'M-FR001F': 'Almond 4pk-1L',
    'M-FR001U': 'Almond 1 L',
    'M-FR003F': 'Apple 4pk-1L',
    'M-FR003U': 'Apple 1 L',
    'M-FR006A': 'Blackberry 12pk-1L',
    'M-FR006F': 'Blackberry 4pk-1L',
    'M-FR006U': 'Blackberry 1 L',
    'M-FR007F': 'Blue Curacao 4pk-1L',
    'M-FR007U': 'Blue Curacao 1 L',
    'M-FR008F': 'Blueberry 4pk-1L',
    'M-FR008U': 'Blueberry 1 L',
    'M-FR009A': 'Caramel 12pk-1L',
    'M-FR009F': 'Caramel 4pk-1L',
    'M-FR009U': 'Caramel 1 L',
    'M-FR010F': 'Cherry 4pk-1L',
    'M-FR010U': 'Cherry 1 L',
    'M-FR012F': 'Cinnamon 4pk-1L',
    'M-FR012U': 'Cinnamon 1 L',
    'M-FR013F': 'Coconut 4pk-1L',
    'M-FR013U': 'Coconut 1 L',
    'M-FR014F': 'Espresso 4pk-1L',
    'M-FR014U': 'Espresso 1 L',
    'M-FR015F': 'Cranberry 4pk-1L',
    'M-FR015U': 'Cranberry 1 L',
    'M-FR016F': 'Frosted Mint 4pk-1L',
    'M-FR016U': 'Frosted Mint 1 L',
    'M-FR018F': 'Ginger 4pk-1L',
    'M-FR018U': 'Ginger 1 L',
    'M-FR019F': 'Ruby Red Grapefruit 4pk-1L',
    'M-FR019U': 'Ruby Red Grapefruit 1 L',
    'M-FR021F': 'Green Mint 4 Pk-1 L',
    'M-FR021U': 'Green Mint 1 L',
    'M-FR022A': 'Grenadine 12pk-1L',
    'M-FR022F': 'Grenadine 4pk-1L',
    'M-FR022U': 'Grenadine 1L',
    'M-FR023F': 'Hazelnut 4pk-1L',
    'M-FR023U': 'Hazelnut 1 L',
    'M-FR025F': 'Irish Cream 4pk-1L',
    'M-FR025U': 'Irish Cream 1 L',
    'M-FR027F': 'Kiwi 4pk-1L',
    'M-FR027U': 'Kiwi 1 L',
    'M-FR028F': 'Lemon 4pk-1L',
    'M-FR028U': 'Lemon 1 L',
    'M-FR029F': 'Lime 4pk-1L',
    'M-FR029FP': 'TB Lime 4pk-1L W/ Pump',
    'M-FR029U': 'Lime 1 L',
    'M-FR030F': 'Lychee 4pk-1L',
    'M-FR030U': 'Lychee 1 L',
    'M-FR032A': 'Mango 12pk-1L',
    'M-FR032F': 'Mango 4pk-1L',
    'M-FR032U': 'Mango 1 L',
    'M-FR034F': 'Orange 4pk-1L',
    'M-FR034U': 'Orange 1 L',
    'M-FR035F': 'Passion Fruit 4pk-1L',
    'M-FR035U': 'Passion Fruit 1 L',
    'M-FR036F': 'Peach 4pk-1L',
    'M-FR036U': 'Peach 1 L',
    'M-FR038F': 'Pineapple 4pk - 1L',
    'M-FR038FP': 'Pineapple 4pk-1L W/ Pump',
    'M-FR038U': 'Pineapple 1L',
    'M-FR039F': 'Pistachio 4pk-1L',
    'M-FR039U': 'Pistachio 1L',
    'M-FR040A': 'Raspberry 12pk-1L',
    'M-FR040F': 'Raspberry 4pk-1L',
    'M-FR040U': 'Raspberry 1 L',
    'M-FR042A': 'Strawberry 12pk-1L',
    'M-FR042F': 'Strawberry 4pk-1L',
    'M-FR042U': 'Strawberry 1 L',
    'M-FR043F': 'Swiss Chocolate 4pk-1L',
    'M-FR043U': 'Swiss Chocolate  1 L',
    'M-FR045A': 'Vanilla 12pk-1L',
    'M-FR045F': 'Vanilla 4pk-1L',
    'M-FR045U': 'Vanilla 1 L',
    'M-FR046F': 'Banana 4pk-1L',
    'M-FR046U': 'Banana 1 L',
    'M-FR047F': 'Amaretto 4pk - 1L',
    'M-FR047U': 'Amaretto 1L',
    'M-FR048F': 'Macadamia Nut 4pk-1L',
    'M-FR048U': 'Macadamia Nut 1L',
    'M-FR049F': 'Granny Smith Apple 4pk-1L',
    'M-FR049U': 'Granny Smith Apple 1L',
    'M-FR050F': 'Peppermint  4pk-1L',
    'M-FR050U': 'Peppermint 1 L',
    'M-FR051F': 'Toffee Nut  4 Pk-1 L',
    'M-FR051U': 'Toffee Nut 1 L',
    'M-FR053F': 'Mojito Mix 4pk-1L',
    'M-FR053U': 'Mojito Mix 1 L',
    'M-FR056F': 'Rose 4pk-1L',
    'M-FR056U': 'Rose 1L',
    'M-FR059F': 'Watermelon 4pk-1L',
    'M-FR059U': 'Watermelon 1 L',
    'M-FR060F': 'Gingerbread 4pk-1L',
    'M-FR060U': 'Gingerbread 1 L',
    'M-FR061F': 'Lavender 4pk-1L',
    'M-FR061U': 'Lavender 1L',
    'M-FR062F': 'Dark Chocolate 4pk-1L',
    'M-FR062U': 'Dark Chocolate 1 L',
    'M-FR063F': 'White Chocolate 4pk-1L',
    'M-FR063U': 'White Chocolate 1 L',
    'M-FR064F': 'Praline 4pk-1L',
    'M-FR064U': 'Praline 1L',
    'M-FR065F': 'Desert Pear 4pk-1L',
    'M-FR065U': 'Desert Pear 1 L',
    'M-FR066F': 'Guava 4pk-1L',
    'M-FR066U': 'Guava 1 L',
    'M-FR067F': 'Key Lime Pie 4pk-1L',
    'M-FR067U': 'Key Lime Pie 1 L',
    'M-FR068F': 'Red Passion Fruit 4pk-1L',
    'M-FR068U': 'Red Passion Fruit 1 L',
    'M-FR069F': 'Blood Orange 4pk-1L',
    'M-FR069U': 'Blood Orange 1 L',
    'M-FR070F': 'Lemongrass 4pk-1L',
    'M-FR070U': 'Lemongrass 1 L',
    'M-FR072F': 'Wild Strawberry 4pk-1L',
    'M-FR072U': 'Wild Strawberry 1 L',
    'M-FR075F': 'Pomegranate 4pk-1L',
    'M-FR075U': 'Pomegranate 1 L',
    'M-FR081F': 'Red Sangria Mix 4pk-1L',
    'M-FR081U': 'Red Sangria Mix 1 L',
    'M-FR084F': 'Honey Syrup 4pk-1L',
    'M-FR084U': 'Honey Syrup 1L',
    'M-FR087F': 'Candied Orange 4pk-1L',
    'M-FR087U': 'Candied Orange 1L',
    'M-FR092F': 'Blue Cotton Candy 4pk-1L',
    'M-FR092U': 'Blue Cotton Candy 1L',
    'M-FR095F': 'Cucumber 4pk-1L',
    'M-FR095U': 'Cucumber 1 L',
    'M-FR101F': 'White Sangria Mix 4pk-1L',
    'M-FR101U': 'White Sangria Mix 1 L',
    'M-FR105F': 'Pumpkin Spice 4pk-1L',
    'M-FR105U': 'Pumpkin Spice 1 L',
    'M-FR112F': 'Butterscotch 4pk-1L',
    'M-FR112U': 'Butterscotch 1L',
    'M-FR113F': 'Maple Spice 4pk - 1L',
    'M-FR113U': 'Maple Spice 1L',
    'M-FR114F': 'Wildberry 4pk-1L',
    'M-FR114U': 'Wildberry 1 L',
    'M-FR122F': 'Spicy Mango  4pk-1L',
    'M-FR122U': 'Spicy Mango 1 L',
    'M-FR123F': 'Chipotle Pineapple 4pk-1L',
    'M-FR123U': 'Chipotle Pineapple 1 L',
    'M-FR125F': 'Blackberry Sangria Mix 4pk-1L',
    'M-FR125U': 'Blackberry Sangria 1 L',
    'M-FR132F': 'Rock Melon Cantaloupe 4pk-1L',
    'M-FR132U': 'Rock Melon Cantaloupe  1L',
    'M-FR133F': 'Huckleberry 4pk-1L',
    'M-FR133U': 'Huckleberry 1 L',
    'M-FR135F': 'Cold Brew Coffee Concentrate 4pk-1L',
    'M-FR135U': 'Cold Brew Coffee Concentrate 1L',
    'M-FR136F': 'Wild Raspberry 4pk-1L',
    'M-FR136U': 'Wild Raspberry 1 L',
    'M-FR138F': 'Peach Mango 4pk-1L',
    'M-FR144F': 'Wild Grape 4pk-1L',
    'M-FR144U': 'Wild Grape 1L',
    'M-FR145F': 'Toasted Marshmallow 4pk-1L',
    'M-FR145U': 'Toasted Marshmallow 1L',
    'M-FR146F': 'Habanero Lime 4pk-1L',
    'M-FR146U': 'Habanero Lime 1 L',
    'M-FR147F': 'Elderflower 4pk-1L',
    'M-FR147U': 'Elderflower 1 L',
    'M-FR148F': 'White Peach 4pk-1L',
    'M-FR148U': 'White Peach 1L',
    'M-FR150F': 'French Raspberry 4pk-1L',
    'M-FR150U': 'French Raspberry 1L',
    'M-FR162F': 'Pumpkin Caramel 4pk-1L',
    'M-FR162U': 'Pumpkin Caramel 1L',
    'M-FR177F': 'Hibiscus 4pk-1L',
    'M-FR177U': 'Hibiscus 1 L',
    'M-FR190F': 'French Vanilla 4pk-1L',
    'M-FR190FP2': 'French Vanilla 4pk-1L W/ Pump',
    'M-FR190U': 'French Vanilla 1 L',
    'M-FR192F': 'Creme Caramel 4pk-1L',
    'M-FR192U': 'Creme Caramel 1 L',
    'M-FR193F': 'Roasted Hazelnut 4pk-1L',
    'M-FR193U': 'Roasted Hazelnut 1 L',
    'M-FR195F': 'Blue Raspberry 4pk-1L',
    'M-FR195U': 'Blue Raspberry 1 L',
    'M-FR197F': 'Spiced Brown Sugar 4pk-1L',
    'M-FR197U': 'Spiced Brown Sugar 1 L',
    'M-FR204F': 'Classic Pumpkin Pie 4pk-1L',
    'M-FR204U': 'Classic Pumpkin Pie 1ltr',
    'M-FR206F': 'Raspberry Lime 4pk-1L',
    'M-FR206U': 'Raspberry Lime 1L',
    'M-FR210F': 'Salted Caramel 4pk-1L',
    'M-FR210U': 'Salted Caramel 1L',
    'M-FR211F': 'Cupcake 4pk - 1L',
    'M-FR211U': 'Cupcake 1L',
    'M-FR212F': 'Cinnamon Bun 4pk-1L',
    'M-FR212U': 'Cinnamon Bun 1L',
    'M-FR213F': 'Hickory Smoke 4pk-1L',
    'M-FR213U': 'Hickory Smoke 1L',
    'M-FR217F': 'Maple Pancake 4pk-1L',
    'M-FR217U': 'Maple Pancake 1L',
    'M-FR218F': 'Tiki Blend 4pk-1L',
    'M-FR219F': 'South Seas Blend 4pk-1L',
    'M-FR219U': 'South Seas Blend 1L',
    'M-FR220F': 'Toasted Marshmallow Graham',
    'M-FR220U': 'Toasted Marshmallow Graham 1L',
    'M-FR223F': 'Black Raspberry 4pk-1L',
    'M-FR223U': 'Black Raspberry 1L',
    'M-FR224F': 'Vanilla Creme 4pk-1L',
    'M-FR224U': 'Vanilla Creme 1L',
    'M-FR225F': 'Orange Tangerine 4pk-1L',
    'M-FR225U': 'Orange Tangerine 1L',
    'M-FR226F': 'Old Fashioned Root Beer 4pk-1L',
    'M-FR226U': 'Old Fashioned Root Beer 1L',
    'M-FR230F': 'Ginger Beer 4pk-1L',
    'M-FR230U': 'Ginger Beer 1L',
    'M-FR232F': 'Exotic Citrus 4pk-1L',
    'M-FR232U': 'Exotic Citrus 1L',
    'M-FR237F': 'Honey Mango 4pk-1L',
    'M-FR237U': 'Honey Mango 1L',
    'M-FR238F': 'Stone Fruit 4pk-1L',
    'M-FR238U': 'Stone Fruit 1L',
    'M-FR241F': 'Pumpkin Pie 4pk-1L',
    'M-FR241U': 'Pumpkin Pie 1L',
    'M-FR245F': 'French Hazelnut 4pk-1L',
    'M-FR245U': 'French Hazelnut 1L',
    'M-FR246F': 'Hawaiian Island 4pk-1L',
    'M-FR246U': 'Hawaiian Island 1L',
    'M-FR247F': 'Cookie Butter 4pk-1L',
    'M-FR247U': 'Cookie Butter 1L',
    'M-FR252F': 'Wild Blackberry 4pk-1L',
    'M-FR252U': 'Wild Blackberry 1L',
    'M-FR255A': 'Classic Watermelon 12pk-1L',
    'M-FR255F': 'Classic Watermelon 4pk-1L',
    'M-FR255U': 'Classic Watermelon 1L',
    'M-FR256F': 'Tiramisu 4pk-1L',
    'M-FR256U': 'Tiramisu 1L',
    'M-FR258F': 'Butter Pecan 4pk-1L',
    'M-FR258U': 'Butter Pecan 1L',
    'M-FR259F': 'Tart Cherry 4pk-1L',
    'M-FR259U': 'Tart Cherry 1L',
    'M-FR264F': 'Chicory 4pk-1L',
    'M-FR267F': 'Winter Citrus 4pk-1L',
    'M-FR267U': 'Winter Citrus 1L',
    'M-FR268F': 'Caramel Apple Butter 4pk-1L',
    'M-FR268U': 'Caramel Apple Butter 1L',
    'M-FR274F': 'Brown Butter 4pk-1L',
    'M-FR274U': 'Brown Butter 1L',
    'M-FR275F': 'Brown Butter Toffee 4p-1L',
    'M-FR275U': 'Brown Butter Toffee 1L',
    'M-FR278F': 'Vanilla Spice 4pk-1L',
    'M-FR278U': 'Vanilla Spice 1L',
    'M-FR282F': 'S\'mores 4pk-1L',
    'M-FR282U': 'S\'mores 1L',
    'M-FR288F': 'Candy Corn 4pk-1L',
    'M-FR288U': 'Candy Corn 1L',
    'M-FR290F': 'Golden Turmeric 4pk-1L',
    'M-FR290U': 'Golden Turmeric 1L',
    'M-FR293A': 'Tropical Peach 12pk-1L',
    'M-FR293U': 'Tropical Peach 1L',
    'M-FR297F': 'Pineberry 4pk-1L',
    'M-FR297U': 'Pineberry 1L',
    'M-FR298F': 'Dragon Fruit 4pk-1L',
    'M-FR298U': 'Dragon Fruit 1L',
    'M-FR318F': 'Strawberry Rose 4pk-1L',
    'M-FR318U': 'Strawberry Rose 1L',
    'M-FR319F': 'Honey Jasmine 4pk-1L',
    'M-FR319U': 'Honey Jasmine 1L',
    'M-FR321F': 'Lavender Lemon 4pk-1L',
    'M-FR321U': 'Lavender Lemon 1L',
    'M-FR329F': 'Black Sugar 4pk-1L',
    'M-FR329U': 'Black Sugar 1L',
    'M-FR330F': 'Slush Base 4pk-1L',
    'M-FR337F': 'Pomegranate Peach 4pk-1L',
    'M-FR338F': 'Holiday Spice 4pk-1L',
    'M-FR339F': 'Holiday Peppermint 4pk-1L',
    'M-FR340F': 'Hot Honey Syrup 4pk-1L',
    'M-FR340U': 'Hot Honey Syrup 1L',
    'M-FR341F': 'Spicy Agave 4pk-1L',
    'M-FR341U': 'Spicy Agave 1L',
    'M-FR345F': 'Orange Dream 4pk-1L',
    'M-FR348F': 'Black Label Vanilla 4pk-1L',
    'M-FR348U': 'Black Label Vanilla 1L',
    'M-FR349F': 'Fall Pumpkin Spice 4pk-1L',
    'M-FR352FP': 'TB Mango Peach Syrup 4pk-1L W/ Pump',
    'M-FR352U': 'TB Mango Peach Syrup 1L',
    'M-FR357F': 'Turmeric Ginger 4pk-1L',
    'M-FR358F': 'Sugar Plum 4pk-1L',
    'M-FR360F': 'Sea Salt Caramel 4pk-1L',
    'M-FR361F': 'Ube 4pk-1L',
    'M-FR361U': 'Ube 1L',
    'M-FR362F': 'Strawberry Pineapple 4pk-1L',
    'M-FR362U': 'Strawberry Pineapple 1L',
    'M-FR363F': 'Pomegranate Acai 4pk-1L',
    'M-FR364F': 'Juicy Blue Raspberry 4pk-1L',
    'M-FR364U': 'Juicy Blue Raspberry 1L',
    'M-FR367F': 'Spicy Chocolate 4pk-1L (Peet\'s)',
    'M-FR368F': 'Wild Blueberry 4pk-1L',
    'M-FR369F': 'Harvest Pumpkin Pie 4pk-1L',
    'M-FR370F': 'Holiday Cranberry 4pk-1L',
    'M-FR373FP': 'TB Strawberry Passion Fruit Syrup 4pk-1L W/ Pump',
    'M-FR373U': 'TB Strawberry Passion Fruit Syrup 1L',
    'M-FR374FP': 'TB Dragon Fruit Berry Syrup 4pk-1L W/ Pump',
    'M-FR374U': 'TB Dragon Fruit Berry Syrup 1L',
    'M-FR378F': 'Yuzu Pineapple 4pk-1L',
    'M-FR378U': 'Yuzu Pineapple 1L',
    'M-FR379F': 'Strawberry Blossom 4pk-1L',
    'M-FR390F': 'Pink Cotton Candy 4pk-1L',
    'M-FR391F': 'Rainbow Sherbet 4pk-1L',
    'M-FR392F': 'Launch Pop 4pk-1L',
    'M-FR394F': 'Chocolate Cake 4pk1-L',
    'M-FR394U': 'Chocolate Cake 1L',
    'M-FR395F': 'Dark Cherry 4pk-1L',
    'M-FR395U': 'Dark Cherry 1L',
    'M-FR396F': 'Citrus Cardamom 4pk-1L',
    'M-FR399F': 'Passionfruit Pineapple Blue Raspberry Blend 4pk-1L',
    'M-FR400F': 'Toasted Coconut 4pk-1L',
    'M-FR400U': 'Toasted Coconut 1L',
    'M-FR405FP': 'TB Cookie Dough 4pk-1L W/ Pump',
    'M-FR410FP': 'Tropical Blue 4pk-1L W/ Pump',
    'M-FS001F': 'Sugar Free Almond 4pk-1L',
    'M-FS001U': 'Sugar Free Almond 1L',
    'M-FS006F': 'Sugar Free Blackberry 4pk-1L',
    'M-FS006U': 'Sugar Free Blackberry 1L',
    'M-FS009F': 'Sugar Free Caramel 4pk-1L',
    'M-FS009U': 'Sugar Free Caramel 1 L',
    'M-FS013F': 'Sugar Free Coconut 4pk-1L',
    'M-FS013U': 'Sugar Free Coconut 1L',
    'M-FS023F': 'Sugar Free Hazelnut 4pk-1L',
    'M-FS023U': 'Sugar Free Hazelnut 1 L',
    'M-FS024F': 'Sugar Free Chocolate  4pk-1L',
    'M-FS024U': 'Sugar Free Chocolate  1 L',
    'M-FS032F': 'Sugar Free Mango 4pk-1L',
    'M-FS032U': 'Sugar Free Mango 1L',
    'M-FS036F': 'Sugar Free Peach 4pk-1L',
    'M-FS036U': 'Sugar Free Peach 1 L',
    'M-FS040F': 'Sugar Free Raspberry 4pk-1L',
    'M-FS040U': 'Sugar Free Raspberry 1L',
    'M-FS042F': 'Sugar Free Strawberry 4pk-1L',
    'M-FS042U': 'Sugar Free Strawberry 1 L',
    'M-FS044F': 'Sugar Free Triple Sec 4pk-1L',
    'M-FS044U': 'Sugar Free Triple Sec 1 L',
    'M-FS045F': 'Sugar Free Vanilla 4pk-1L',
    'M-FS045U': 'Sugar Free Vanilla 1 L',
    'M-FS061F': 'Sugar Free Lavender 4pk-1L',
    'M-FS061U': 'Sugar Free Lavender 1L',
    'M-FS065F': 'Sugar Free Desert Pear 4pk-1L',
    'M-FS075F': 'Sugar Free Pomegranate 4pk-1L',
    'M-FS075U': 'Sugar Free Pomegranate 1 L',
    'M-FS083F': 'Sugar Free Sweetener 4pk-1L',
    'M-FS083FP': 'Sugar Free Sweetener 4pk-1L W/ Pump',
    'M-FS083FP2': 'Sugar Free Sweetener 4pk-1L W/ Pump',
    'M-FS083U': 'Sugar Free Sweetener 1 L',
    'M-FS190F': 'Sugar Free French Vanilla 4pk-1L',
    'M-FS190U': 'Sugar Free French Vanilla 1L',
    'M-FS195F': 'Sugar Free Blue Raspberry 4pk-1L',
    'M-FS195U': 'Sugar Free Blue Raspberry 1L',
    'M-FS337F': 'Sugar Free Pomegranate Peach 4pk-1L',
    'M-FT080F': 'Chai Tea 4pk-1L',
    'M-FT080U': 'Chai Tea Concentrate 1 L',
    'M-FT216F': 'Iced Coffee Concentrate 4pk-1L',
    'M-FT216U': 'Iced Coffee Concentrate 1L',
    'M-FT322F': 'Energy Boost Concentrate 4pk-1L',
    'M-FT322U': 'Energy Boost Concentrate 1L',
    'M-FT342F': 'Maple Pumpkin Cold Brew 4pk-1L',
    'M-FT342U': 'Maple Pumpkin Cold Brew 1L',
    'M-FT366F': 'Matcha Green Tea 4pk-1L',
    'M-FT366U': 'Matcha Green Tea 1L',
    'M-FT415F': 'Green Tea Concentrate 4pk-1L',
    'M-FX325F': 'Total Immunity Boost 4pk-1L',
    'M-FX325FP': 'Total Immunity Boost 4pk-1L W/ Pump',
    'M-FX325U': 'Total Immunity Boost 1L',
    'M-FX326F': 'Energy Boost 4pk-1L',
    'M-FX326FP': 'Energy Boost 4pk-1L W/ Pump',
    'M-FX326U': 'Energy Boost 1L',
    'M-FX346F': 'Hydration Boost 4pk-1L (Wawa)',
    'M-FX346U': 'Hydration Boost 1L (Wawa)',
    'M-FX356FP': 'Hydration Boost 4pk-1L W/ Pump',
    'M-FX356U': 'Hydration Boost 1L',
    'M-FX372F': 'Very Berry Immunity 4pk-1L',
    'M-GBP032U': 'Mango Puree 500 Lb Drum',
    'M-GC009FP': 'Caramel Sauce 4pk-64oz',
    'M-GC009U': 'Caramel Sauce 64 fl oz',
    'M-GC062FP': 'Dark Chocolate Sauce 4pk-64oz',
    'M-GC062U': 'Dark Chocolate Sauce 64 fl oz',
    'M-GC063FP': 'White Chocolate Sauce 4pk-64oz',
    'M-GC063U': 'White Chocolate Sauce 64 fl oz',
    'M-GC221FP': 'Dulce de Leche Sauce 4pk-64oz',
    'M-GC221U': 'Dulce de Leche 64oz',
    'M-GC300FP': 'Sea Salt Caramel Toffee Sauce 4pk-64oz',
    'M-GC300U': 'Sea Salt Caramel Toffee Sauce 64oz',
    'M-GF062FP': 'Sugar Free Dark Chocolate Sauce 4pk-64oz',
    'M-GF062U': 'Sugar Free Dark Chocolate Sauce 64 floz',
    'M-GO157F': 'Organic Agave Nectar 4pk 64oz',
    'M-GO157U': 'Organic Agave 64oz',
    'M-GR069FP': 'Blood Orange 4pk-64oz w/ Pump',
    'M-GR192FP': 'Creme Caramel 4pk-64oz W/ Pump',
    'M-GR253F': 'Margarita Mix 4pk-64oz',
    'M-GR253U': 'Margarita Mix 64oz',
    'M-GR259FP': 'Tart Cherry 4pk-64oz W/ Pump',
    'M-GR272FP': 'Mango Strawberry 4pk-64oz W/ Pump',
    'M-GR272U': 'Mango Strawberry 64oz',
    'M-GR320FP': 'Peet\'s Blended Base 4pk-64oz w/ Pump',
    'M-GR320U': 'Peet\'s Blended Base 64oz',
    'M-GR333FP': 'Neutral Base 4pk-64oz W/ Pump',
    'M-GR333U': 'Neutral Base 64oz',
    'M-GR353F': 'Red\'s Margarita Mix 4pk-64oz',
    'M-GR353U': 'Red\'s Margarita Mix 64oz',
    'M-GR355FP': 'Mango Passion Fruit 4pk-64oz W/ Pump',
    'M-GR377FP': 'Blueberry Cotton Candy 4pk-64oz W/ Pump',
    'M-GR393FP': 'Churro 4pk-64oz W/ Pump',
    'M-GT080FP': 'Chai Tea 4pk-64oz W/ Pump',
    'M-GVS384FP': 'BRL SF Berry Blue Flvrd Energy 4pk-64oz W/Pump',
    'M-GVS384U': 'BRL SF Berry Blue Flvrd Energy 64oz',
    'M-GVS385FP': 'BRL SF Dragon Frt Pnk Flvr Enrgy 4pk-64oz W/Pump',
    'M-GVS385U': 'BRL SF Dragon Frt Pnk Flvr Enrgy 64oz',
    'M-GVS386FP': 'BRL SF Glacier Clear Energy 4pk-64oz W/ Pump',
    'M-GVS386U': 'BRL SF Glacier Clear Energy 64oz',
    'M-GVX381FP': 'BRL Yumberry Red Natural Energy 4pk-64oz W/ Pump',
    'M-GVX381U': 'BRL Yumberry Red Natural Energy 64oz',
    'M-GVX382FP': 'BRL Starfruit Yellow Nat Energy 4pk-64oz W/ Pump',
    'M-GVX382U': 'BRL Starfruit Yellow Natural Energy 64oz',
    'M-GVX383FP': 'BRL Powerfruit Purple Nat Energy 4pk-64oz W/ Pump',
    'M-GVX383U': 'BRL Powerfruit Purple Natural Energy 64oz',
    'M-GVX384FP': 'BRL Berry Blue Natural Energy 4pk-64oz W/ Pump',
    'M-GVX384U': 'BRL Berry Blue Natural Energy 64oz',
    'M-GVX385FP': 'BRL Dragon Fruit Pink Nat Energy 4pk-64oz W/ Pump',
    'M-GVX385U': 'BRL Dragon Fruit Pink Natural Energy 64oz',
    'M-GVX386FP': 'BRL Glacier Clear Natural Energy 4pk-64oz W/ Pump',
    'M-GVX386U': 'BRL Glacier Clear Natural Energy 64oz',
    'M-GVX407FP': 'BRL Orange Cream Nat Energy 4pk-64oz W/ Pump',
    'M-GVX407U': 'BRL Orange Cream Natural Energy 64oz',
    'M-GX312F': 'HomeCrafted Strawberry Ginger Lmnd 4pk-64oz',
    'M-KC009B': 'Caramel Sauce 6pk-12oz',
    'M-KC009U': 'Caramel Sauce 12 fl oz',
    'M-KC062B': 'Dark Chocolate Sauce 6pk-12oz',
    'M-KC062U': 'Dark Chocolate Sauce 12 fl oz',
    'M-KC063B': 'White Chocolate Sauce 6pk-12oz',
    'M-KC063U': 'White Chocolate Sauce 12 floz',
    'M-KC221B': 'Dulce de Leche Sauce 6pk-12oz',
    'M-KC221U': 'Dulce de Leche Sauce 12oz',
    'M-KC300B': 'Sea Salt Caramel Toffee Sauce 6pk-12oz',
    'M-KC300U': 'Sea Salt Caramel Toffee Sauce 12oz',
    'M-KF062B': 'SF Dark Chocolate Sauce 6pk-12floz',
    'M-KF062U': 'SF Dark Chocolate Sauce 12 floz',
    'M-LR003U': 'Apple 30 Gal. Drum',
    'M-LR006U': 'Blackberry 30 Gal. Drum',
    'M-LR013U': 'Coconut 30 Gal. Drum',
    'M-LR018U': 'Ginger 30 Gal. Drum',
    'M-LR032U': 'Mango 30 Gal. Drum',
    'M-LR036U': 'Peach 30 Gal. Drum',
    'M-LR038U': 'Pineapple 30 Gal. Drum',
    'M-LR040U': 'Raspberry 30 Gal. Drum',
    'M-LR042U': 'Strawberry 30 Gal. Drum',
    'M-LR043U': 'Swiss Chocolate 30 Gal. Drum',
    'M-LR045U': 'Vanilla 30 Gal. Drum',
    'M-LR049U': 'Granny Smith Apple 30 Gal. Drum',
    'M-LR059U': 'Watermelon 30 Gal. Drum',
    'M-LR061U': 'Lavender 30 Gal. Drum',
    'M-LR095U': 'Cucumber 30 Gal. Drum',
    'M-LR133U': 'Huckleberry 30 Gal. Drum',
    'M-LR145U': 'Toasted Marshmallow 30 Gal. Drum',
    'M-LR210U': 'Salted Caramel 30 Gal. Drum',
    'M-LR224U': 'Vanilla Creme 30 Gal. Drum',
    'M-LR228U': 'Chocolate Fudge 30 Gal.  Drum',
    'M-LR241U': 'Pumpkin Pie 30 Gal. Drum',
    'M-LR255U': 'Classic Watermelon 30 Gal. Drum',
    'M-NL157D': 'Agave Nectar 50ml Glass 120pk',
    'M-NL157L': 'Agave Nectar 115pk-50ml Glass',
    'M-NL157U': 'Agave Nectar 50 mL glass',
    'M-NR000D': 'Pure Cane 120pk-50ml Glass',
    'M-NR000L': 'Pure Cane 115pk-50ml Glass',
    'M-NR000U': 'Pure Cane 50ml Glass',
    'M-NR006D': 'Blackberry 120pk-50ml Glass',
    'M-NR006L': 'Blackberry 115pk-50ml Glass',
    'M-NR006U': 'Blackberry 50 mL Glass',
    'M-NR009D': 'Caramel 120pk-50ml Glass',
    'M-NR009L': 'Caramel 115pk-50ml Glass',
    'M-NR009U': 'Caramel 50ml Glass',
    'M-NR013D': 'Coconut 120pk-50ml Glass',
    'M-NR013L': 'Coconut 115pk-50ml Glass',
    'M-NR013U': 'Coconut 50ml Glass',
    'M-NR015L': 'Cranberry 115pk-50ml Glass',
    'M-NR022L': 'Grenadine 115pk-50ml Glass',
    'M-NR022U': 'Grenadine 50ml Glass',
    'M-NR023D': 'Hazelnut 120pk-50ml Glass',
    'M-NR023L': 'Hazelnut 115pk-50ml Glass',
    'M-NR023U': 'Hazelnut 50 mL Glass',
    'M-NR025D': 'Irish Cream 120pk-50ml Glass',
    'M-NR025L': 'Irish Cream 115pk-50ml Glass',
    'M-NR025U': 'Irish Cream 50ml Glass',
    'M-NR032D': 'Mango 120pk-50ml Glass',
    'M-NR032L': 'Mango 115pk-50ml Glass',
    'M-NR032U': 'Mango 50 mL Glass',
    'M-NR035D': 'Passion Fruit 120pk-50ml Glass',
    'M-NR035L': 'Passion Fruit 115pk-50ml Glass',
    'M-NR035U': 'Passion Fruit 50ml Glass',
    'M-NR040D': 'Raspberry 120pk-50ml Glass',
    'M-NR040L': 'Raspberry 115pk-50ml Glass',
    'M-NR040U': 'Raspberry 50mL Glass',
    'M-NR042D': 'Strawberry 120pk-50ml Glass',
    'M-NR042L': 'Strawberry 115pk-50ml Glass',
    'M-NR042U': 'Strawberry 50 mL Glass',
    'M-NR045D': 'Vanilla 120pk-50ml Glass',
    'M-NR045L': 'Vanilla 115pk-50ml Glass',
    'M-NR045U': 'Vanilla 50 mL Glass',
    'M-NR047D': 'Amaretto 120pk-50ml Glass',
    'M-NR047L': 'Amaretto 115pk-50ml Glass',
    'M-NR047U': 'Amaretto 50ml Glass',
    'M-NR048D': 'Macadamia Nut 50ml 120pk-50ml Glass',
    'M-NR048L': 'Macadamia Nut 115pk-50ml Glass',
    'M-NR048U': 'Macadamia Nut 50ml Glass',
    'M-NR050D': 'Peppermint 120pk-50ml Glass',
    'M-NR050L': 'Peppermint 115pk-50ml Glass',
    'M-NR050U': 'Peppermint 50 mL Glass',
    'M-NR053D': 'Mojito Mix 120pk-50ml Glass',
    'M-NR053L': 'Mojito Mix 115pk-50ml Glass',
    'M-NR053U': 'Mojito Mix 50 mL Glass',
    'M-NR055D': 'Violet 120pk-50ml Glass',
    'M-NR055L': 'Violet 115pk-50ml Glass',
    'M-NR055U': 'Violet 50ml Glass',
    'M-NR056D': 'Rose 120pk-50ml Glass',
    'M-NR056L': 'Rose 115pk-50ml Glass',
    'M-NR056U': 'Rose 50ml Glass',
    'M-NR060D': 'Gingerbread 120pk-50ml Glass',
    'M-NR060L': 'Gingerbread 115pk-50ml Glass',
    'M-NR060U': 'Gingerbread 50ml Glass',
    'M-NR061D': 'Lavender 120pk-50ml Glass',
    'M-NR061L': 'Lavender 115pk-50ml Glass',
    'M-NR061U': 'Lavender 50ml Glass',
    'M-NR062D': 'Dark Chocolate 120pk-50ml Glass',
    'M-NR062U': 'Dark Chocolate 50ml Glass',
    'M-NR075D': 'Pomegranate 120pk-50ml Glass',
    'M-NR075L': 'Pomegranate 115pk-50ml Glass',
    'M-NR075U': 'Pomegranate 50 mL Glass',
    'M-NR105D': 'Pumpkin Spice 120pk-50ml Glass',
    'M-NR105L': 'Pumpkin Spice 115pk-50ml Glass',
    'M-NR105U': 'Pumpkin Spice 50 mL Glass',
    'M-NR145D': 'Toasted Marshmallow 120pk-50ml Glass',
    'M-NR145L': 'Toasted Marshmallow 115pk-50ml Glass',
    'M-NR145U': 'Toasted Marshmallow 50ml Glass',
    'M-NR147L': 'Elderflower 115pk-50ml Glass',
    'M-NR147U': 'Elderflower 50ml Glass',
    'M-NR177D': 'Hibiscus 120pk-50ml Glass',
    'M-NR177L': 'Hibiscus 115pk-50ml Glass',
    'M-NR177U': 'Hibiscus 50ml Glass',
    'M-NR212L': 'Cinnamon Bun 115pk-50ml Glass',
    'M-NR212U': 'Cinnamon Bun 50ml Glass',
    'M-NR241L': 'Pumpkin Pie 115pk-50ml Glass',
    'M-NR241U': 'Pumpkin Pie 50ml',
    'M-NR247L': 'Cookie Butter 115pk-50ml Glass',
    'M-NR247U': 'Cookie Butter 50ml',
    'M-NR268D': 'Caramel Apple Butter 120pk-50ml Glass',
    'M-NR268L': 'Caramel Apple Butter 115pk-50ml Glass',
    'M-NR268U': 'Caramel Apple Butter 50ml',
    'M-NR275D': 'Brown Butter Toffee 120pk-50ml Glass',
    'M-NR275L': 'Brown Butter Toffee 115pk-50ml Glass',
    'M-NR275U': 'Brown Butter Toffee 50ml Glass',
    'M-NR318D': 'Strawberry Rose 120pk-50ml Glass',
    'M-NR318U': 'Strawberry Rose 50ml Glass',
    'M-NR319D': 'Honey Jasmine 120pk-50ml Glass',
    'M-NR319U': 'Honey Jasmine 50ml Glass',
    'M-NR321D': 'Lavender Lemon 120pk-50ml Glass',
    'M-NR321U': 'Lavender Lemon 50ml',
    'M-NR341D': 'Spicy Agave 120pk-50ml Glass',
    'M-NR341L': 'Spicy Agave 115pk-50ml Glass',
    'M-NR341U': 'Spicy Agave 50ml Glass',
    'M-NR361D': 'Ube 120pk-50ml',
    'M-NR361L': 'Ube 115pk-50ml Glass',
    'M-NR361U': 'Ube 50ml Glass',
    'M-NR378L': 'Yuzu Pineapple 115pk-50ml Glass',
    'M-NR378U': 'Yuzu Pineapple 50ml Glass',
    'M-NR400L': 'Toasted Coconut 115pk-50ml Glass',
    'M-NR400U': 'Toasted Coconut 50ml Glass',
    'M-NS013L': 'Sugar Free Coconut 115pk-50ml Glass',
    'M-NS013U': 'Sugar Free Coconut 50ml Glass',
    'M-NS042U': 'SF Strawberry 50ml Glass',
    'M-NT216D': 'Iced Coffee Concentrate 120pk-50ml Glass',
    'M-NT216L': 'Iced Coffee Concentrate 115pk-50ml Glass',
    'M-NT216U': 'Iced Coffee Concentrate 50ml Glass',
    'M-NVS384L': 'BRL SF Berry Blue Flavored Energy 115pk-50ml',
    'M-NVS384U': 'BRL SF Berry Blue Flavored Energy 50ml',
    'M-NVS385L': 'BRL SF Dragon Fruit Pink Flvrd Energy 115pk-50ml',
    'M-NVS385U': 'BRL SF Dragon Fruit Pink Flavored Energy 50ml',
    'M-NVS386L': 'BRL SF Glacier Clear Energy 115pk-50ml',
    'M-NVS386U': 'BRL SF Glacier Clear Energy 50ml',
    'M-NVX381L': 'BRL Yumberry Red Natural Energy 115pk-50ml',
    'M-NVX381U': 'BRL Yumberry Red Natural Energy 50ml',
    'M-NVX382L': 'BRL Starfruit Yellow Natural Energy 115pk-50ml',
    'M-NVX382U': 'BRL Starfruit Yellow Natural Energy 50ml',
    'M-NVX383L': 'BRL Powerfruit Purple Natural Energy 115pk-50ml',
    'M-NVX383U': 'BRL Powerfruit Purple Natural Energy 50ml',
    'M-NVX384L': 'BRL Berry Blue Natural Energy 115pk-50ml',
    'M-NVX384U': 'BRL Berry Blue Natural Energy 50ml',
    'M-NVX385L': 'BRL Dragon Fruit Pink Natural Energy 115pk-50ml',
    'M-NVX385U': 'BRL Dragon Fruit Pink Natural Energy 50ml',
    'M-NVX386L': 'BRL Glacier Clear Natural Energy 115pk-50ml',
    'M-NVX386U': 'BRL Glacier Clear Natural Energy 50ml',
    'M-NVX407L': 'BRL Orange Cream Natural Energy 115pk-50ml',
    'M-NVX407U': 'BRL Orange Cream Natural Energy 50ml',
    'M-NX326U': 'Energy Boost 50ml Glass',
    'M-RP006F': 'Blackberry Puree 4pk-1L',
    'M-RP006U': 'Blackberry Puree 1L',
    'M-RP008F': 'Blueberry Puree 4pk-1L',
    'M-RP008U': 'Blueberry Puree 1L',
    'M-RP013F': 'Coconut Puree 4pk -1L',
    'M-RP013U': 'Coconut Puree 1L',
    'M-RP018F': 'Ginger Puree 4pk-1L',
    'M-RP018U': 'Ginger Puree 1L',
    'M-RP032F': 'Mango Puree 4pk-1L',
    'M-RP032U': 'Mango Puree 1 L',
    'M-RP035F': 'Passion Fruit Puree 4pk-1L',
    'M-RP035U': 'Passion Fruit Puree 1 L',
    'M-RP036F': 'Peach Puree 4pk-1L',
    'M-RP036FP2': 'Peach Puree 4pk-1L W/ Pump (Flat Cap)',
    'M-RP036U': 'Peach Puree 1 L',
    'M-RP038F': 'Pineapple Puree 4pk-1L',
    'M-RP038FP': 'Pineapple Puree 4pk-1L W/ Pump',
    'M-RP038U': 'Pineapple Puree 1L',
    'M-RP040F': 'Raspberry Puree 4pk-1L',
    'M-RP040U': 'Raspberry Puree 1 L',
    'M-RP042F': 'Strawberry Puree 4pk-1L',
    'M-RP042FP2': 'Strawberry Puree 4pk-1L W/ Pump (Flat Cap)',
    'M-RP042U': 'Strawberry Puree 1 L',
    'M-RP046F': 'Banana Puree 4pk-1L',
    'M-RP046U': 'Banana Puree 1 L',
    'M-RP049F': 'Granny Smith Apple Puree 4pk - 1L',
    'M-RP049U': 'Granny Smith Apple Puree 1L',
    'M-RP059F': 'Watermelon Puree 4pk-1L',
    'M-RP059U': 'Watermelon Puree 1L',
    'M-RP066F': 'Guava Puree 4pk-1L',
    'M-RP066U': 'Guava Puree 1L',
    'M-RP069F': 'Blood Orange Puree 4pk-1L',
    'M-RP069U': 'Blood Orange Puree 1L',
    'M-RP114F': 'Wildberry Puree 4pk-1L',
    'M-RP114FP': 'Wildberry Puree 4pk-1L Wth Pump',
    'M-RP114U': 'Wildberry Puree 1 L',
    'M-RP142F': 'Black Cherry Puree 4pk-1L',
    'M-RP142U': 'Black Cherry Puree 1L',
    'M-RP248F': 'Yuzu Puree 4pk-1L',
    'M-RP248U': 'Yuzu Puree 1L',
    'M-RP249F': 'Spiced Pumpkin Puree 4pk-1L',
    'M-RP249U': 'Spiced Pumpkin Puree 1L',
    'M-RP285F': 'Strawberry Rhubarb Puree 4pk-1L',
    'M-RP285U': 'Strawberry Rhubarb Puree 1L',
    'M-TR025U': 'Irish Cream 1,000 L Tote',
    'M-TR035U': 'Passion Fruit 1,000 L Tote',
    'M-TR045U': 'Vanilla 1,000 L Tote',
    'M-TR063U': 'White Chocolate 1,000 L Tote',
    'M-TR298U': 'Dragon Fruit 1,000 L Tote',
    'M-VJ006FP': 'Blackberry Concentrated Flavor 4pk-375mL',
    'M-VJ006U': 'Blackberry Concentrated Flavor 375mL',
    'M-VJ008FP': 'Blueberry Concentrated Flavor 4pk-375mL',
    'M-VJ008U': 'Blueberry Concentrated Flavor 375mL',
    'M-VJ013FP': 'Coconut Concentrated Flavor 4pk-375mL',
    'M-VJ013U': 'Coconut Concentrated Flavor 375mL',
    'M-VJ018FP': 'Ginger Concentrated Flavor 4pk-375mL',
    'M-VJ018U': 'Ginger Concentrated Flavor 375mL',
    'M-VJ032FP': 'Mango Concentrated Flavor 4pk-375mL',
    'M-VJ032U': 'Mango Concentrated Flavor 375mL',
    'M-VJ035FP': 'Passion Fruit Concentrated Flavor 4pk-375mL',
    'M-VJ035U': 'Passion Fruit Concentrated Flavor 375mL',
    'M-VJ036FP': 'Peach Concentrated Flavor 4pk-375mL',
    'M-VJ036U': 'Peach Concentrated Flavor 375mL',
    'M-VJ038FP': 'Pineapple Concentrated Flavor 4pk-375mL',
    'M-VJ038U': 'Pineapple Concentrated Flavor 375mL',
    'M-VJ040FP': 'Raspberry Concentrated Flavor 4pk-375mL',
    'M-VJ040U': 'Raspberry Concentrated Flavor 375mL',
    'M-VJ042FP': 'Strawberry Concentrated Flavor 4pk-375mL',
    'M-VJ042U': 'Strawberry Concentrated Flavor 375mL',
    'M-VJ045FP': 'Vanilla Concentrated Flavor 4pk-375mL',
    'M-VJ045U': 'Vanilla Concentrated Flavor 375mL',
    'M-VJ059FP': 'Watermelon Concentrated Flavor 4pk-375mL',
    'M-VJ059U': 'Watermelon Concentrated Flavor 375mL',
    'M-VJ095FP': 'Cucumber Concentrated Flavor 4pk-375mL',
    'M-VJ095U': 'Cucumber Concentrated Flavor 375mL',
    'M-VJ188FP': 'Mint Concentrated Flavor 4pk-375mL',
    'M-VJ188U': 'Mint Concentrated Flavor 375mL',
    'M-VJ235FP': 'Basil Concentrated Flavor 4pk-375mL',
    'M-VJ235U': 'Basil Concentrated Flavor 375mL',
    'M-VJ236FP': 'Habanero Concentrated Flavor 4pk-375mL',
    'M-VJ236U': 'Habanero Concentrated Flavor 375mL',
    'M-VJ260FP': 'Jalapeno Concentrated Flavor 4pk-375mL',
    'M-VJ260U': 'Jalapeno Concentrated Flavor 375mL',
    'M-VJ262FP': 'Oak Barrel Concentrated Flavor 4pk-375mL',
    'M-VJ262U': 'Oak Barrel Concentrated Flavor 375mL',
    'M-VJ292FP': 'Lime Concentrated Flavor 4pk-375ml',
    'M-VJ292U': 'Lime Concentrated Flavor 375ml',
    'M-VJ371FP': 'Bourbon Conc Flavor 4pk-375ml',
    'MB-AR009A': 'Caramel 12pk-750mL (Canada)',
    'MB-AR023A': 'Hazelnut 12pk-750mL (Canada)',
    'MB-AR045A': 'Vanilla 12pk-750mL (Canada)',
    'MB-AS045A': 'Sugar Free Vanilla 12pk-750mL (Canada)',
    'MB-AS045U': 'Sugar Free Vanilla 750 mL (Canada)',
    'MB-FL157U': 'Organic Agave Nectar 1L (Canada)',
    'MB-FR009F': 'Caramel 4pk-1L (Canada)',
    'MB-FR009U': 'Caramel 1L (Canada)',
    'MB-FR023F': 'Hazelnut 4pk-1L (Canada)',
    'MB-FR023U': 'Hazelnut 1L (Canada)',
    'MB-FR045F': 'Vanilla 4pk-1L (Canada)',
    'MB-FR045U': 'Vanilla 1L (Canada)',
    'MB-FR193FP': 'Roasted Hazelnut 4pk-1L W/ Pump (Canada)',
    'MB-FR256FP': 'Tiramisu Flavored Syrup 4pk-1L W/ Pump (Canada)',
    'MB-FR365FP': 'White Chocolate Pistachio 4pk-1L W/ Pump (Canada)',
    'MB-FR406FP': 'Chocolate Pistachio 4pk-1L W/ Pump (Canada)',
    'MB-FS045F': 'Sugar Free Vanilla 4pk-1L (Canada)',
    'MB-FS045U': 'Sugar Free Vanilla 1L (Canada)',
    'MB-GC009FP': 'Caramel Sauce 4pk-64oz (Canada)',
    'MB-GC009U': 'Caramel Sauce 64 fl oz (Canada)',
    'MB-GC062FP': 'Dark Chocolate Sauce 4pk-64oz (Canada)',
    'MB-GC062U': 'Dark Chocolate Sauce 64 fl oz (Canada)',
    'MB-GC063FP': 'White Chocolate Sauce 4pk-64oz (Canada)',
    'MB-GC063U': 'White Chocolate Sauce 64 fl oz (Canada)',
    'MB-GVX386FP': 'BRL Glacier Clear Nat Engy 4pk-64oz W Pmp (Canada)',
    'MB-RP032FP': 'Mango Puree 4pk-1L W/ Pump (Canada)',
    'MB-RP328FP': 'Strawberry Puree 4pk-1L W/ Pump (V3) (Canada)',
    'MB-RP380FP': 'Strawberry Puree 4pk-1L W/ Pump (V4) (Canada)',
    'MF-FR340F': 'Hot Honey Syrup 4pk-1L (UK)',
    'MF-RP059F': 'Watermelon Puree 4pk-1L Flat Cap (UK)',
    'MKFC-BS283U': 'Sugar Free Lemonade 3L BIB',
    'OG-FR076A': 'OG Bellini 12pk-1L',
    'OG-FR076U': 'Bellini Mix 1 Liter',
    'OG-FR100A': 'OG Berry Sangria 12pk-1L',
    'OG-FR100U': 'OG Berry Sangria 1L',
    'P001-L1': 'Brilliance Orange Cream Letter',
    'P001-L2': 'TruFlavour Culinary Letter',
    'P114': 'Miscellaneous Item',
    'P120': 'Plain 8oz Coffee Cup (25pk bag)',
    'P122': 'Black Unwrapped 7.75" Straw',
    'P123': 'Black Unwrapped Straw 5.75" (150pk box)',
    'P125-A': 'Monin Logo Panache Tumbler',
    'P128': 'Iced Coffee Cold Brew Concentrate Sheet',
    'P131': 'Microfiber Screen Cloth',
    'P133': 'Monin Logo 8oz Cup - Insulated (25pk Bag)',
    'P136': 'Premium Syrups Sale Sheet',
    'P137': 'Fruit Purees Sale Sheet',
    'P138': 'Beverage Concentrates Sale Sheet',
    'P139': 'Sweeteners Sale Sheet',
    'P140': 'Organic Syrups Sale Sheet',
    'P141': 'Gourmet Sauce Sales Sheet',
    'P142': 'Sugar Free Syrups Sale Sheet',
    'P147': 'Coconut + Sugar Free Coconut Sale Sheet',
    'P148': 'TruFlavour - Full Range Brochure 2026',
    'P148RB': 'TruFlavour - Recipe Fan Booklet',
    'P149A': 'TruFlavour Product Insert',
    'P151': '4 Bottle Tote Bag',
    'P153': '12pk Wheeled Bag',
    'P157': 'Monin 50oz Thermos',
    'P159': 'GFS Pick 4',
    'P160': 'Monin Post-It Notes',
    'P161': 'Monin Sharpie',
    'P162': 'Monin Mint Tin',
    'P163': 'Brown/White Monin Notebook',
    'P170-CC': 'Natural Smoothie Counter Card',
    'P170-P': 'Natural Smoothie Posters',
    'P181': 'Flavor Your Day Counter Card - Coffee',
    'P186': '16oz Coffee Ground Eco Tumbler',
    'P187': '4oz Monin Branded Hot Cup (50pk bag)',
    'P192A': 'EOS Lip Balm - Vanilla Bean',
    'P192B': 'EOS Lip Balm - Strawberry Sorbet',
    'P192C': 'Lip Butter - Coconut Water',
    'P192D': 'Lip Butter - Peppermint',
    'P192E': 'Lip Butter - Vanilla',
    'P193': 'GFS Pick 4 Puree Sheet',
    'P194': 'Monin Tervis Tumbler With Lid',
    'P198': 'Monin Selfie Tripod Kits',
    'P201': 'Monin Paper Straw 7.75" Unwrapped (100pk Bag)',
    'P206': 'Plain 4oz Hot Cup (50pk bag)',
    'P207': 'Monin Squeeze Bottle',
    'P208': 'Monin Journal',
    'P209': 'Monin Measuring Jigger (Acrylic) 3oz',
    'P210': 'Monin Measuring Jigger (Stainless) 1-2oz',
    'P212': 'Monin Logo 20oz Plastic Shaker',
    'P221-20PK': 'Pump Cover 20pk - 750ml & 1L Pumps',
    'P225': 'Custom 3ml Pump Black for 375ml Conc Flavor',
    'P225-250PK': 'Custom 3ml Pump Black 250pk for 375ml Conc Flavor',
    'P226': 'Custom 3ml Pump Black for 1L Conc Flavor',
    'P230': 'Glass Bottle Pump 1/4 oz White With Topper',
    'P230-200PK': 'Glass Bottle Pump 1/4 oz White With Topper 200Pk',
    'P230-4PK': 'Glass Bottle Pump 1/4oz 4pk',
    'P230-72PK': '72pk of P230 Pumps (6-12pk Bags per case)',
    'P230B': 'Glass Bottle Pump 1/4 oz Black With Topper',
    'P231': '11 Bottle Display Rack-750ml/1L',
    'P233': 'Monin Phone PopSocket',
    'P238': 'Long Straw 9.429"',
    'P240': 'Plastic Bottle Pump 1/4 oz White W/ Topper',
    'P240-20PK': 'TB Plastic Bottle Pump 1/4 oz White 20pk',
    'P240-250PK': 'Plastic Bottle Pump 1/4 oz White With Topper 250pk',
    'P240-4PK': 'Plastic Bottle Pump 1/4 oz White 4pk',
    'P240-CK': 'Plastic Bottle Pump 1/4 oz White 20pk',
    'P240-R': 'Plastic Bottle Pump 1/4 oz White - Removed Topper',
    'P240-R-250PK': 'Plastic Bottle Pump 1/4 oz White No Topper 250pk',
    'P240-TB5': 'TB Plastic Bottle Pump 1/4 oz White 5pk',
    'P240B': 'Plastic Bottle Pump 1/4 oz Black With Topper',
    'P240B-250PK': 'Plastic Bottle Pump 1/4 oz Black With Topper 250pk',
    'P240R-10ML': 'Plastic Bottle Pump 10mL Red',
    'P241L': '64oz Bottle Pump 1oz White Long Straw',
    'P242L': '64oz Bottle Pump 3/4 oz White Long Straw',
    'P243BL': '64oz Bottle Pump 1/2 oz Black Long Straw',
    'P243L': '64oz Bottle Pump 1/2 oz White Long Straw',
    'P243L-75PK': '64oz Bottle Pump 1/2 oz White Long Straw 75pk',
    'P244BL': '64oz Bottle Pump 1/4 oz Black Long Straw',
    'P244L': '64oz Bottle Pump 1/4 oz White Long Straw',
    'P245': 'Puree Pump 1/2 oz Black With Topper',
    'P246-R': 'Puree Pump 1/4 oz Black - Removed Topper',
    'P247': 'Trifold Brochure',
    'P251': 'Culinary-Breakfast Sell Sheet',
    'P255': 'La Flavour Sell Sheet',
    'P267': 'GFS Conc Flv Pick 4 Sell Sheet',
    'P272': 'Natural Zero Sell Sheet',
    'P274': 'Margarita Mix Sell Sheet',
    'P276': 'Monin Sunglasses',
    'P277': 'Concentrated Flavor Water Sell Sheet',
    'P281': 'Monin Webcam Cover',
    'P282': 'C-Store Rack Insert',
    'P285': 'Bev Dispenser Sale Sheet',
    'P296': 'Rack and Pump Sale Sheet',
    'P298': 'Happy Hour Menu',
    'P299C': 'Small Shelf Talker - HomeCrafted',
    'P317': 'C Store Coffee Sale Sheet',
    'P318': 'C Store Tea Sale Sheet',
    'P319': 'C Store Iced Coffee Sale Sheet',
    'P335': 'Monin Bar Spoon',
    'P336': 'Monin Mini Popcorn Box',
    'P344': 'Kraft Self-Seal StayFlat Mailer #10 - 7x9"',
    'P350': 'Industry Hub Postcard',
    'P351': 'How to Beer Sales Sheet',
    'P355': 'Pineberry Sale Sheet',
    'P356': 'Dragon Fruit Sale Sheet',
    'P365': 'Monin Chef Journal',
    'P368': 'Monin Spritz Glass (2pk.)',
    'P368-B': 'Monin Spritz Glass Box',
    'P371': 'Stress Ball - Pineberry Shape',
    'P372': 'Black Monin Tote Bag W/ Bottle Print',
    'P380': 'Pumpkin Spice Sale Sheet',
    'P387': 'Hard Seltzer Infusions Sale Sheet',
    'P391': 'Pineapple Key Chain',
    'P392': 'Monin Beverage Coozie',
    'P393': 'Monin Bottle Pin',
    'P400': '5pk Sampler Sales Sheet',
    'P410': '2022 Spring Summer Merch Kit',
    'P419': 'Burger and Fry Sales Sheet',
    'P420': 'Counter Card Wooden Block Stand - Wide Slot',
    'P425': '2022/2023 Fall/Winter Merch Kit',
    'P428': 'HomeCrafted Mixers Sell Sheet',
    'P429': 'Immunity Boost Buttons',
    'P430': '2023 Spring/Summer Merch Kit',
    'P430-CC': '2023 Spring/Summer Counter Cards',
    'P430-SR': '2023 Spring/Summer Sgntr Recipe, Hub & Contest',
    'P432': 'Granita Sell Sheet',
    'P433': 'Monin Birthday Card',
    'P435': 'Italian Sale Sheet',
    'P436': 'Pizza Sale Sheet',
    'P437': 'BBQ Sale Sheet',
    'P438': 'Mexican Sale Sheet',
    'P439': 'Asian Sale Sheet',
    'P440': '2023/2024 Fall/Winter Merch Kit',
    'P440-CC': '2023/2024 Fall/Winter Counter Cards',
    'P448': 'Beverage Boost Trifold Sale Sheet',
    'P449': 'Cocktail Mix Concentrates Sell Sheet',
    'P453': 'Strawberry Rose Sale Sheet',
    'P455': '2020/2021 Fall /Winter Merch Kit',
    'P455-SR': '2020/2021 Fall /Winter Recipe Cards',
    'P457': 'Meijer Rack Header  6”x23.5”',
    'P458': 'Meijer Rack Insert Slider 7.3125”x2.875"',
    'P461': 'Florals Sale Sheet',
    'P463': 'Strawberry Rose Counter Card - NAB',
    'P464': 'Strawberry Rose Counter Card - Coffee',
    'P470': '2024 Spring/Summer Merch Kit',
    'P470-CA': '2024 Spring/Summer Merch Kit (Canada)',
    'P470-CC': '2024 Spring/Summer Counter Cards',
    'P470-SR': '2024 Spring/Summer Sgntr Recipe, Hub & Contest',
    'P471': 'ZCNF Seasonal SS - Peppermint & Pumpkin Spice',
    'P472': 'Energy Boost Small Sale Sheet',
    'P473': 'Total Immunity Boost Small Sale Sheet',
    'P475': '2024/2025 Fall/Winter Merch Kit',
    'P475-CA': '2024/2025 Fall/Winter Merch Kit (Canada)',
    'P475-CC': '2024/2025 Fall/Winter Counter Cards',
    'P480': '2025 Spring/Summer Merch Kit',
    'P480-CA': '2025 Spring/Summer Merch Kit (Canada)',
    'P480-CC': '2025 Spring/Summer Counter Cards',
    'P484': 'Generic Window Cling 5" x7"',
    'P485': '2025/2026 Fall/Winter Merch Kit',
    'P485-CA': '2025/2026 Fall/Winter Merch Kit (Canada)',
    'P485-CC': '2025/2026 Fall/Winter Counter Cards',
    'P501': 'Flavor List',
    'P509': 'Twist USB 1 GB',
    'P510': 'Monin Stainless Boston Shaker – 25 & 18oz Tins',
    'P515': 'Monin 8oz Martini Shaker',
    'P517': 'Monin Notecard 4.25 x 5.5',
    'P517A': 'Monin Notecard 5 x 7',
    'P518': 'Monin Envelopes - A2',
    'P518A': 'Monin Envelopes - A7',
    'P519': 'Monin.com Thank You Card',
    'P519A': 'Monin.com Thank You Card - Holiday Edition',
    'P519C': 'Monin.com Thank You Card - FOTY 2024 (Ube)',
    'P519D': 'Monin.com Thank You Card - FOTY 2025 (Yuzu)',
    'P519E': 'Monin.com Thank You Card-FOTY 2026 Toast. Coconut',
    'P520-AH': 'Monin Thank You Note - Alex Helvie',
    'P520-AM': 'Monin Thank You Note - Angie Mikeska',
    'P520-CB': 'Monin Thank You Note - Catherine Bach',
    'P520-CC': 'Monin Thank You Note - Chris Casaccio',
    'P520-CL': 'Monin Thank You Note - Caroline Lovaglio',
    'P520-CM': 'Monin Thank You Note - Cynthia Milana',
    'P520-CM2': 'Monin Thank You Note - Christine Moore',
    'P520-CP': 'Monin Thank You Note -Courtney Payk',
    'P520-DR': 'Monin Thank You Note - Donna Ritchie',
    'P520-DS2': 'Monin Thank You Note - Don Salemi',
    'P520-EP': 'Monin Thank You Note - Emily Prats',
    'P520-HR': 'Monin Thank You Note - Heidi Rosati',
    'P520-HR2': 'Monin Thank You Note - Heather Rady',
    'P520-KA': 'Monin Thank You Note - Kelly Applegate',
    'P520-KM': 'Monin Thank You Note - Ken Mate',
    'P520-KP': 'Monin Thank You Note - Kristina Patterson',
    'P520-LA': 'Monin Thank You Note - Loretta Arthur',
    'P520-LB': 'Monin Thank You Note - Lisa Brown',
    'P520-LC': 'Monin Thank You Note - Lee Ann Carson',
    'P520-LR': 'Monin Thank You Note - Leslie Rhodes',
    'P520-MG': 'Monin Thank You Note - Mary Jo Gigliotti',
    'P520-MH': 'Monin Thank You Note - Michael Harrison',
    'P520-MN': 'Monin Thank You Note - Margot Nouailletas',
    'P520-MY': 'Monin Thank You Note - Matt Yost',
    'P520-NH': 'Monin Thank You Note - Nicole Hannon',
    'P520-NR': 'Monin Thank You Note - Nancy Reynolds',
    'P520-PJ': 'Monin Thank You Note -Patrick Johnson',
    'P520-RM': 'Monin Thank You Note - Ryan Maher',
    'P520-RS': 'Monin Thank You Note - Rebecca Smith',
    'P520-SB': 'Monin Thank You Note - Steven Bishop',
    'P520-SG': 'Monin Thank You Note - Stephen Gray',
    'P520-SK': 'Monin Thank You Note - Spencer Kearns',
    'P520-SL': 'Monin Thank You Note - Samantha Lopez',
    'P520-TM': 'Monin Thank You Note - Tony Munoz',
    'P520-TM2': 'Monin Thank You Note - Trey McInvale',
    'P520-VC': 'Monin Thank You Note - Vittorio Caputi',
    'P524': 'Culinary Trifold',
    'P526-A': 'Sorry - Package 1',
    'P526-B': 'Sorry - Package 2',
    'P530': 'Total Immunity Boost Shot Label',
    'P531': 'Energy Boost Shot Label',
    'P542': 'Bubble Tea Sale Sheet',
    'P547': 'Yuzu Puree Sale Sheet',
    'P552': 'Neutral Beverage Base Sale Sheet',
    'P556': 'Monin Moscow Mule Mug',
    'P558': 'Monin Bev-Naps (bundle of 200)',
    'P559': 'CSR Pillars - One Sheeter',
    'P564': 'Plastic Tiered 6 Bottle Rack',
    'P566': '4 Bottle Display Rack-375ml (Without Ring)',
    'P567': '6 Bottle Display Rack-2 wide, 3 high (no inserts)',
    'P570': 'Monin Black Hat (White Monin Panache Logo)',
    'P571': 'Monin Bib Apron',
    'P577': 'Fdsv Coffee Sheet',
    'P580': 'Flavored Tea/Lemonade Sales Sheet',
    'P585': '4 Bottle Display Rack-750ml/1L',
    'P611S': 'Logo T-shirt (Gray) Small',
    'P613L': 'Logo T-shirt (Cream) Large',
    'P613M': 'Logo T-shirt (Cream) Medium',
    'P613S': 'Logo T-shirt (Cream) Small',
    'P613XL': 'Logo T-shirt (Cream) X Large',
    'P613XXL': 'Logo T-shirt (Cream) XXLarge',
    'P613XXXL': 'Logo T-shirt (Cream) XXXLarge',
    'P614L': 'Monin Panache T-shirt Large',
    'P614M': 'Monin Panache T-shirt Medium',
    'P614S': 'Monin Panache T-shirt Small',
    'P614XL': 'Monin Panache T-shirt X Large',
    'P614XXL': 'Monin Panache T-shirt XX Large',
    'P619': 'Blends Sale Sheet',
    'P622': 'Monin 15 oz. Campfire Mug',
    'P623': 'Monin 17oz. Majestic Coffee Mug',
    'P626': '6 Bottle Display Rack 1 wide, 6 high',
    'P631': '5.5 x 8.5 Acrylic Sign Holder',
    'P633': '4 Bottle Display Rack Label-375ml',
    'P636': 'Bee Initiative Card',
    'P638B': '2024 Trends Brochure',
    'P638C': '2025 Trends Brochure',
    'P638D': '2026 Trends Booklet',
    'P639': 'Monin 16”x25” Microfiber Golf Towel W/ Corner Hook',
    'P641': 'Monin Waist Apron W/ 3-Pocket Pouch',
    'P643': 'Tea & Lemonade Profit Sheet',
    'P646': 'Hot Honey Sell Sheet',
    'P648': 'Agave Sell Sheet',
    'P653': 'Maple Pumpkin Cold Brew Sale Sheet',
    'P654': 'Monin Key Fob Lanyard',
    'P656A': 'Car Air Freshener - Vanilla',
    'P657': 'US Direct Stocking Sheet',
    'P659': '11 Bottle Display Rack Label',
    'P661': 'Why LTO? Sale Sheet',
    'P664': 'Monin Travel Mug 12oz - White',
    'P665': 'Monin Travel Mug 12oz - Orange',
    'P666': 'Monin Counter Card Wooden Block Stand-Narrow Slot',
    'P669': 'Plain Plastic Cup 3oz Softside (100pk Bag)',
    'P670': 'Clear Plastic Cup 5oz Softside (50pk Bag)',
    'P671': 'Monin Logo Plastic Cup 3oz Softside (100pk Bag)',
    'P672': 'Frosted Monin Tumbler 12oz',
    'P673': 'Plastic 16oz Cup (50pk bag)',
    'P684': 'Plain 16oz Insulated Hot Cup (25pk bag)',
    'P704': 'Dark Chocolate Sauce Pump 1/4 oz - Topper Removed',
    'P707': 'Black Sugar Sale Sheet',
    'P708': 'Ube 2024 FOTY Sale Sheet',
    'P709': 'Ube 2024 FOTY Small Sale Sheet',
    'P717': '3 Bottle Display Rack - 64oz Brilliance (Blue)',
    'P718': 'Brilliance Sale Sheet',
    'P719': 'Brilliance 8-Panel Brochure',
    'P724B': 'FOTY Intro Box Kit - Yuzu',
    'P724C': 'FOTY Intro Box Kit - Toasted Coconut',
    'P726': 'Sugar Free Lavender Sale Sheet',
    'P727': 'Sugar Free Lavender Small Sale Sheet',
    'P728': 'Hydration Boost Sale Sheet',
    'P729': 'Hydration Boost Small Sale Sheet',
    'P730L': 'Monin Mens Polo Large - Black',
    'P730M': 'Monin Mens Polo Medium - Black',
    'P730S': 'Monin Mens Polo Small - Black',
    'P730XL': 'Monin Mens Polo XL - Black',
    'P730XXL': 'Monin Mens Polo 2XL - Black',
    'P730XXXL': 'Monin Mens Polo 3XL - Black',
    'P730XXXXL': 'Monin Mens Polo 4XL - Black',
    'P731L': 'Monin Mens Polo Large - Cool Grey',
    'P731M': 'Monin Mens Polo Medium - Cool Grey',
    'P731XL': 'Monin Mens Polo XL - Cool Grey',
    'P731XXL': 'Monin Mens Polo 2XL - Cool Grey',
    'P731XXXXL': 'Monin Mens Polo 4XL - Cool Grey',
    'P732L': 'Monin Mens Polo Large - Navy',
    'P732M': 'Monin Mens Polo Medium - Navy',
    'P732XL': 'Monin Mens Polo XL - Navy',
    'P732XXL': 'Monin Mens Polo 2XL - Navy',
    'P732XXXL': 'Monin Mens Polo 3XL - Navy',
    'P733L': 'Monin Mens Polo Large - Gym Blue',
    'P733M': 'Monin Mens Polo Medium - Gym Blue',
    'P733XL': 'Monin Mens Polo XL - Gym Blue',
    'P733XXL': 'Monin Mens Polo 2XL - Gym Blue',
    'P733XXXL': 'Monin Mens Polo 3XL - Gym Blue',
    'P734L': 'Monin Mens Polo Large - Vivid Pink',
    'P734M': 'Monin Mens Polo Medium - Vivid Pink',
    'P734XXL': 'Monin Mens Polo 2XL - Vivid Pink',
    'P734XXXL': 'Monin Mens Polo 3XL - Vivid Pink',
    'P735L': 'Monin Mens Polo Large - University Red',
    'P735M': 'Monin Mens Polo Medium - University Red',
    'P735XL': 'Monin Mens Polo XL - University Red',
    'P735XXL': 'Monin Mens Polo 2XL - University Red',
    'P735XXXL': 'Monin Mens Polo 3XL - University Red',
    'P736L': 'Monin Womens Polo Large - Blacktop',
    'P736XXXL': 'Monin Womens Polo 3XL - Blacktop',
    'P737L': 'Monin Womens Polo Large - Diesel Grey',
    'P737M': 'Monin Womens Polo Medium - Diesel Grey',
    'P737XL': 'Monin Womens Polo XL - Diesel Grey',
    'P737XXL': 'Monin Womens Polo 2XL - Diesel Grey',
    'P737XXXL': 'Monin Womens Polo 3XL - Diesel Grey',
    'P738L': 'Monin Womens Polo Large - Enzyme Blue',
    'P738M': 'Monin Womens Polo Medium - Enzyme Blue',
    'P738XL': 'Monin Womens Polo XL - Enzyme Blue',
    'P738XXL': 'Monin Womens Polo 2XL - Enzyme Blue',
    'P739L': 'Monin Womens Polo Large - Ripped Red',
    'P739M': 'Monin Womens Polo Medium - Ripped Red',
    'P739S': 'Monin Womens Polo Small - Ripped Red',
    'P739XL': 'Monin Womens Polo XL - Ripped Red',
    'P739XXL': 'Monin Womens Polo 2XL - Ripped Red',
    'P741': 'Monin Logo 8oz Clear Cup',
    'P742': 'Monin Beanie Hat',
    'P743': 'Monin Leather Coaster',
    'P744': '2024 New Products Sale Sheet',
    'P746': 'Monin Logo 8oz Clear Cup (50pk bag)',
    'P747B': 'Monin Desk Calendar 2024',
    'P747C': 'Monin Desk Calendar 2025',
    'P747D': 'Monin Desk Calendar 2026',
    'P748': 'Matcha Green Tea Concentrate Sale Sheet',
    'P749': 'Matcha Green Tea Concentrate Small Sale Sheet',
    'P751': 'Dirty Soda Sale Sheet',
    'P752': 'Monin 20 oz. Clear Plastic Measuring Cup',
    'P753': 'Monin Cool Swipe Screen Cleaner',
    'P754': 'Monin Bottle Brush Cleaner',
    'P756': 'Monin 9" Cutting Board',
    'P757': 'Monin Toddy Tie Cord Organizer',
    'P758': 'Monin Stainles Steel Coffee Scoop/Clip',
    'P759': 'Brilliance Launch Kit',
    'P759A': 'Brilliance Launch Kit 2',
    'P760': 'Black Lanyard Double-Sided Monin Logo - No Panache',
    'P769': '4 or 6 Bottle Rack Front Label (1 Btl Wide)',
    'P775': 'Monin Corp Folder (Almond)',
    'P780': 'Dark Chocolate Sauce Pump, Brown With Topper',
    'P781': 'SF Dark Choc Sauce Pump, Lt. Brown With Topper',
    'P782': 'Caramel, Dulce & Sea Salt Sauce Pmp Amber W Topper',
    'P783': 'White Choc Sauce Pump, Cream With Topper',
    'P785': 'New Cap Generic Pourers (dozen bag)',
    'P787': '2 Bottle Display Rack - 64oz',
    'P796': 'Sauce Display Rack Front Label',
    'P804': 'Monin Black Gift Bag',
    'P814': 'Monin Black Gift Bag - Small',
    'P819': 'Brilliance Corkcicle Tumbler 16oz',
    'P821': 'Brilliance 8.5" Straw W/ Lightning Bolt Topper',
    'P822': '3.5" x 7" Microfiber Sunglass Drawstring Bag',
    'P823': 'Headband & Wristband Combo Set',
    'P824': 'Brilliance Lightning Bolt Stress Toy',
    'P827-R1': 'Retail 5pk 50ml Cocktail Sampler',
    'P827-R1A': 'Retail 5pk 50ml Cocktail Sampler 12pk Case',
    'P827-R3': 'Retail 5pk 50ml Coffee Sampler',
    'P827-R3A': 'Retail 5pk 50ml Coffee Sampler 12pk Case',
    'P827-R4': 'Retail 5pk 50ml Holiday Cheer Sampler',
    'P827-R4A': 'Retail 5pk 50ml Holiday Cheer Sampler 12pk Case',
    'P827-R5': 'Retail 5pk 50ml Floral Sampler',
    'P827-R5A': 'Retail 5pk 50ml Floral Sampler 12pk Case',
    'P827-R6': 'Retail 5pk 50ml Autumn Harvest Sampler',
    'P827-R6A': 'Retail 5pk 50ml Autumn Harvest Sampler 12pk Case',
    'P827-R7': 'Retail 5pk 50ml Brilliance Sampler',
    'P827-R7A': 'Retail 5pk 50ml Brilliance Sampler 12pk Case',
    'P827-R8': 'Retail 5pk 50ml Sugar Free Sampler',
    'P827-R8A': 'Retail 5pk 50ml Sugar Free Sampler 12pk Case',
    'P831': 'Supplies on the Fly Sale Sheet',
    'P832': 'Smoothie Mix Sell Sheet',
    'P842': 'Brilliance Sport Towel',
    'P843A': 'Brilliance Lapel Pin 1.5" W/ 3x2" Backer Card',
    'P843B': 'Brilliance Lapel Pin 1.5" Without Backer Card',
    'P846L': 'Brilliance Crop Top Large - Royal',
    'P846M': 'Brilliance Crop Top Medium - Royal',
    'P846S': 'Brilliance Crop Top Small - Royal',
    'P846XL': 'Brilliance Crop Top XL - Royal',
    'P846XXL': 'Brilliance Crop Top 2XL - Royal',
    'P846XXXL': 'Brilliance Crop Top 3XL - Royal',
    'P847L': 'Brilliance Dri-Fit Tee Large - White',
    'P847M': 'Brilliance Dri-Fit Tee Medium - White',
    'P847S': 'Brilliance Dri Fit Tee Small - White',
    'P847XL': 'Brilliance Dri-Fit Tee XL - White',
    'P847XXL': 'Brilliance Dri-Fit Tee 2XL - White',
    'P847XXXL': 'Brilliance Dri-Fit Tee 3XL - White',
    'P848A': 'Brilliance Napkins-Logo Only (200pk bundle)',
    'P848B': 'Brilliance Napkins-Edison Quote (200pk bundle)',
    'P848C': 'Brilliance Napkins-iPhone Quote (200pk bundle)',
    'P848D': 'Brilliance Napkins-Van Gogh Quote (200pk bundle)',
    'P849A': 'Brilliance Athletic Crew Sports Socks - Pink',
    'P849B': 'Brilliance Athletic Crew Sports Socks - Blue',
    'P851': 'Brilliance Pen',
    'P852': 'Brilliance Display Rack Front Label',
    'P853': 'Brilliance Operator Rebate Sale Sheet',
    'P854': 'Brilliance Silipint Straight Up Pint Glass 16oz',
    'P856': 'Brilliance Broker Incentive Sale Sheet',
    'P857': 'Brilliance Core Coffee Distrib. Rebate Sale Sheet',
    'P858': 'Brilliance 16oz Cup (50pk Bag)',
    'P859': '25 x 50ml Straight Rim Bell Jigger',
    'P861': 'Sparkling Yuzu Small Candle Tin',
    'P862': 'Brilliance Tote',
    'P864': 'Brilliance 50ml Sampler Pack Insert- Sales and Brokers',
    'P866': 'Yuzu Pineapple Sale Sheet',
    'P867': 'Blue Raspberry & SF Blue Raspberry Sale Sheet',
    'P868B': 'TruFlavour by Monin Apron - Dark Green',
    'P869': 'Monin Yuzu Yellow 5"x7" Journal Notebook',
    'P874': 'Handcrafted Sodas Sell Sheet',
    'P876': 'Monin Panache Pen',
    'P878A': 'Brilliance Powerfruit Cntr Card 5.5x8.5"',
    'P878B': 'Brilliance Berry Blue Cntr Card 5.5x8.5"',
    'P878C': 'Brilliance Starfruit Cntr Card 5.5x8.5"',
    'P878D': 'Brilliance Chilled & Charged Cntr Card 5.5x8.5"',
    'P878E': 'Brilliance Refresh & Recharge Cntr Card 5.5x8.5"',
    'P878F': 'Brilliance Flavored Cntr Card 5.5x8.5"',
    'P878G': 'Brilliance Electric Energy Cntr Card 5.5x8.5"',
    'P879A': 'Brilliance Window Cling 5.5" x 8.5"',
    'P881A': 'Brilliance Flavored Energy Poster 11x17',
    'P881B': 'Brilliance Electric Energy Poster 11x17',
    'P882': 'GFS Brilliance Pick 4 Program Sale Sheet',
    'P884': 'Clear View Poly Mailers - 14 x 17"',
    'P891': 'Monin Bamboo Rimmer',
    'P892': 'Monin Coconut Bowl',
    'P893': 'Monin Silicone Spatula W/ Wood Handle',
    'P909A': 'TruFlavour Launch Kit - Hybrid',
    'P909B': 'TruFlavour Launch Kit - Culinary',
    'P918B': '2026 FOTY Sale Sheet - Toasted Coconut',
    'P919A': '2025 FOTY Small Sale Sheet - Yuzu',
    'P923A': '2026 FOTY Standard Sample Letter-Toasted Coconut',
    'P926': 'TruFlavour by Monin Square Coaster Set in Tin',
    'P927': 'TruFlavour by Monin Jute Tote',
    'PM007-A': '1L Puree Bottle',
    'PM007-C': '1L Puree Bottle',
    'PM008-B': '1L Puree Btl Pour Cap w/ Seal',
    'PM014': 'Puree 4pk 1L Box - Black Pour Caps',
    'PM024': 'Partition 12pk - 750m Glass',
    'PM036': '12pk 750ml Glass Box',
    'PM036W2': '12pk 750ml Glass Wrap Box C Flute',
    'PM052-B': '43mm White TE Cap for 64oz (CSI)',
    'PM056-A': 'Plastic Caps Black 28 mm TE Un',
    'PM059-A': '1 liter PET bottle',
    'PM075': '4pk 1 Liter Box "B" Flute',
    'PM087-A': '43mm Silver Cap',
    'PM097-A': 'Signature 64oz Bottle',
    'PM098': 'Sauce Box 4pk-64oz',
    'PM113': 'Waste Material - Unusable Syrup',
    'PM121': 'Smoothie Mixed Case Shipper',
    'PM123': '50ml Glass 120pk Box',
    'PM141': 'White Puree 4pk Box – Silver Flat Caps',
    'PM143A': '1pk Plain Shipper Box Kraft - Hinged Pulp',
    'PM143B': '2pk Shipper Box - Hinged Pulp',
    'PM143C': '3pk Shipper Box - Hinged Pulp',
    'PM143D': '4pk Shipper Box - Hinged Pulp',
    'PM143E': '6pk Shipper Box - Hinged Pulp',
    'PM143F': '8pk Shipper Box - Hinged Pulp',
    'PM143G': '12pk Shipper Box - Hinged Pulp',
    'PM166': 'Conc Flavors 375ml 4pk Box (Larger Bottle)',
    'PM172': '**Branded 12 x 5 x 4 Kraft Box**',
    'PM181': '2oz Square PET Bottle W/ 26mm 1914 Neck',
    'PM183': '10oz Squat Round PET Bottle W/ 38mm DBJ Neck',
    'PM184': '33mm DBJ White HDPE Tamper Evident Screw Cap',
    'PM192': 'Universal Custom Hinged Pulp - 750ml & 1L',
    'PM204': '26mm Natural HDPE Cap',
    'PM603-B': '**3pk 50mL Glass Box-Carton Only (Clear PVC)**',
    'PM603-I': '3pk 50mL Glass Box-Insert Only (Clear PVC)',
    'PM607-B': 'Customizable 50ml Sampler Box (3-4 Btl)',
    'PM607-IB': '4 Bottle Insert for Customizable Box',
    'PM789': 'Merchandising Kit Box 11 x 8 x 3"',
    'PM826': '5pk 50ml Glass Box',
    'PM827-R3': 'Retail 5pk 50ml Coffee Recipe Card',
    'PM827-R4': 'Retail 5pk 50ml Holiday Recipe Card',
    'PM827-R7': 'Retail 5pk 50ml Brilliance Recipe Card',
    'PM827-R8A': 'Retail 5pk 50ml Sugar Free Recipe Card (Automatn)',
    'RI072': 'Nat. Blueberry Flavor WONF 14% 7.92640',
    'RI346': 'Watermelon Juice Conc, Cloudy',
    'RI607': 'Nat Caramelized Sugar Type PG Free',
    'RI891': 'proVontage EXP',
    'WBS-FR045U': 'Vanilla 1 L',
    'WBS-FR061U': 'Lavender 1L',
    'WBS-FR195U': 'Blue Raspberry 1 L',
}

def lookup_item_code(description, pack_size=''):
    """Try to find a Monin item code from a description and optional pack size.
    Uses ITEM_LOOKUP reverse search. Returns best match code or None.
    pack_size: '4/1 LT', '12/750', '750 ML', '1 L', etc.
    """
    if not description:
        return None
    desc_norm = description.upper().strip()

    # Determine pack type from pack_size string
    pack_hint = ''
    if pack_size:
        ps = pack_size.upper().replace(' ', '')
        if re.search(r'^12/', ps): pack_hint = '12pk'
        elif re.search(r'^6/', ps): pack_hint = '6pk'
        elif re.search(r'^4/', ps): pack_hint = '4pk'
        elif re.search(r'^1/', ps) or re.search(r'750ML|750$', ps): pack_hint = 'single'
        elif re.search(r'1L$|1LT$', ps): pack_hint = 'single'

    # Pack suffix map for Monin codes
    pack_suffix = {'4pk': 'F', '12pk': 'A', '6pk': 'B', 'single': 'U'}

    candidates = []
    for code, item_desc in ITEM_LOOKUP.items():
        item_norm = item_desc.upper()
        # Check if key words from the distributor description appear in item description
        words = [w for w in re.split(r'[\s/]+', desc_norm) if len(w) > 2
                 and w not in ('THE','AND','FOR','WITH','MONIN','BEVERAGE','SYRUP','PUREE','SAUCE')]
        if not words:
            continue
        matches = sum(1 for w in words if w in item_norm)
        if matches >= max(1, len(words) - 1):
            candidates.append((code, item_desc, matches))

    if not candidates:
        return None

    # If pack_hint given, prefer items whose description matches the pack type
    pack_kw_map = {'4pk': '4PK', '12pk': '12PK', '6pk': '6PK', 'single': None}
    if pack_hint and pack_hint != 'single':
        kw = pack_kw_map.get(pack_hint, '')
        preferred = [c for c in candidates if kw and kw in c[1].upper()]
        if preferred:
            candidates = preferred
    elif pack_hint == 'single':
        preferred = [c for c in candidates if '4PK' not in c[1].upper()
                     and '12PK' not in c[1].upper() and '6PK' not in c[1].upper()]
        if preferred:
            candidates = preferred

    # Sort by match count desc, then code length asc (shorter = more specific)
    candidates.sort(key=lambda x: (-x[2], len(x[0])))
    return candidates[0][0]


# ─── EXACT TELLUS SHEET1 COLUMNS (20 cols, no blank first col) ───────────────
S1_COLS = [
    'Source', 'Vendor #', 'Program #', 'Customer Reference',
    'Payee Type', 'Distributor ID', 'Operator ID', 'Closing Method',
    'Billback Date', 'BB Start Date', 'BB End Date',
    'Item Number', 'Item Volume Qty', 'Item Dollar Amount',
    'Associated Distributor ID', 'UOM', 'Trade Indicator',
    'Lump Sum Indicator', 'File Name', 'Component'
]

# ─── DEFAULT SUPPLIER CONFIG ─────────────────────────────────────────────────
DEFAULT_SUPPLIER_CONFIG = {
    'KAST':       {'program_num': '1004089', 'dist_id': '134810000', 'trade': 'D'},
    'SOFO':       {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'PFS':        {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'LABATT':     {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'Y_HATA':     {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'BEK':        {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'NICH_CO':    {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'SHAMROCK':   {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'DOT_CBBB':   {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'MCLANE':     {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'S_AND_W':    {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'BLAIR_CANDY':{'program_num': '',        'dist_id': '',           'trade': 'D'},
    'CHENEY':      {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'HARBOR':      {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'MARTIN_BROS':  {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'DOT_FOODS_BB': {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'DRISCOLL':     {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'TANKERSLEY':    {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'CHRIST_PANOS':  {'program_num': '',        'dist_id': '',           'trade': 'D',
        'did_map': {
            '300514': 'M-FR006F',
            '300513': 'M-FR008F',
            '380169': 'M-RP008F',
            '380177': 'M-FR112F',
            '300503': 'M-FR009F',
            '380178': 'M-FR268F',
            '300505': 'M-FT080F',
            '380176': 'M-FR012F',
            '300533': 'M-FR247F',
            '300144': 'M-FR190F',
            '300143': 'M-FR021F',
            '300140': 'M-FR023F',
            '300512': 'M-FR032F',
            '300504': 'M-FR036F',
            '300507': 'M-FR050F',
            '380173': 'M-FR039F',
            '300500': 'M-FR075F',
            '300502': 'M-FR040F',
            '300171': 'M-GC009FP',
            '300149': 'M-GC062FP',
            '300159': 'M-GC063FP',
            '300510': 'M-FR042F',
            '300158': 'M-RP042F',
            '300501': 'M-FR045F',
            '300508': 'M-FS045F',
            '300509': 'M-FR136F',
        }
    },
    'HENRY_FOODS':  {'program_num': '',        'dist_id': '',           'trade': 'D',
        'did_map': {
            '1282177': 'P585',
            '1282193': 'P240',
            '1454826': 'M-AR009A',
            '1454867': 'M-AR063A',
            '1454909': 'M-AR045A',
            '1454925': 'M-AR023A',
            '1454928': 'M-AR084A',
            '1454933': 'M-AO157B',
            '1454982': 'M-AS009A',
        }
    },
    'DELCO_FOODS':  {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'CHEFS_WH':     {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'ATLAS':        {'program_num': '',        'dist_id': '',           'trade': 'D'},
    'KOHL_WH':      {'program_num': '',        'dist_id': '',           'trade': 'D',
        'product_map': {
            '399093': 'M-FR049F',
            '399159': 'M-AT080A',
            '464377': 'M-FR114F',
            '501756': 'M-FR145F',
            '522704': 'M-FR092F',
            '538963': 'M-FR195F',
            '601578': 'M-FR197F',
            '641090': 'M-FR061F',
            '656163': 'M-FR252F',
            '63511':  'P240',
            '65228':  'P245',
            '67896':  'P230',
        }
    },
    'UNKNOWN':      {'program_num': '',        'dist_id': '',           'trade': 'D'},
}

# ─── DATE HELPERS ─────────────────────────────────────────────────────────────
def to_yyyymmdd(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    if isinstance(val, str):
        val = val.strip().replace('/', '-')
        for fmt in ('%m-%d-%Y','%Y-%m-%d','%m-%d-%y','%Y%m%d','%d-%b-%Y'):
            try:
                return datetime.strptime(val, fmt).strftime('%Y%m%d')
            except: pass
        digits = re.sub(r'\D','',val)
        if len(digits) == 8: return digits
        return val
    if isinstance(val, datetime):
        return val.strftime('%Y%m%d')
    if hasattr(val, 'strftime'):
        return val.strftime('%Y%m%d')
    return str(val)

def parse_date_range(text):
    m = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*[-–to]+\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', str(text))
    if m:
        return to_yyyymmdd(m.group(1).replace('-','/')), to_yyyymmdd(m.group(2).replace('-','/'))
    return '', ''

def month_name_to_range(period_str):
    """Convert 'December 2025' → ('20251201', '20251231')"""
    import calendar
    months = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
              'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}
    s = str(period_str).strip().lower()
    for name, num in months.items():
        if name in s:
            m = re.search(r'(\d{4})', s)
            if m:
                year = int(m.group(1))
                last = calendar.monthrange(year, num)[1]
                start = f'{year}{num:02d}01'
                end   = f'{year}{num:02d}{last:02d}'
                return start, end
    return '', ''

def clean_amount(val):
    if val is None: return 0.0
    s = str(val).replace('$','').replace(',','').strip()
    neg = s.startswith('(') and s.endswith(')')
    s = s.strip('()')
    try:
        v = float(s)
        return -v if neg else v
    except: return 0.0

# ─── TELLUS ROW BUILDER (returns dict keyed by S1_COLS) ──────────────────────
def make_row(source='', program_num='', customer_ref='', dist_id='',
             bill_date='', start_date='', end_date='',
             item='', qty=0, amount=0.0, trade='D', operator_id=''):
    dist_id_raw = str(dist_id).strip()
    dist_id_int = dist_id_raw if dist_id_raw else None  # always string — preserves leading zeros
    try:
        qty_int = round(float(str(qty))) if qty not in ('', None) else 0
    except:
        qty_int = 0
    try:
        amt_float = round(float(str(amount)), 4) if amount not in ('', None) else 0.0
    except:
        amt_float = 0.0
    return {
        'Source':                  str(source),
        'Vendor #':                ' ',
        'Program #':               str(program_num) if program_num else ' ',
        'Customer Reference':      str(customer_ref),
        'Payee Type':              'D',
        'Distributor ID':          dist_id_int,
        'Operator ID':             str(operator_id) if operator_id else ' ',
        'Closing Method':          'D',
        'Billback Date':           str(bill_date) if bill_date else TODAY,
        'BB Start Date':           str(start_date),
        'BB End Date':             str(end_date),
        'Item Number':             str(item),
        'Item Volume Qty':         qty_int,
        'Item Dollar Amount':      amt_float,
        'Associated Distributor ID': dist_id_int,
        'UOM':                     'CS',
        'Trade Indicator':         str(trade),
        'Lump Sum Indicator':      None,
        'File Name':               None,
        'Component':               ' ',
    }

# ─── FORMAT DETECTOR ─────────────────────────────────────────────────────────
def detect_supplier(filename):
    fn = os.path.basename(filename).upper()
    if 'KAST' in fn: return 'KAST'
    if 'SOFO' in fn: return 'SOFO'
    if 'PFS' in fn: return 'PFS'
    if re.search(r'BIDRRPT|BID[_-]?REPT|BIDBILL', fn): return 'PFS'
    if 'BLAIR' in fn: return 'BLAIR_CANDY'
    if 'LABATT' in fn: return 'LABATT'
    if 'SHAMROCK' in fn: return 'SHAMROCK'
    if re.search(r'S[\s_]*AND[\s_]*W|S\s*&\s*W', fn): return 'S_AND_W'
    if 'BEK' in fn: return 'BEK'
    if 'NICH' in fn: return 'NICH_CO'
    if 'CBBB' in fn: return 'DOT_CBBB'
    if 'TANKERSLEY' in fn: return 'TANKERSLEY'
    if re.search(r'CHRIST.*PANOS|PANOS.*CHRIST', fn): return 'CHRIST_PANOS'
    if re.search(r"HENRY.{0,4}FOOD|PURCHASE.DETAIL", fn): return 'HENRY_FOODS'
    if 'DELCO' in fn: return 'DELCO_FOODS'
    if re.search(r'CHEFS.{0,6}WH|CHEFSWAREHOUSE|DAIRYLAND', fn): return 'CHEFS_WH'
    if re.search(r'ATLAS[\s_]?WHOLESALE|ATLAS[\s_]?FOOD', fn): return 'ATLAS'
    if re.search(r'Y[\s.]?HATA|Y_HATA|TM\s+\d{6}', fn): return 'Y_HATA'
    if 'DRISCOLL' in fn: return 'DRISCOLL'
    if 'KOHL' in fn: return 'KOHL_WH'
    if 'HARBOR' in fn: return 'HARBOR'
    if 'SUPPLIER BILLBACK' in fn or 'SUPPLIER_BILLBACK' in fn: return 'HARBOR'
    m = re.search(r'CUST\s+(\d+)', fn.replace('  ',' '))
    if m:
        cust = m.group(1)
        if cust in ('15101', '141699') and fn.endswith('.XLSX'): return 'MCLANE_OR_DOT'
        if fn.endswith('.PDF'): return 'PFS_STYLE'
    return 'UNKNOWN'

# ─── PARSERS ─────────────────────────────────────────────────────────────────

def parse_bek_or_nich(filepath, source_name, cfg, customer_ref):
    # PDFs use Trackmax format — route through content-based detector
    if filepath.lower().endswith('.pdf'):
        return parse_supplier_billback_pdf(filepath, cfg, customer_ref)
    rows = []
    try:
        df = pd.read_csv(filepath, sep='\t', dtype=str, low_memory=False)
        df.columns = df.columns.str.strip()
        for _, r in df.iterrows():
            item = str(r.get('MID','')).strip()
            if not item or item == 'nan' or not item.startswith('M-'): continue
            inv = str(r.get('InvoiceNumber','')).strip()
            bill = to_yyyymmdd(str(r.get('InvoiceDate','')).strip())
            start, end = parse_date_range(str(r.get('ProgramPeriod','')))
            qty = r.get('Quantity','0')
            amt = clean_amount(r.get('ExtendedAmount','0'))
            rows.append(make_row(
                source=source_name,
                program_num=cfg['program_num'],
                customer_ref=customer_ref or inv,
                dist_id=cfg['dist_id'],
                bill_date=bill,
                start_date=start,
                end_date=end,
                item=item,
                qty=qty,
                amount=amt,
                trade=cfg['trade']
            ))
    except Exception as e:
        rows.append({'_error': str(e)})
    return rows

def parse_shamrock(filepath, cfg, customer_ref):
    """Shamrock Foods XLSX billback parser.

    Supports two formats:
      A) Old format — header row contains 'MFG PROD #'; simple per-row M-code + amount.
      B) New format — header row 1 contains 'MFG ID' (col AA), 'PA NAME' (col H),
         'BRANCH' (col D). Each row is one invoice line; aggregates per operator × M-code.
         Emits Trade=O with Operator ID = PA Name.
    """
    rows = []
    try:
        wb = __import__('openpyxl').load_workbook(filepath, data_only=True)
        ws = wb.active
        header_row = None
        col_map = {}

        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_vals = [str(v).strip() if v is not None else '' for v in row]
            row_upper = [v.upper() for v in row_vals]
            # Old format: "MFG PROD #" / "MFG PROD#"
            if any('MFG' in c and 'PROD' in c for c in row_upper):
                header_row = i
                col_map = {v.upper(): j for j, v in enumerate(row_vals)}
                break
            # New format: "MFG ID" column present alongside "PA NAME" / "BRANCH"
            if 'MFG ID' in row_upper and ('PA NAME' in row_upper or 'PA NUMBER' in row_upper):
                header_row = i
                col_map = {v.upper(): j for j, v in enumerate(row_vals)}
                break

        if header_row is None:
            return [{'_error': 'SHAMROCK: header not found'}]

        def gc(key_options):
            for k in key_options:
                for ck, ci in col_map.items():
                    if k in ck: return ci
            return None

        # ── New format (MFG ID / PA NAME / BRANCH) ───────────────────────────
        if 'MFG ID' in col_map:
            ci_item     = col_map.get('MFG ID')
            ci_qty      = gc(['QUANTITY', 'QTY'])
            ci_amt      = gc(['TOTAL ALLOWANCE', 'ALLOWANCE/CASE', 'ALLOW'])
            ci_bill     = gc(['PAYMENT DATE', 'INVOICE DATE', 'BILL DATE'])
            ci_operator = gc(['PA NAME', 'PA NUMBER'])
            ci_branch   = gc(['BRANCH'])
            ci_ref      = gc(['MEMO NUMBER', 'MEMO NUM'])

            from collections import defaultdict
            # Key: (operator, item_code) → {qty, amt, bill_date, source, cref}
            agg_qty = defaultdict(float)
            agg_amt = defaultdict(float)
            agg_meta = {}  # first-seen bill_date + source per key

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                rv = [str(v).strip() if v is not None else '' for v in row]
                if not any(rv): continue
                item = rv[ci_item].strip().upper() if ci_item is not None else ''
                if not item or not item.startswith('M-'): continue

                operator = rv[ci_operator].strip() if ci_operator is not None else ''
                qty_raw  = rv[ci_qty] if ci_qty is not None else 0
                amt_raw  = rv[ci_amt] if ci_amt is not None else 0
                bill     = to_yyyymmdd(rv[ci_bill] if ci_bill is not None else '')
                branch   = rv[ci_branch].strip() if ci_branch is not None else 'Shamrock'
                cref     = rv[ci_ref].strip() if ci_ref is not None else (customer_ref or '')

                # Use customer_ref override if provided, else from Memo Number column
                if customer_ref:
                    cref = customer_ref

                key = (operator, item)
                agg_qty[key] += float(qty_raw or 0)
                agg_amt[key] += clean_amount(amt_raw)
                if key not in agg_meta:
                    agg_meta[key] = {'bill': bill, 'source': branch, 'cref': cref}

            for (operator, item), meta in agg_meta.items():
                rows.append(make_row(
                    source=meta['source'],
                    program_num=cfg['program_num'],
                    customer_ref=meta['cref'],
                    dist_id=cfg['dist_id'],
                    bill_date=meta['bill'],
                    start_date='',
                    end_date='',
                    item=item,
                    qty=round(agg_qty[(operator, item)], 4),
                    amount=round(agg_amt[(operator, item)], 2),
                    trade='O',
                    operator_id=operator
                ))

        # ── Old format (MFG PROD #) ───────────────────────────────────────────
        else:
            ci_item   = gc(['MFG PROD'])
            ci_qty    = gc(['QUANTITY', 'QTY'])
            ci_amt    = gc(['NET ALLOW', 'ALLOW', 'AMOUNT'])
            ci_bill   = gc(['INVOICE DATE', 'BILL DATE'])
            ci_start  = gc(['START', 'FROM'])
            ci_end    = gc(['END', 'TO', 'THRU'])
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                rv = [str(v).strip() if v is not None else '' for v in row]
                if not any(rv): continue
                item = rv[ci_item] if ci_item is not None else ''
                if not item or not item.upper().startswith('M-'): continue
                qty  = rv[ci_qty]  if ci_qty  is not None else 0
                amt  = clean_amount(rv[ci_amt] if ci_amt is not None else 0)
                bill = to_yyyymmdd(rv[ci_bill]  if ci_bill  is not None else '')
                start= to_yyyymmdd(rv[ci_start] if ci_start is not None else '')
                end  = to_yyyymmdd(rv[ci_end]   if ci_end   is not None else '')
                rows.append(make_row(
                    source='Shamrock',
                    program_num=cfg['program_num'],
                    customer_ref=customer_ref,
                    dist_id=cfg['dist_id'],
                    bill_date=bill,
                    start_date=start,
                    end_date=end,
                    item=item,
                    qty=qty,
                    amount=amt,
                    trade=cfg['trade']
                ))
    except Exception as e:
        rows.append({'_error': str(e)})
    return rows

def parse_dot_cbbb(filepath, cfg, customer_ref):
    rows = []
    try:
        wb = __import__('openpyxl').load_workbook(filepath, data_only=True)
        ws = wb.active
        header_row = None
        col_map = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_vals = [str(v).strip().upper() if v else '' for v in row]
            if 'ITEM #' in row_vals or 'ITEM#' in row_vals or any('ITEM' in c and '#' in c for c in row_vals):
                header_row = i
                col_map = {v: j for j, v in enumerate(row_vals)}
                break
        if header_row is None:
            return [{'_error': 'DOT_CBBB: header not found'}]
        def gc(keys):
            for k in keys:
                for ck, ci in col_map.items():
                    if k in ck: return ci
            return None
        ci_item  = gc(['ITEM #','ITEM#'])
        ci_qty   = gc(['QTY','QUANTITY','CASES'])
        ci_amt   = gc(['ALLOW AMT','ALLOWANCE','AMOUNT'])
        ci_bill  = gc(['INV DATE','INVOICE DATE','DATE'])
        ci_start = gc(['START'])
        ci_end   = gc(['END','THRU'])
        for row in ws.iter_rows(min_row=header_row+1, values_only=True):
            row_vals = [str(v).strip() if v is not None else '' for v in row]
            if not any(row_vals): continue
            item = row_vals[ci_item] if ci_item is not None else ''
            if not item or not item.upper().startswith('M-'): continue
            qty  = row_vals[ci_qty]  if ci_qty  is not None else 0
            amt  = clean_amount(row_vals[ci_amt]  if ci_amt  is not None else 0)
            bill = to_yyyymmdd(row_vals[ci_bill]  if ci_bill is not None else '')
            start= to_yyyymmdd(row_vals[ci_start] if ci_start is not None else '')
            end  = to_yyyymmdd(row_vals[ci_end]   if ci_end  is not None else '')
            rows.append(make_row(
                source='DOT_CBBB',
                program_num=cfg['program_num'],
                customer_ref=customer_ref,
                dist_id=cfg['dist_id'],
                bill_date=bill,
                start_date=start,
                end_date=end,
                item=item,
                qty=qty,
                amount=amt,
                trade=cfg['trade']
            ))
    except Exception as e:
        rows.append({'_error': str(e)})
    return rows

def parse_mclane(filepath, cfg, customer_ref):
    rows = []
    try:
        wb = __import__('openpyxl').load_workbook(filepath, data_only=True)
        ws = wb.active
        header_row = None
        col_map = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_vals = [str(v).strip().upper() if v else '' for v in row]
            if any('M-CODE' in c or 'MCODE' in c or ('ITEM' in c and 'CODE' in c) for c in row_vals):
                header_row = i
                col_map = {v: j for j, v in enumerate(row_vals)}
                break
        if header_row is None:
            return [{'_error': 'MCLANE: header not found'}]
        def gc(keys):
            for k in keys:
                for ck, ci in col_map.items():
                    if k in ck: return ci
            return None
        ci_item  = gc(['M-CODE','MCODE','ITEM CODE','ITEM #'])
        ci_qty   = gc(['QTY','QUANTITY','CASES'])
        ci_amt   = gc(['ALLOW','AMOUNT','NET AMT'])
        ci_bill  = gc(['INV DATE','INVOICE DATE','DATE'])
        ci_start = gc(['START'])
        ci_end   = gc(['END','THRU'])
        for row in ws.iter_rows(min_row=header_row+1, values_only=True):
            row_vals = [str(v).strip() if v is not None else '' for v in row]
            if not any(row_vals): continue
            item = row_vals[ci_item] if ci_item is not None else ''
            if not item or not item.upper().startswith('M-'): continue
            qty  = row_vals[ci_qty]  if ci_qty  is not None else 0
            amt  = clean_amount(row_vals[ci_amt]  if ci_amt  is not None else 0)
            bill = to_yyyymmdd(row_vals[ci_bill]  if ci_bill is not None else '')
            start= to_yyyymmdd(row_vals[ci_start] if ci_start is not None else '')
            end  = to_yyyymmdd(row_vals[ci_end]   if ci_end  is not None else '')
            rows.append(make_row(
                source='McLane',
                program_num=cfg['program_num'],
                customer_ref=customer_ref,
                dist_id=cfg['dist_id'],
                bill_date=bill,
                start_date=start,
                end_date=end,
                item=item,
                qty=qty,
                amount=amt,
                trade=cfg['trade']
            ))
    except Exception as e:
        rows.append({'_error': str(e)})
    return rows

def parse_sw_pdf(filepath, cfg, customer_ref, source_override=''):
    """S&W Wholesale Foods Trackmax PDF format.
    Columns: Inv.Number Inv.Date CustomerID CustomerName Brand PackSize Description
             ProductID(6-digit) DID(####CS) UPC Quantity TotalWeight ProgramAmount AmountDue
    Sums by ProductID since one product appears across many store invoices.
    """
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            all_text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
        # No full-text dedup here — short lines like "STRAWBERRY Cost" repeat legitimately.
        # Instead we deduplicate at the match level using a seen_matches set below.

        # Normalize M-codes where PDF extraction emits a space instead of a dash
        # e.g. "M FR210F" → "M-FR210F"
        all_text = re.sub(r'\bM\s+([A-Z]{2}\d{3,4}[A-Z0-9]*)\b', r'M-\1', all_text)

        # Dates
        bill_date = start_date = end_date = ''
        m = re.search(r'(?:generated|posted)\s+on\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: bill_date = to_yyyymmdd(m.group(1))
        m = re.search(r'between\s+(\d{1,2}/\d{1,2}/\d{4})\s+and\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: start_date, end_date = to_yyyymmdd(m.group(1)), to_yyyymmdd(m.group(2))

        # Invoice / customer ref
        inv_num = ''
        m = re.search(r'Invoice\s+Number[:\s]+(\d+)', all_text, re.I)
        if m: inv_num = m.group(1).strip()
        cref = customer_ref or inv_num

        # Source name: use user-selected distributor if provided, else auto-detect from IMPORTANT! header
        if source_override:
            source_name = source_override
        else:
            source_name = 'S&W'
            m2 = re.search(r'IMPORTANT!\s*\n([^\n]+)', all_text)
            if m2: source_name = m2.group(1).strip()

        # Line pattern: positive invoice lines
        # Format: ... MONIN <PackSize> <Description> <ProductID> <DID>CS <UPC> <Qty> <Wt> BB To .../unit $<Amt>
        line_pat = re.compile(
            r'MONIN\s+([\d/.]+(?:\s+\d+)?(?:\s*(?:LITER|LITERS|LTR|LT|ML|L|OZ|EA|Z))?)\s+'  # PackSize (Z = oz abbrev)
            r'(?:\S+\s+)*?'                            # Description words (non-greedy)
            r'([A-Z]-[A-Z0-9]+|[A-Z]\d{3,5}|\d{5,6})\s+'  # Product ID
            r'(?:\*N\*|\w{3,8})\s+'                   # DID (*N* = no distributor ID, or alphanumeric)
            r'(?:(?:\d{6,16}|[A-Z]-[A-Z0-9]+)\s+){0,2}'  # UPC: 0-2 elements (barcode OR M-code in UPC slot)
            r'(\d[\d,]*\.?\d*)\s+'                     # Quantity (must start with digit — rejects credit lines)
            r'[\d,.]+\s+'                              # Total Weight
            r'(?:\$[\d,.]+\s+){0,2}'                   # 0-2 pre-amounts
            r'(?:BB\s+To\s+[\d.]+/unit|[\d.]+\s*%\s+of\s+(?:FOB|Del)(?:Charges)?(?:\s+Cost)?)\s+'
            r'\$([\d,.]+)',                             # Amount Due
            re.I
        )

        # Credit/return line pattern — qty and amount are in parentheses: (1.00) (weight) ($7.98)
        credit_pat = re.compile(
            r'MONIN\s+([\d/.]+(?:\s+\d+)?(?:\s*(?:LITER|LITERS|LTR|LT|ML|L|OZ|EA|Z))?)\s+'
            r'(?:\S+\s+)*?'
            r'([A-Z]-[A-Z0-9]+|[A-Z]\d{3,5}|\d{5,6})\s+'
            r'(?:\*N\*|\w{3,8})\s+'
            r'(?:(?:\d{6,16}|[A-Z]-[A-Z0-9]+)\s+){0,2}'  # UPC: 0-2 elements (barcode OR M-code in UPC slot)
            r'\((\d[\d,]*\.?\d*)\)\s+'   # qty in parens
            r'\([\d,.]+\)\s+'            # weight in parens
            r'(?:(?:\$[\d,.]+|\(\$[\d,.]+\))\s+){0,2}'  # 0-2 amounts: plain $X.XX or ($X.XX) for credit lines
            r'(?:BB\s+To\s+[\d.]+/unit|[\d.]+\s*%\s+of\s+(?:FOB|Del)(?:Charges)?(?:\s+Cost)?)\s+'
            r'\(\$([\d,.]+)\)',           # amount in parens — negative
            re.I
        )

        # Processing fee pattern: "MONIN: <non-M-code> qty wt BB To <neg-rate>/unit $<amt>"
        fee_pat = re.compile(
            r'MONIN:\s+(\S+)\s+[\d.]+\s+[\d.]+\s+BB\s+To\s+[-\d.]+/unit\s+\$([\d,.]+)',
            re.I
        )

        prod_pack  = {}   # prod_id → pack size
        prod_desc  = {}   # prod_id → description text (for numeric ID → M-code lookup)
        prod_did   = {}   # prod_id → DID value (for did_map lookup)
        did_map    = cfg.get('did_map', {})

        # Aggregate by product ID
        from collections import defaultdict
        totals_qty = defaultdict(float)
        totals_amt = defaultdict(float)
        seen_matches = set()   # deduplicate page-boundary repeated MONIN lines

        for m in line_pat.finditer(all_text):
            # Key = full source line (includes invoice# before MONIN), so identical
            # product lines from different invoices are NOT deduplicated.
            line_start = all_text.rfind('\n', 0, m.start()) + 1
            match_key = all_text[line_start:m.end()][:200]
            if match_key in seen_matches:
                continue
            seen_matches.add(match_key)
            pack_size = m.group(1).strip()
            prod_id   = m.group(2).upper()
            # If ProductID is numeric, check product_map first, then try UPC-slot M-code extraction
            # (Kohl Wholesale: 6-digit ProductID, M-code lives in UPC column)
            if re.match(r'^\d+$', prod_id):
                product_map = cfg.get('product_map', {})
                # Strip leading zeros for product_map lookup (PDF may pad to 6 digits)
                prod_id_stripped = prod_id.lstrip('0') or prod_id
                if prod_id_stripped in product_map:
                    prod_id = product_map[prod_id_stripped]
                elif prod_id in product_map:
                    prod_id = product_map[prod_id]
                else:
                    upc_m = re.search(
                        rf'{re.escape(prod_id)}\s+(?:\*N\*|\w{{3,8}})\s+(?:\d+\s+)?([A-Z]-[A-Z0-9]+|[A-Z]\d{{3,5}})',
                        m.group(0), re.I)
                    if upc_m:
                        prod_id = upc_m.group(1).upper()
            qty       = float(m.group(3))
            amt       = float(m.group(4).replace(',', ''))
            totals_qty[prod_id] += qty
            totals_amt[prod_id] += amt
            # Capture DID (the \w{3,8} group right after product ID)
            if prod_id not in prod_did:
                did_m2 = re.search(rf'{re.escape(prod_id)}\s+(\w{{3,8}})\s', m.group(0))
                if did_m2:
                    prod_did[prod_id] = did_m2.group(1)
            if prod_id not in prod_pack:
                prod_pack[prod_id] = pack_size
            # Extract description (text between PackSize and ProductID) for numeric ID lookup.
            # pdfplumber sometimes wraps description words to the next line, so we also
            # grab the line immediately following the match position.
            if prod_id not in prod_desc and re.match(r'^\d+$', prod_id):
                dm = re.search(
                    rf'MONIN\s+{re.escape(pack_size)}\s+(.*?)\s+{re.escape(prod_id)}\b',
                    m.group(0), re.I | re.S)
                desc_text = re.sub(r'\s+', ' ', dm.group(1).strip()) if dm else ''
                # If description is missing or just one very short word, check next line
                words_found = [w for w in desc_text.split()
                               if w.upper() not in ('SYRUP','SAUCE','PUREE','MONIN','')]
                if not words_found:
                    next_line_m = re.search(
                        rf'{re.escape(prod_id)}\b[^\n]*\n([A-Z][A-Z ]+?)(?:\s+Cost|\s+\d|\s*$)',
                        all_text, re.I)
                    if next_line_m:
                        desc_text = (desc_text + ' ' + next_line_m.group(1).strip()).strip()
                if desc_text:
                    prod_desc[prod_id] = desc_text

        # Subtract credit/return lines (qty and amount in parentheses)
        seen_credits = set()
        for m in credit_pat.finditer(all_text):
            line_start = all_text.rfind('\n', 0, m.start()) + 1
            credit_key = all_text[line_start:m.end()][:200]
            if credit_key in seen_credits:
                continue
            seen_credits.add(credit_key)
            prod_id = m.group(2).upper()
            qty     = float(m.group(3))
            amt     = float(m.group(4).replace(',', ''))
            totals_qty[prod_id] -= qty
            totals_amt[prod_id] -= amt

        if not totals_qty:
            return [{'_error': 'S&W PDF: no product rows found — format may have changed'}]

        # Resolve numeric product IDs to M-codes, then aggregate to avoid duplicates
        # (multiple numeric IDs can resolve to the same M-code)
        resolved_qty = defaultdict(float)
        resolved_amt = defaultdict(float)
        resolved_pack = {}
        for prod_id in totals_qty:
            item_code = prod_id
            if re.match(r'^\d+$', prod_id):
                # 0. product_map (ProductID → M-code, e.g. Kohl Wholesale)
                product_map = cfg.get('product_map', {})
                prod_id_stripped = prod_id.lstrip('0') or prod_id
                pm_key = prod_id_stripped if prod_id_stripped in product_map else prod_id
                if pm_key in product_map:
                    item_code = product_map[pm_key]
                    resolved_qty[item_code] += totals_qty[prod_id]
                    resolved_amt[item_code] += totals_amt[prod_id]
                    if item_code not in resolved_pack:
                        resolved_pack[item_code] = prod_pack.get(prod_id, '')
                    continue
                # 1. DID map (explicit, most accurate — used for Christ Panos etc.)
                did_val = prod_did.get(prod_id, '')
                if did_val and did_val in did_map:
                    item_code = did_map[did_val]
                else:
                    # 2. Description-based lookup (fallback for other numeric-ID distributors)
                    pack_size = prod_pack.get(prod_id, '')
                    flavor    = prod_desc.get(prod_id, '')
                    looked_up = lookup_item_code(flavor, pack_size) if flavor else None
                    if looked_up:
                        item_code = looked_up
            resolved_qty[item_code] += totals_qty[prod_id]
            resolved_amt[item_code] += totals_amt[prod_id]
            if item_code not in resolved_pack:
                resolved_pack[item_code] = prod_pack.get(prod_id, '')

        for item_code in resolved_qty:
            rows.append(make_row(
                source=source_name,
                program_num=cfg['program_num'],
                customer_ref=cref,
                dist_id=cfg['dist_id'],
                bill_date=bill_date,
                start_date=start_date,
                end_date=end_date,
                item=item_code,
                qty=round(resolved_qty[item_code], 4),
                amount=round(resolved_amt[item_code], 2),
                trade=cfg['trade']
            ))
        # Warn about processing fees (no M-code — can't go into Tellus automatically)
        for m in fee_pat.finditer(all_text):
            prod = m.group(1)
            amt  = float(m.group(2).replace(',', ''))
            rows.append({'_warning': True, 'code': prod, 'desc': 'Processing Fee (no M-code)', 'amount': amt})

    except Exception as e:
        rows.append({'_error': f'S&W PDF: {e}'})
    return rows


def parse_sw(filepath, cfg, customer_ref):
    """S&W Wholesale Foods — routes PDFs to parse_sw_pdf, Excel to legacy parser.
    Excel format header: Mfq. Product Code | Total case qty. | Incentive $$ ...
    Date range is extracted from the header area (Invoice Activity Date row).
    Amount per row = Total case qty × Incentive $$.
    Items with numeric Mfq. codes (no M-prefix) are flagged as warnings.
    """
    # Route PDFs to dedicated PDF parser
    if filepath.lower().endswith('.pdf'):
        return parse_sw_pdf(filepath, cfg, customer_ref)

    rows = []
    try:
        wb = __import__('openpyxl').load_workbook(filepath, data_only=True)
        ws = wb.active

        # Extract date range from header area (before data section)
        start_date = end_date = bill_date = TODAY
        inv_num = customer_ref
        for row in ws.iter_rows(max_row=30, values_only=True):
            for cell in row:
                if not cell or not isinstance(cell, str): continue
                m = re.search(r'(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})', cell)
                if m:
                    start_date = to_yyyymmdd(m.group(1))
                    end_date   = to_yyyymmdd(m.group(2))
                if 'Invoice Number' in cell or 'Invoice #' in cell:
                    pass  # handled by adjacent cell below
            # Check for invoice number in adjacent cells
            row_list = list(row)
            for j, cell in enumerate(row_list):
                if cell and isinstance(cell, str) and 'Invoice Number' in cell:
                    if j+1 < len(row_list) and row_list[j+1]:
                        inv_num = inv_num or str(row_list[j+1]).strip()

        cref = customer_ref or inv_num

        # Find column header row
        header_row = None
        col_map = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_vals = [str(v).strip().upper() if v else '' for v in row]
            if any('MFQ' in c and 'PRODUCT' in c for c in row_vals):
                header_row = i
                col_map = {v: j for j, v in enumerate(row_vals)}
                break
        if header_row is None:
            return [{'_error': 'S&W: column header not found'}]

        def gc(keys):
            for k in keys:
                for ck, ci in col_map.items():
                    if k in ck: return ci
            return None

        ci_item = gc(['MFQ. PRODUCT CODE', 'MFQ PRODUCT', 'MFG. PRODUCT'])
        ci_qty  = gc(['TOTAL CASE QTY', 'CASE QTY', 'QTY', 'CASES', 'QUANTITY'])
        ci_inc  = gc(['INCENTIVE'])          # Incentive $$ per case
        ci_desc = gc(['PRODUCT # : NAME', 'PRODUCT NAME', 'DESCRIPTION'])

        item_map = cfg.get('item_map', {})

        for row in ws.iter_rows(min_row=header_row+1, values_only=True):
            if not any(row): continue
            raw_item = str(row[ci_item]).strip() if ci_item is not None and row[ci_item] else ''
            if not raw_item or raw_item.upper() in ('NONE', ''): continue

            qty  = float(row[ci_qty] or 0) if ci_qty  is not None else 0
            inc  = clean_amount(str(row[ci_inc] or 0)) if ci_inc is not None else 0.0
            amt  = round(qty * inc, 4)

            desc = str(row[ci_desc]).strip() if ci_desc is not None and row[ci_desc] else raw_item

            if re.match(r'[A-Z]-[A-Z0-9]+', raw_item, re.I):
                rows.append(make_row(
                    source='S&W',
                    program_num=cfg['program_num'],
                    customer_ref=cref,
                    dist_id=cfg['dist_id'],
                    bill_date=bill_date,
                    start_date=start_date,
                    end_date=end_date,
                    item=raw_item.upper(),
                    qty=qty,
                    amount=amt,
                    trade=cfg['trade']
                ))
            else:
                mapped = item_map.get(raw_item.strip())
                if mapped and re.match(r'[A-Z]-[A-Z0-9]+', str(mapped), re.I):
                    rows.append(make_row(
                        source='S&W',
                        program_num=cfg['program_num'],
                        customer_ref=cref,
                        dist_id=cfg['dist_id'],
                        bill_date=bill_date,
                        start_date=start_date,
                        end_date=end_date,
                        item=mapped.upper(),
                        qty=qty,
                        amount=amt,
                        trade=cfg['trade']
                    ))
                else:
                    rows.append({'_warning': True, 'code': raw_item, 'desc': desc[:60], 'amount': amt})
    except Exception as e:
        rows.append({'_error': f'S&W: {e}'})
    return rows

def parse_kast_po_report(filepath, cfg, customer_ref):
    """KAST PO Receipts Report format.
    Each line: DATE  PO#  REQ#  PROD#  Description M-CODE  QTY  CS  WEIGHT  UNITCOST  CS  EXTCOST
    M-code is embedded in the description; amount = ExtCost * 4%.
    """
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            all_text = '\n'.join(page.extract_text() or '' for page in pdf.pages)

        # Bill date — first date in the header (report run date)
        bill_date = ''
        m = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', all_text)
        if m: bill_date = to_yyyymmdd(m.group(1))

        # Date range from "Receipt Date (MM-DD-YY - MM-DD-YY)"
        start_date = end_date = ''
        dr = re.search(r'Receipt\s+Date\s*\((\d{2}-\d{2}-\d{2})\s*-\s*(\d{2}-\d{2}-\d{2})\)', all_text, re.I)
        if dr:
            def yy_to_yyyy(s):
                # MM-DD-YY → YYYYMMDD
                parts = s.split('-')
                if len(parts) == 3:
                    mm, dd, yy = parts
                    yyyy = '20' + yy if int(yy) < 50 else '19' + yy
                    return f'{yyyy}{mm}{dd}'
                return ''
            start_date = yy_to_yyyy(dr.group(1))
            end_date   = yy_to_yyyy(dr.group(2))

        # Parse each line for M-code + qty + extended cost
        # Pattern: M-CODE  QTY  CS  WEIGHT  UNITCOST  CS  EXTCOST
        # After M-code there are always: int  CS  int  decimal  CS  decimal
        line_pat = re.compile(
            r'([A-Z]-[A-Z0-9]+)\s+(\d+)\s+CS\s+[\d,]+\s+([\d,.]+)\s+CS\s+([\d,.]+)',
            re.I
        )
        seen = set()
        for line in all_text.splitlines():
            # Skip totals / headers
            if re.search(r'Total For|Receipt Date|Product Description|Page \d', line, re.I):
                continue
            m = line_pat.search(line)
            if not m:
                continue
            item     = m.group(1).upper()
            qty      = int(m.group(2))
            ext_cost = clean_amount(m.group(4))
            amount   = round(ext_cost * 0.04, 4)
            rows.append(make_row(
                source='Kast',
                program_num=cfg['program_num'],
                customer_ref=customer_ref,
                dist_id=cfg['dist_id'],
                bill_date=bill_date,
                start_date=start_date,
                end_date=end_date,
                item=item,
                qty=qty,
                amount=amount,
                trade=cfg['trade']
            ))
    except Exception as e:
        rows.append({'_error': f'KAST PO Report: {e}'})
    return rows


def parse_kast(filepath, cfg, customer_ref):
    # Detect format: PO Receipts Report vs invoice
    try:
        with pdfplumber.open(filepath) as pdf:
            first_page = pdf.pages[0].extract_text() or ''
        if 'PO Receipts Report' in first_page or 'Receipt Date' in first_page:
            return parse_kast_po_report(filepath, cfg, customer_ref)
    except Exception:
        pass

    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            bill_date = start_date = end_date = ''
            for page in pdf.pages:
                text = page.extract_text() or ''
                if not bill_date:
                    m = re.search(r'(?:Invoice\s*Date|Date)[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text, re.I)
                    if m: bill_date = to_yyyymmdd(m.group(1))
                if not start_date:
                    s, e = parse_date_range(text)
                    if s: start_date, end_date = s, e

                # Table rows: M-code  desc  qty  ext_cost
                item_pat = re.compile(
                    r'([A-Z]-[A-Z0-9]+)\s+.{0,50?}\s+(\d+)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)',
                    re.I
                )
                for m in item_pat.finditer(text):
                    item     = m.group(1).upper()
                    qty      = m.group(2)
                    ext_cost = clean_amount(m.group(4))
                    amt      = round(ext_cost * 0.04, 4)
                    rows.append(make_row(
                        source='Kast',
                        program_num=cfg['program_num'],
                        customer_ref=customer_ref,
                        dist_id=cfg['dist_id'],
                        bill_date=bill_date,
                        start_date=start_date,
                        end_date=end_date,
                        item=item,
                        qty=qty,
                        amount=amt,
                        trade=cfg['trade']
                    ))

                # Table extraction fallback
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row: continue
                        cells = [str(c).strip() if c else '' for c in row]
                        item_cell = next((c for c in cells if re.match(r'[A-Z]-[A-Z0-9]+', c, re.I)), None)
                        if not item_cell: continue
                        nums = [clean_amount(c) for c in cells if re.match(r'[\d,]+\.?\d*$', c.replace(',',''))]
                        if len(nums) >= 3:
                            qty = int(nums[0]) if nums[0] > 0 else 1
                            ext_cost = nums[-1]
                            amt = round(ext_cost * 0.04, 4)
                            # avoid duplicates from text extraction
                            already = any(r.get('Item Number') == item_cell.upper() for r in rows[-10:])
                            if not already:
                                rows.append(make_row(
                                    source='Kast',
                                    program_num=cfg['program_num'],
                                    customer_ref=customer_ref,
                                    dist_id=cfg['dist_id'],
                                    bill_date=bill_date,
                                    start_date=start_date,
                                    end_date=end_date,
                                    item=item_cell.upper(),
                                    qty=qty,
                                    amount=amt,
                                    trade=cfg['trade']
                                ))
    except Exception as e:
        rows.append({'_error': str(e)})
    return rows

def parse_dot_foods_bb_xlsx(filepath, cfg, customer_ref):
    """DOT Foods BB format delivered as XLSX (used by SOFO and others).
    Structure: header rows with Invoice#/Date, then a data table with BB Dept / BB Vendor #
    columns. Monin rows identified by Vendor Number == 140469."""
    rows = []
    try:
        raw = pd.read_excel(filepath, sheet_name=0, header=None, nrows=15)

        # Extract invoice date and number from the top header rows
        bill_date = ''
        inv_num   = ''
        header_row_idx = None
        for i, row in raw.iterrows():
            vals = [str(v).strip() for v in row if str(v) not in ('nan', 'None', '')]
            joined = ' '.join(vals).lower()
            if 'invoice date' in joined:
                for v in vals:
                    d = to_yyyymmdd(v)
                    if d: bill_date = d; break
            if re.search(r'invoice\s*#', joined):
                for v in vals:
                    if re.match(r'\d{6,}$', v): inv_num = v; break
            if 'bb dept' in joined:
                header_row_idx = i
                break

        if header_row_idx is None:
            return [{'_error': 'DOT Foods BB XLSX: could not find BB Dept header row'}]

        df = pd.read_excel(filepath, sheet_name=0, header=header_row_idx)

        # Filter to Monin rows only (Vendor Number = 140469)
        df['Vendor Number'] = pd.to_numeric(df.get('Vendor Number', pd.Series(dtype=float)),
                                             errors='coerce')
        monin = df[df['Vendor Number'] == 140469].copy()
        if monin.empty:
            return [{'_error': 'DOT Foods BB XLSX: no Monin rows found (Vendor Number 140469)'}]

        # Date range from Invoice Date column
        dates = [to_yyyymmdd(v) for v in monin.get('Invoice Date', []) if to_yyyymmdd(v)]
        start_date = min(dates) if dates else bill_date
        end_date   = max(dates) if dates else bill_date

        if not customer_ref and inv_num:
            customer_ref = inv_num

        # Aggregate by Vendor Item# (M-code)
        from collections import defaultdict
        totals_qty = defaultdict(float)
        totals_amt = defaultdict(float)
        for _, row in monin.iterrows():
            code = str(row.get('Vendor Item#', '') or '').strip().upper()
            if not code or code == 'NAN': continue
            totals_qty[code] += float(row.get('Qty', 0) or 0)
            totals_amt[code] += float(row.get('Ded', 0) or 0)

        if not totals_amt:
            return [{'_error': 'DOT Foods BB XLSX: no product rows after filtering'}]

        for code, amt in totals_amt.items():
            rows.append(make_row(
                source='SOFO', program_num=cfg['program_num'],
                customer_ref=customer_ref, dist_id=cfg['dist_id'],
                bill_date=bill_date, start_date=start_date, end_date=end_date,
                item=code, qty=totals_qty[code], amount=amt, trade=cfg['trade'],
            ))
    except Exception as e:
        rows.append({'_error': f'DOT Foods BB XLSX: {e}'})
    return rows


def parse_sofo(filepath, cfg, customer_ref):
    """SOFO Foods — may arrive as DOT Foods BB format or Trackmax format; detect by content."""
    if filepath.lower().endswith('.pdf'):
        try:
            with pdfplumber.open(filepath) as pdf:
                first_page = pdf.pages[0].extract_text() or ''
        except Exception as e:
            return [{'_error': f'SOFO: could not read PDF: {e}'}]
        # SOFO files that come through DOT Foods use the BB Dept/BB Vendor format
        if 'BB Dept' in first_page and 'BB Vendor' in first_page:
            return parse_dot_foods_bb(filepath, cfg, customer_ref)
        # Trackmax-hosted SOFO files
        if 'Powered byTrackmax' in first_page or ('Product ID' in first_page and 'DID' in first_page):
            return parse_trackmax(filepath, cfg, customer_ref, source_name='SOFO')
    # XLSX/XLS — detect format by content
    if filepath.lower().endswith(('.xlsx', '.xls')):
        try:
            df_peek = pd.read_excel(filepath, header=None, nrows=15)
            flat = ' '.join(str(v) for row in df_peek.values for v in row if str(v) != 'nan')
            if 'BB Dept' in flat:
                return parse_dot_foods_bb_xlsx(filepath, cfg, customer_ref)
            if 'Cheney Invoice No' in flat:
                return parse_cheney(filepath, cfg, customer_ref)
        except Exception:
            pass
        return [{'_error': 'SOFO XLSX: unrecognised format — please share the file to add support.'}]

    # Native SOFO PDF format (direct M-code columns)
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            bill_date = start_date = end_date = ''
            for page in pdf.pages:
                text = page.extract_text() or ''
                if not bill_date:
                    m = re.search(r'(?:Date)[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text, re.I)
                    if m: bill_date = to_yyyymmdd(m.group(1))
                if not start_date:
                    s, e = parse_date_range(text)
                    if s: start_date, end_date = s, e
                # Conservative pattern: M-code then explicit qty (small integer) then amounts
                item_pat = re.compile(
                    r'([A-Z]-[A-Z0-9]+)\s+\S+\s+(\d{1,4})\s+\$?([\d,.]+)\s+\$?([\d,.]+)',
                    re.I
                )
                for m in item_pat.finditer(text):
                    item = m.group(1).upper()
                    qty  = m.group(2)
                    amt  = clean_amount(m.group(4))
                    rows.append(make_row(
                        source='SOFO',
                        program_num=cfg['program_num'],
                        customer_ref=customer_ref,
                        dist_id=cfg['dist_id'],
                        bill_date=bill_date,
                        start_date=start_date,
                        end_date=end_date,
                        item=item,
                        qty=qty,
                        amount=amt,
                        trade=cfg['trade']
                    ))
    except Exception as e:
        rows.append({'_error': str(e)})
    return rows

def _pfs_normalize_mcode(raw):
    """Normalize a PFS M-code: add missing dash (MRP036F → M-RP036F), uppercase."""
    code = raw.strip().upper()
    if re.match(r'^[A-Z]-[A-Z]', code):
        return code  # already has dash
    # Bare code like MRP036F, KKfr079F etc — insert dash after first letter
    if re.match(r'^[A-Z]{2,3}\d{3}', code):
        return code[0] + '-' + code[1:]
    return code

def parse_pfs_bid_pdf(filepath, cfg, customer_ref):
    """PFS 'Bid Bill-Back Report' (BIBBRPT10R) PDF format.
    Each Bid section covers one operator. Emits one row per operator × M-code
    with Trade=O and Operator ID set to the Bid description (e.g. 'ZAXBY POWELL').
    Detail line format: <6-digit vendor#> <desc> <13/14-digit UPC> <M-code> <ContractId> <Qty> <Each> <Allow> <Extended>
    """
    rows = []
    try:
        all_text = ''
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                all_text += (page.extract_text() or '') + '\n'

        # Date range: "From: M/D/YY" and "Thru: M/D/YY" on the header line
        dr = re.search(r'From:\s+(\d{1,2}/\d{1,2}/\d{2,4}).*?Thru:\s+(\d{1,2}/\d{1,2}/\d{2,4})', all_text, re.I | re.S)
        start_date = to_yyyymmdd(dr.group(1)) if dr else ''
        end_date   = to_yyyymmdd(dr.group(2)) if dr else ''
        bill_date  = end_date or start_date

        # Customer ref from filename: Inv_0123003_... → 0123003
        inv_m = re.search(r'Inv[_-](\d+)', os.path.basename(filepath), re.I)
        cust_ref = inv_m.group(1) if inv_m else (customer_ref or '')

        mcode_pat = re.compile(r'\b([A-Z]-?[A-Z]{1,2}\d{3,4}[A-Z0-9]*)\b', re.I)
        # Bid section header: "Bid...: 3 !! ZAXBY POWELL" or "Bid...: 148 P TERRY'S MAIN BID"
        bid_hdr   = re.compile(r'Bid\.\.\.\s*:\s*\d+\s+(?:!!\s+)?(.+)', re.I)
        skip_pat  = re.compile(r'Totals|Bid\.\.\.|Vendor\s+\d+\s+Totals|END\s+OF\s+REPORT', re.I)

        # bid_data: {operator_name: {mcode: [qty, amt]}}
        from collections import defaultdict, OrderedDict
        bid_data = OrderedDict()
        current_op = ''

        for line in all_text.splitlines():
            bm = bid_hdr.match(line.strip())
            if bm:
                current_op = bm.group(1).strip()
                if current_op not in bid_data:
                    bid_data[current_op] = {}
                continue
            if skip_pat.search(line): continue
            # Data lines start with a 6-digit vendor item number
            if not re.match(r'^\s*\d{6}\s+', line): continue
            # Must contain a 13-14 digit UPC
            upc_m = re.search(r'\b\d{13,14}\b', line)
            if not upc_m: continue
            # M-code immediately after UPC
            after_upc = line[upc_m.end():]
            mc = mcode_pat.search(after_upc)
            if not mc: continue
            mcode = _pfs_normalize_mcode(mc.group(1))
            # Last 4 numbers on the line are: Qty, Each, Allow, Extended
            after_mcode = after_upc[mc.end():]
            nums = re.findall(r'[\d,.]+', after_mcode)
            if len(nums) < 4: continue
            try:
                qty      = float(nums[-4].replace(',', ''))
                extended = float(nums[-1].replace(',', ''))
            except Exception:
                continue
            if extended == 0: continue  # skip $0.00 rows
            if not current_op:
                current_op = 'Unknown'
                bid_data[current_op] = {}
            d = bid_data[current_op]
            if mcode not in d:
                d[mcode] = [0.0, 0.0]
            d[mcode][0] += qty
            d[mcode][1] += extended

        for operator, mcodes in bid_data.items():
            for mcode in sorted(mcodes):
                qty, amt = mcodes[mcode]
                rows.append(make_row(
                    source='PFS',
                    program_num=cfg['program_num'],
                    customer_ref=cust_ref,
                    dist_id=cfg['dist_id'],
                    bill_date=bill_date, start_date=start_date, end_date=end_date,
                    item=mcode, qty=qty, amount=amt,
                    trade='O',
                    operator_id=operator
                ))
        if not rows:
            rows.append({'_error': 'PFS Bid Bill-Back PDF: no Monin items found'})
    except Exception as e:
        rows.append({'_error': f'PFS Bid Bill-Back PDF: {e}\n{traceback.format_exc()}'})
    return rows

def parse_pfs(filepath, cfg, customer_ref):
    """PFS Roma / PFG — uses Trackmax format (same engine as Driscoll Foods)."""
    if filepath.lower().endswith('.pdf'):
        # Check for PFS Bid Bill-Back Report (BIBBRPT10R) format first
        first_page, err = _pdfplumber_extract_first_page(filepath)
        if err is None and ('Bid Bill-Back Report' in first_page or 'BIBBRPT' in first_page):
            return parse_pfs_bid_pdf(filepath, cfg, customer_ref)
        # PFS Roma uses Powered-by-Trackmax format; route through the generic Trackmax parser.
        return parse_trackmax(filepath, cfg, customer_ref, source_name='PFS')
    # Non-PDF (CSV/XLSX) — PFS BidRRpt format
    # Columns: Distribution Center | Vendor# | Vendor Name | Date From | Date Thru |
    #          Total Due | Bid | Bid Description | Item | Pack | Size | Description |
    #          UPC | Vendor Item# | Contract ID | Qty | Each | Allowance | Extended Amt
    rows = []
    try:
        import pandas as pd
        if filepath.lower().endswith('.csv'):
            df = None
            for enc in ('utf-8', 'latin-1', 'cp1252'):
                try:
                    df = pd.read_csv(filepath, dtype=str, encoding=enc)
                    break
                except UnicodeDecodeError:
                    continue
            if df is None:
                rows.append({'_error': 'PFS CSV: could not decode file (tried utf-8, latin-1, cp1252)'})
                return rows
        else:
            df = pd.read_excel(filepath, dtype=str)
        df.columns = [c.strip() for c in df.columns]

        # Locate key columns (case-insensitive, strip whitespace)
        col_map = {c.strip().lower(): c for c in df.columns}
        def col(name): return col_map.get(name.lower())

        mcode_col  = col('vendor item#') or col('item no') or col('item number')
        qty_col    = col('qty') or col('quantity')
        amt_col    = col('extended amt') or col('extended amount') or col('extended')
        date_from  = col('date from') or col('from')
        date_thru  = col('date thru') or col('thru') or col('date through')
        # Bid Description column holds the operator/chain name per row
        bid_desc_col = col('bid description') or col('operator') or col('bid desc') or col('chain')

        if not mcode_col or not qty_col or not amt_col:
            rows.append({'_error': f'PFS CSV: expected columns Vendor Item#, Qty, Extended Amt — found: {list(df.columns)}'})
            return rows

        # Dates from first data row
        start_date = end_date = bill_date = ''
        if date_from and not df[date_from].dropna().empty:
            start_date = to_yyyymmdd(df[date_from].dropna().iloc[0].strip())
        if date_thru and not df[date_thru].dropna().empty:
            end_date = to_yyyymmdd(df[date_thru].dropna().iloc[0].strip())
        bill_date = end_date or start_date

        # bid_data: {operator_name: {mcode: [qty, amt]}}
        # If no Bid Description column, use a single key '' (aggregate all)
        from collections import OrderedDict
        bid_data = OrderedDict()

        for _, r in df.iterrows():
            raw_code = str(r.get(mcode_col, '') or '').strip()
            if not raw_code or raw_code.lower() == 'nan': continue
            # Accept M-codes: already-dashed form (M-RPO40F) or no-dash form (MRP036F)
            # Use a broad check: [A-Z]-[A-Z0-9]{3,} OR [A-Z]{2,3}\d{3}
            if not (re.match(r'^[A-Z]-[A-Z0-9]{3,}', raw_code, re.I) or
                    re.match(r'^[A-Z]{2,3}\d{3}', raw_code, re.I)):
                continue
            item = _pfs_normalize_mcode(raw_code)
            try: qty = float(str(r.get(qty_col, 0) or 0).replace(',', ''))
            except: qty = 0.0
            try: amt = float(str(r.get(amt_col, 0) or 0).replace(',', '').strip())
            except: amt = 0.0
            if amt == 0: continue

            operator = ''
            if bid_desc_col:
                raw_op = str(r.get(bid_desc_col, '') or '').strip()
                if raw_op.lower() != 'nan':
                    # Strip leading !! / ¢¢ / other non-word prefixes (e.g. "!! ZAXBY" → "ZAXBY")
                    operator = re.sub(r'^[^A-Z0-9\'(]+', '', raw_op, flags=re.I).strip()

            if operator not in bid_data:
                bid_data[operator] = {}
            d = bid_data[operator]
            if item not in d:
                d[item] = [0.0, 0.0]
            d[item][0] += qty
            d[item][1] += amt

        has_operators = bid_desc_col and any(k for k in bid_data)
        for operator, mcodes in bid_data.items():
            for item in sorted(mcodes):
                qty, amt = mcodes[item]
                rows.append(make_row(
                    source='PFS', program_num=cfg['program_num'],
                    customer_ref=customer_ref, dist_id=cfg['dist_id'],
                    bill_date=bill_date, start_date=start_date, end_date=end_date,
                    item=item, qty=qty, amount=amt,
                    trade='O' if has_operators else cfg['trade'],
                    operator_id=operator if has_operators else ''
                ))
        if not rows:
            rows.append({'_error': 'PFS CSV: no M-code rows found in Vendor Item# column'})
    except Exception as e:
        rows.append({'_error': f'PFS CSV: {e}'})
    return rows

def parse_yhata(filepath, cfg, customer_ref):
    # Y. Hata uses Trackmax-format PDFs — route through the content-based dispatcher
    if filepath.lower().endswith('.pdf'):
        return parse_supplier_billback_pdf(filepath, cfg, customer_ref)
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            bill_date = start_date = end_date = ''
            for page in pdf.pages:
                text = page.extract_text() or ''
                if not bill_date:
                    m = re.search(r'(?:Date|Dated)[:\s]+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text, re.I)
                    if m: bill_date = to_yyyymmdd(m.group(1))
                if not start_date:
                    s, e = parse_date_range(text)
                    if s: start_date, end_date = s, e

                # Pattern: M-code  desc  UPC(8-20 digits)  qty  price  ext  allowance
                detail_pat = re.compile(
                    r'([A-Z]-[A-Z0-9]+)\s+.{0,60}?\s+\d{8,20}\s+(\d+)\s+\$?([\d,.]+)\s+\$?([\d,.]+)\s+\(?\$?([\d,.]+)\)?',
                    re.I
                )
                detail_pat2 = re.compile(
                    r'([A-Z]-[A-Z0-9]+)\s+.{0,60}?\s+\d+\s+(\d+)\s+\$?([\d,.]+)\s+\$?([\d,.]+)\s+\(?\$?([\d,.]+)\)?',
                    re.I
                )
                matched = set()
                for pat in [detail_pat, detail_pat2]:
                    for m in pat.finditer(text):
                        item = m.group(1).upper()
                        if item in matched: continue
                        matched.add(item)
                        qty  = m.group(2)
                        amt  = clean_amount(m.group(5))
                        rows.append(make_row(
                            source='Y.Hata',
                            program_num=cfg['program_num'],
                            customer_ref=customer_ref,
                            dist_id=cfg['dist_id'],
                            bill_date=bill_date,
                            start_date=start_date,
                            end_date=end_date,
                            item=item,
                            qty=qty,
                            amount=amt,
                            trade=cfg['trade']
                        ))
    except Exception as e:
        rows.append({'_error': str(e)})
    return rows

def parse_labatt(filepath, cfg, customer_ref):
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            bill_date = start_date = end_date = ''
            for page in pdf.pages:
                text = page.extract_text() or ''
                if not bill_date:
                    m = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})', text)
                    if m: bill_date = to_yyyymmdd(m.group(1))
                if not start_date:
                    s, e = parse_date_range(text)
                    if s: start_date, end_date = s, e

                item_pat = re.compile(
                    r'([A-Z]-[A-Z0-9]+)\s+\S+\s+(\d+)\s+\$?([\d,.]+)\s+\$?([\d,.]+)',
                    re.I
                )
                for m in item_pat.finditer(text):
                    item = m.group(1).upper()
                    qty  = m.group(2)
                    amt  = clean_amount(m.group(4))
                    rows.append(make_row(
                        source='Labatt',
                        program_num=cfg['program_num'],
                        customer_ref=customer_ref,
                        dist_id=cfg['dist_id'],
                        bill_date=bill_date,
                        start_date=start_date,
                        end_date=end_date,
                        item=item,
                        qty=qty,
                        amount=amt,
                        trade=cfg['trade']
                    ))
    except Exception as e:
        rows.append({'_error': str(e)})
    return rows

def parse_trackmax(filepath, cfg, customer_ref, source_name=''):
    """Generic Trackmax 'Supplier Billback' PDF parser (Driscoll Foods, PFS Roma, etc.).
    Columns: Inv.Number Inv.Date CustomerID CustomerName Brand PackSize Description
             ProductID(M-code) DID UPC Quantity TotalWeight TotalCharges ProgramAmount AmountDue
    Credit/return lines have parenthesized quantities and amounts, e.g. (1.00) and ($9.04).
    These must be captured as NEGATIVE amounts to match the Grand Total.
    source_name: the row Source label; auto-detected from PDF header if empty.
    """
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            all_text = '\n'.join(page.extract_text() or '' for page in pdf.pages)

        # Auto-detect distributor name from header ("IMPORTANT!\n<Company Name>")
        if not source_name:
            m = re.search(r'IMPORTANT!\s*\n([^\n]+)', all_text)
            source_name = m.group(1).strip() if m else 'Trackmax Dist'

        # Statement date and program date range
        bill_date = start_date = end_date = ''
        m = re.search(r'(?:generated|posted)\s+on\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: bill_date = to_yyyymmdd(m.group(1))
        m = re.search(r'between\s+(\d{1,2}/\d{1,2}/\d{4})\s+and\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m:
            start_date = to_yyyymmdd(m.group(1))
            end_date   = to_yyyymmdd(m.group(2))
        # Also check "Start Date: / Stop Date:" header format
        if not start_date:
            m2 = re.search(r'Start\s+Date[:\s]+(\d{1,2}/\d{1,2}/\d{4}).*?Stop\s+Date[:\s]+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I | re.S)
            if m2:
                start_date = to_yyyymmdd(m2.group(1))
                end_date   = to_yyyymmdd(m2.group(2))

        # Invoice number for customer ref
        inv_num = ''
        m2 = re.search(r'Our\s+Invoice\s+Number[:\s]+(\w+)', all_text, re.I)
        if m2: inv_num = m2.group(1).strip()
        cref = customer_ref or inv_num

        # Data line pattern — matches both regular and credit/return lines:
        #   Regular:  019926 02/13/2026 ... M-FR035F 350377 10738337884136 1.00  11.53 $31.44 BB To 22.400/unit $9.04
        #   Credit:   X97707 02/20/2026 ... M-FR035F 350377 10738337884136 (1.00) (11.53) ($31.44) BB To 22.400/unit ($9.04)
        #   Fully Excluded (Y.Hata): ...M-FR036F 256271 0110738337060868 (1.00) (12.50) ($31.44) ($31.44) -10.00 % of FOB ($3.14)
        # Also handles: UPCs up to 18 digits; M-codes with hyphen dropped by pdfplumber (e.g. MFR066F)
        line_pat = re.compile(
            r'([A-Z]-?[A-Z][A-Z0-9]+)\s+\d{4,9}\s+(?:\d{8,18}\s+)?'  # M-code (hyphen optional), DID, UPC optional (up to 18d)
            r'(\([\d.]+\)|[\d.]+)\s+'                    # qty: positive or (negative)
            r'[\d,.()]+\s+'                              # weight (ignore, may have comma e.g. 1,268.40)
            r'(?:\(\$[\d,.]+\)|\$[-\d,.]+)\s+'          # Total Charges 1 (ignore, may be negative $-X.XX)
            r'(?:(?:\(\$[\d,.]+\)|\$[-\d,.]+)\s+)?'     # Total Charges 2 (optional — Fully Excluded Sales format)
            r'\S.*?'                                     # Program Amount text
            r'(\(\$[\d,.]+\)|\$[-\d,.]+)',               # Amount Due (last $ value)
            re.I
        )

        seen_lines = set()
        for line in all_text.splitlines():
            if re.search(r'^Totals for|^Grand Total|Powered by|Invoice is Due', line, re.I):
                continue
            m = line_pat.search(line)
            if not m:
                continue
            if line in seen_lines:
                continue
            seen_lines.add(line)

            item = m.group(1).upper()
            # Restore missing hyphen that pdfplumber sometimes drops (e.g. MFR066F → M-FR066F)
            if len(item) > 1 and item[1] != '-':
                item = item[0] + '-' + item[1:]
            raw_qty = m.group(2)
            raw_amt = m.group(3)

            # Parenthesized qty means a return (negative)
            if raw_qty.startswith('('):
                qty = -float(raw_qty.strip('()'))
            else:
                qty = float(raw_qty)

            amount = clean_amount(raw_amt)  # handles ($9.04) and $-2.80 → negative

            rows.append(make_row(
                source=source_name,
                program_num=cfg['program_num'],
                customer_ref=cref,
                dist_id=cfg['dist_id'],
                bill_date=bill_date,
                start_date=start_date,
                end_date=end_date,
                item=item,
                qty=qty,
                amount=amount,
                trade=cfg['trade']
            ))

        if not rows:
            rows.append({'_error': f'{source_name} PDF: no M-code rows found'})
    except Exception as e:
        rows.append({'_error': f'Trackmax parser: {e}'})
    return rows


def parse_driscoll(filepath, cfg, customer_ref):
    """Driscoll Foods wrapper — uses generic Trackmax parser."""
    return parse_trackmax(filepath, cfg, customer_ref, source_name='Driscoll Foods')


def parse_henrys_foods(filepath, cfg, customer_ref):
    """Henry's Foods Trackmax PDF — uses '1-NNN ML' pack size format and 5-6 digit numeric product IDs.
    Columns: InvNum PO# InvDate RcvdDate SupplierName Brand PackSize Description ProductID DID UPC Qty Weight FOB DEL Rate AmtDue
    Resolves numeric product IDs to M-codes via did_map in supplier config.
    """
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            all_text = '\n'.join(page.extract_text() or '' for page in pdf.pages)

        # Dates
        bill_date = start_date = end_date = ''
        m = re.search(r'(?:generated|posted)\s+on\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: bill_date = to_yyyymmdd(m.group(1))
        m = re.search(r'between\s+(\d{1,2}/\d{1,2}/\d{4})\s+and\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: start_date, end_date = to_yyyymmdd(m.group(1)), to_yyyymmdd(m.group(2))
        if not start_date:
            m2 = re.search(r'Start\s+Date[:\s]+(\d{1,2}/\d{1,2}/\d{4}).*?Stop\s+Date[:\s]+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I | re.S)
            if m2: start_date, end_date = to_yyyymmdd(m2.group(1)), to_yyyymmdd(m2.group(2))

        # Invoice ref and source name
        inv_num = ''
        m = re.search(r'Invoice\s+Number[:\s]+(\d+)', all_text, re.I)
        if m: inv_num = m.group(1).strip()
        cref = customer_ref or inv_num

        source_name = "Henry's Foods"
        m2 = re.search(r'IMPORTANT!\s*\n([^\n]+)', all_text)
        if m2: source_name = m2.group(1).strip()

        # Line pattern for Henry's Foods '1-NNN ML' pack size format.
        # Format: MONIN 1-750 ML CARAMEL SYRUP 399109 1454826 738337006814 24.00 81.50 $142.25 $142.24 1.00 % of Del $1.42
        line_pat = re.compile(
            r'MONIN\s+(\d+-\S+(?:\s+(?:ML|LT|CT|RACK|OZ|EA))?)\s+'  # PackSize: 1-750 ML, 1-RACK, 1-1 CT
            r'(?:\S+\s+)*?'                     # Description words (non-greedy)
            r'(\d{5,6})\s+'                     # Product ID (5-6 digit numeric)
            r'(\d{3,9})\s+'                     # DID
            r'(?:\d{6,18}\s+)?'                 # UPC optional (6-18 digits)
            r'(\d[\d,]*\.?\d*)\s+'              # Quantity
            r'[\d,.]+\s+'                       # Total Weight
            r'(?:\$[\d,.]+\s+){0,2}'            # 0-2 pre-amounts (FOB + DEL)
            r'[\d.]+\s*%\s+of\s+Del(?:Charges)?\s+'  # Rate: X.XX % of Del
            r'\$([\d,.]+)',                     # Amount Due
            re.I
        )

        did_map = cfg.get('did_map', {})
        from collections import defaultdict
        totals_qty = defaultdict(float)
        totals_amt = defaultdict(float)
        prod_pack  = {}
        prod_did   = {}
        seen_matches = set()

        for m in line_pat.finditer(all_text):
            line_start = all_text.rfind('\n', 0, m.start()) + 1
            match_key  = all_text[line_start:m.end()][:200]
            if match_key in seen_matches:
                continue
            seen_matches.add(match_key)
            pack_size = m.group(1).strip()
            prod_id   = m.group(2)
            did_val   = m.group(3)
            qty       = float(m.group(4).replace(',', ''))
            amt       = float(m.group(5).replace(',', ''))
            totals_qty[prod_id] += qty
            totals_amt[prod_id] += amt
            if prod_id not in prod_pack:
                prod_pack[prod_id] = pack_size
            if prod_id not in prod_did:
                prod_did[prod_id] = did_val

        # Resolve numeric product IDs → M-codes via DID map
        resolved_qty  = defaultdict(float)
        resolved_amt  = defaultdict(float)
        resolved_pack = {}
        for prod_id in totals_qty:
            item_code = prod_id
            did_val   = prod_did.get(prod_id, '')
            if did_val and did_val in did_map:
                item_code = did_map[did_val]
            resolved_qty[item_code]  += totals_qty[prod_id]
            resolved_amt[item_code]  += totals_amt[prod_id]
            if item_code not in resolved_pack:
                resolved_pack[item_code] = prod_pack.get(prod_id, '')

        for item_code in resolved_qty:
            rows.append(make_row(
                source=source_name,
                program_num=cfg['program_num'],
                customer_ref=cref,
                dist_id=cfg['dist_id'],
                bill_date=bill_date,
                start_date=start_date,
                end_date=end_date,
                item=item_code,
                qty=resolved_qty[item_code],
                amount=resolved_amt[item_code],
                trade=cfg['trade']
            ))

        if not rows:
            rows.append({'_error': "Henry's Foods PDF: no product rows found"})
    except Exception as e:
        rows.append({'_error': f"Henry's Foods parser: {e}"})
    return rows


def parse_delco_foods(filepath, cfg, customer_ref):
    """Delco Foods 'Supplier Billback' Trackmax PDF.
    Columns: InvNum PO# InvDate Rcvd SupplierName Brand PackSize Description ProductID DID Qty Total FOB DEL Rate Amount
    M-codes appear directly as Product ID; DID format is XXXXXXXXCAS (10 alphanumeric chars).
    Rate: 4.00 % of FOB with two identical dollar columns (FOB + DEL).
    """
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            all_text = '\n'.join(page.extract_text() or '' for page in pdf.pages)

        # Statement / bill date
        bill_date = start_date = end_date = ''
        m = re.search(r'generated\s+on\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: bill_date = to_yyyymmdd(m.group(1))

        # Program date range
        m = re.search(r'between\s+(\d{1,2}/\d{1,2}/\d{4})\s+and\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: start_date, end_date = to_yyyymmdd(m.group(1)), to_yyyymmdd(m.group(2))
        if not start_date:
            m2 = re.search(r'Start\s+Date[:\s]+(\d{1,2}/\d{1,2}/\d{4}).*?Stop\s+Date[:\s]+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I | re.S)
            if m2: start_date, end_date = to_yyyymmdd(m2.group(1)), to_yyyymmdd(m2.group(2))

        # Invoice number
        inv_num = ''
        m = re.search(r'Our\s+Invoice\s+Number[:\s]+(\w+)', all_text, re.I)
        if not m:
            m = re.search(r'Invoice\s+Number[:\s]+(\w+)', all_text, re.I)
        if m: inv_num = m.group(1).strip()
        cref = customer_ref or inv_num

        source_name = 'Delco Foods'

        # Line pattern:
        # MONIN PackSize Description M-code DID Qty Weight $FOB $DEL Rate% $Amount
        # PackSize examples: 12/750  4/1.89  6/12 OZ  4/1 LTR  4/1 L  1/each  6/12OZ  12/750 ml
        # DID examples: 1002081CAS  1004006CAS  1002428EAC  (10 alphanumeric chars)
        line_pat = re.compile(
            r'MONIN\s+'
            r'\S+\s+'                                    # PackSize main token (unit may be on next line)
            r'.+?\s+'                                    # Description words (lazy, no dotall — won't cross newlines)
            r'(M-[A-Z][A-Z0-9]+|P\d{3,4})\s+'          # Product ID (M-code like M-AR045A or P-code like P240)
            r'\w{6,12}\s+'                              # DID (e.g., 1002081CAS)
            r'(\d[\d,]*\.?\d*)\s+'                      # Quantity
            r'[\d,.]+\s+'                               # Total weight
            r'\$[\d,]+\.?\d*\s+'                        # FOB amount
            r'\$[\d,]+\.?\d*\s+'                        # DEL amount
            r'[\d.]+\s*%\s+of\s+FOB\s+'                # Rate (4.00 % of FOB)
            r'\$([\d,]+\.?\d*)',                        # Amount Due
            re.I
        )

        from collections import defaultdict
        totals_qty = defaultdict(float)
        totals_amt = defaultdict(float)
        seen_matches = set()

        for m in line_pat.finditer(all_text):
            # Deduplicate lines that appear on page-boundary continuations
            line_start = all_text.rfind('\n', 0, m.start()) + 1
            match_key  = all_text[line_start:m.end()][:200]
            if match_key in seen_matches:
                continue
            seen_matches.add(match_key)

            item = m.group(1).upper()
            # Fix OCR artifact: pdfplumber sometimes reads '0' as 'O' in the 3-digit
            # numeric section of M-codes (e.g., M-ARO10A → M-AR010A, M-ARO42A → M-AR042A).
            # Pattern M-XX\d\d\dY — only replace O→0 inside the 3-char numeric block.
            ocr_fix = re.match(r'(M-[A-Z]{2})([A-Z0-9]{3})([A-Z].*)$', item)
            if ocr_fix:
                item = ocr_fix.group(1) + ocr_fix.group(2).replace('O', '0') + ocr_fix.group(3)

            qty = float(m.group(2).replace(',', ''))
            amt = float(m.group(3).replace(',', ''))
            totals_qty[item] += qty
            totals_amt[item] += amt

        for item in sorted(totals_qty):
            rows.append(make_row(
                source=source_name,
                program_num=cfg['program_num'],
                customer_ref=cref,
                dist_id=cfg['dist_id'],
                bill_date=bill_date,
                start_date=start_date,
                end_date=end_date,
                item=item,
                qty=totals_qty[item],
                amount=totals_amt[item],
                trade=cfg['trade']
            ))

        if not rows:
            rows.append({'_error': 'Delco Foods PDF: no product rows found'})
    except Exception as e:
        rows.append({'_error': f'Delco Foods parser: {e}'})
    return rows


def parse_chefs_warehouse(filepath, cfg, customer_ref):
    """The Chefs Warehouse / Dairyland 'Supplier Billback' Trackmax PDF.
    M-codes or numeric/alphanumeric product IDs; DID is separate or same as ProductID.
    Rate: 5.00 % of Del with two identical dollar columns (FOB + DEL).
    Two-pass approach: primary for M-codes/P-codes/numeric IDs, secondary for ProductID=DID cases
    (e.g. MONIN925 MONIN925 or 921023N 921023N where product code repeats as the DID).
    """
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            all_text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
        bill_date = start_date = end_date = ''
        m = re.search(r'generated\s+on\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: bill_date = to_yyyymmdd(m.group(1))
        m = re.search(r'between\s+(\d{1,2}/\d{1,2}/\d{4})\s+and\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: start_date, end_date = to_yyyymmdd(m.group(1)), to_yyyymmdd(m.group(2))
        if not start_date:
            m2 = re.search(r'Start\s+Date[:\s]+(\d{1,2}/\d{1,2}/\d{4}).*?Stop\s+Date[:\s]+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I | re.S)
            if m2: start_date, end_date = to_yyyymmdd(m2.group(1)), to_yyyymmdd(m2.group(2))
        inv_num = ''
        m = re.search(r'Our\s+Invoice\s+Number[:\s]+(\w+)', all_text, re.I)
        if not m:
            m = re.search(r'Invoice\s+Number[:\s]+(\w+)', all_text, re.I)
        if m: inv_num = m.group(1).strip()
        cref = customer_ref or inv_num
        source_name = 'The Chefs Warehouse'
        # Primary pattern: handles M-codes (M-AR056A), bare alphanumeric M-codes (MAR147A),
        # P-codes, and pure numeric IDs (10099631).  DID is optional (some lines omit it).
        # An optional parenthetical token e.g. (01) may appear between DID and qty.
        # Uses [ \t]+ (not \s+) throughout to prevent cross-line matching.
        primary_pat = re.compile(
            r'MONIN[ \t]+'
            r'\S+[ \t]+'                           # pack size (e.g. 12/750)
            r'.+?[ \t]+'                           # description words (lazy, no newline cross)
            r'(M-[A-Z][A-Z0-9]+|[A-Z]{2,3}\d{3}[A-Z][A-Z0-9]*|P\d{3,4}|\d{5,10})[ \t]+'  # ProductID
            r'(?:\S+[ \t]+)?'                      # DID (optional — some lines omit it)
            r'(?:\([^)]*\)[ \t]+)?'                # optional parenthetical e.g. (01)
            r'(?:\d{10,16}[ \t]+)?'                # UPC (optional)
            r'(\d[\d,]*\.?\d*)[ \t]+'              # Qty
            r'[\d,.]+[ \t]+'                       # Weight
            r'[$][\d,.]+[ \t]+'                    # FOB $
            r'[$][\d,.]+[ \t]+'                    # DEL $
            r'[\d.]+[ \t]*[%][ \t]+of[ \t]+Del[ \t]+'  # Rate
            r'[$]([\d,.]+)',                        # Amount
            re.I
        )
        # Secondary pattern: ProductID = DID (e.g. MONIN925 MONIN925, 921023N 921023N)
        # Uses backreference \1 to ensure the two tokens are identical.
        # Uses [ \t]+ (not \s+) throughout to prevent cross-line matching.
        secondary_pat = re.compile(
            r'MONIN[ \t]+'
            r'\S+[ \t]+'                           # pack size
            r'(?:\S+[ \t]+)*?'                     # description words (lazy, no newline cross)
            r'([A-Z0-9]{5,12})[ \t]+'              # ProductID (alphanumeric, 5-12 chars)
            r'\1[ \t]+'                             # DID must equal ProductID (backreference)
            r'(?:\d{10,16}[ \t]+)?'                # UPC (optional)
            r'(\d[\d,]*\.?\d*)[ \t]+'              # Qty
            r'[\d,]+\.?\d*[ \t]+'                  # Weight
            r'[$][\d,.]+[ \t]+'                    # FOB $
            r'[$][\d,.]+[ \t]+'                    # DEL $
            r'[\d.]+[ \t]*[%][ \t]+of[ \t]+Del[ \t]+'  # Rate
            r'[$]([\d,.]+)',                        # Amount
            re.I
        )
        from collections import defaultdict
        totals_qty = defaultdict(float)
        totals_amt = defaultdict(float)
        seen_line_starts = set()  # dedup: one match per source line across both passes

        def _add(match_obj):
            ls = all_text.rfind('\n', 0, match_obj.start()) + 1
            if ls in seen_line_starts:
                return
            seen_line_starts.add(ls)
            item = match_obj.group(1).upper()
            qty = float(match_obj.group(2).replace(',', ''))
            amt = float(match_obj.group(3).replace(',', ''))
            totals_qty[item] += qty
            totals_amt[item] += amt

        for m in primary_pat.finditer(all_text):
            _add(m)
        for m in secondary_pat.finditer(all_text):
            _add(m)

        for item in sorted(totals_qty):
            rows.append(make_row(
                source=source_name, program_num=cfg['program_num'],
                customer_ref=cref, dist_id=cfg['dist_id'],
                bill_date=bill_date, start_date=start_date, end_date=end_date,
                item=item, qty=totals_qty[item], amount=totals_amt[item], trade=cfg['trade']
            ))
        if not rows:
            rows.append({'_error': 'Chefs Warehouse PDF: no product rows found'})
    except Exception as e:
        rows.append({'_error': f'Chefs Warehouse parser: {e}'})
    return rows


def parse_martin_bros(filepath, cfg, customer_ref):
    """Martin Bros. Dist. Co. 'Supplier Billback' PDF format.
    Columns: InvNum [PO#] InvDate [Rcvd] Supplier Brand PackSize Desc M-CODE DID UPC Qty
             Weight $Charges [$DEL] Rate $Amount
    M-code is Product ID; billback amount is last $X.XX on the line.
    Rate is either a percentage (e.g. "6.00% of Del Cost") or per-unit (e.g. "BB To 41.410/unit").
    Page-boundary dedup: Martin Bros PDFs sometimes repeat the last line of page N at the top
    of page N+1. Detect this by processing pages individually and skipping the first data line
    of a page when it exactly matches the last data line of the previous page.
    Legitimate within-page duplicates (same item invoiced twice on one page) are preserved.
    """
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            pages_text = [page.extract_text() or '' for page in pdf.pages]
        all_text = '\n'.join(pages_text)

        # Statement date and program date range
        bill_date = start_date = end_date = ''
        m = re.search(r'generated on\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: bill_date = to_yyyymmdd(m.group(1))
        m = re.search(r'between\s+(\d{1,2}/\d{1,2}/\d{4})\s+and\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m:
            start_date = to_yyyymmdd(m.group(1))
            end_date   = to_yyyymmdd(m.group(2))

        # Invoice number for customer ref
        inv_num = ''
        m2 = re.search(r'Invoice\s+Number[:\s]+(\d+)', all_text, re.I)
        if not m2:
            m2 = re.search(r'Our\s+Invoice\s+Number[:\s\n]+(\d+)', all_text, re.I)
        if m2: inv_num = m2.group(1).strip()
        cref = customer_ref or inv_num

        # Each data line: ... M-CODE  DID(5-9d)  UPC(10-14d)  QTY  Weight  $Charges  [$DEL]  Rate  $Amount
        # Rate is either a percentage (e.g. "6.00%") or a per-unit amount (e.g. "BB To 41.410/unit")
        # The second dollar column (DEL) is optional depending on the invoice format.
        line_pat = re.compile(
            r'([A-Z]-[A-Z0-9]+)\s+'         # M-code (group 1)
            r'\d{5,9}\s+'                    # DID
            r'\d{10,14}\s+'                  # UPC
            r'([\d.]+)\s+'                   # Qty (group 2)
            r'[\d.]+\s+'                     # Weight
            r'\$([\d,.]+)'                   # Total charges / FOB (group 3)
            r'(?:\s+\$([\d,.]+))?'           # Optional 2nd $ column / DEL (group 4)
            r'\s+(?:[\d.]+\s*%|BB\s+To\s+[\d.]+/\w+)'  # Rate: percentage OR BB To X/unit
            r'.*?\$([\d,.]+)',               # Amount Due — last $ on the line (group 5)
            re.I
        )

        skip_filter = re.compile(r'^Totals for|^Invoice|Inv\.\s*Number|program activity', re.I)

        def _data_lines(page_text):
            """Return only matchable data lines from a page, in order."""
            result = []
            for ln in page_text.splitlines():
                if skip_filter.search(ln): continue
                if line_pat.search(ln): result.append(ln)
            return result

        prev_last_line = None   # last data line of the previous page
        for page_text in pages_text:
            data_lines = _data_lines(page_text)
            for i, line in enumerate(data_lines):
                # Skip the first data line of this page if it is an exact repeat of the
                # last data line of the previous page (page-boundary carryover).
                if i == 0 and prev_last_line is not None and line == prev_last_line:
                    continue
                m = line_pat.search(line)
                item   = m.group(1).upper()
                qty    = m.group(2)
                amount = clean_amount(m.group(5))
                rows.append(make_row(
                    source='Martin Bros',
                    program_num=cfg['program_num'],
                    customer_ref=cref,
                    dist_id=cfg['dist_id'],
                    bill_date=bill_date,
                    start_date=start_date,
                    end_date=end_date,
                    item=item,
                    qty=qty,
                    amount=amount,
                    trade=cfg['trade']
                ))
            if data_lines:
                prev_last_line = data_lines[-1]
        if not rows:
            rows.append({'_error': 'Martin Bros PDF: no rows extracted'})
    except Exception as e:
        rows.append({'_error': f'Martin Bros: {e}'})
    return rows


def parse_dot_foods_bb(filepath, cfg, customer_ref):
    """DOT Foods 'Supplier Billback' multi-vendor report.
    Columns: BB Dept | BB Vendor # | BB Vendor Name | Chain | Item | Item Description | Vendor Item# | ...
    Aggregates by item using subtotal lines. Vendor Item# may contain M-code or a numeric code.
    Items without M-codes (including non-Monin products) are flagged as warnings.
    """
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            all_text = '\n'.join(p.extract_text() or '' for p in pdf.pages)

        # Date range from invoice dates in file
        bill_date = start_date = end_date = TODAY
        dates_found = re.findall(r'\b(\d{1,2}/\d{1,2}/\d{4})\b', all_text)
        if dates_found:
            from datetime import datetime
            dt_list = []
            for d in dates_found:
                try: dt_list.append(datetime.strptime(d, '%m/%d/%Y'))
                except: pass
            if dt_list:
                start_date = to_yyyymmdd(min(dt_list).strftime('%m/%d/%Y'))
                end_date   = to_yyyymmdd(max(dt_list).strftime('%m/%d/%Y'))

        # Customer ref: use passed value or DocID from filename
        cref = customer_ref
        if not cref:
            m = re.search(r'DocID\s+([\w]+)', os.path.basename(filepath), re.I)
            if m: cref = m.group(1)

        item_map = cfg.get('item_map', {})   # user-supplied DOT item# → M-code

        # Pass 1: build item# → {mcode, desc} from detail lines
        item_mcode = {}
        item_desc  = {}
        for line in all_text.splitlines():
            if re.search(r'\bTotal\b|BB Dept|BB Vendor', line, re.I):
                continue
            # Item# followed by description followed by M-code in Vendor Item# slot
            m = re.search(r'(\d{5,6})([A-Z][A-Z0-9\s/]+?)\s+([A-Z]-[A-Z0-9]+)\s+\d{6,9}', line, re.I)
            if m:
                item_mcode[m.group(1)] = m.group(3).upper()
                item_desc[m.group(1)]  = m.group(2).strip()
                continue
            # Item# followed by description with numeric Vendor Item# — capture desc only
            m2 = re.search(r'(\d{5,6})([A-Z][A-Z0-9\s/.-]+?)\s+\d{4,8}\s+\d{7,9}[A-Z]', line)
            if m2 and m2.group(1) not in item_desc:
                item_desc[m2.group(1)] = m2.group(2).strip()

        # Pass 2: process item-level Total lines (one row per item)
        # Formats: "280930 Total 7 - $ 34.65"  /  "140078 Total 28 280.00 $ 140.00"  /  "280978 Total 1 3 - $ 36.55"
        # Handle variable number of numeric/dash tokens before the final $ amount
        total_pat = re.compile(r'^(\d{5,6})\s+Total\s+(\d+)(?:\s+[-\d,.]+)*\s+\$\s*([\d,.]+)', re.M)
        for m in total_pat.finditer(all_text):
            item_num = m.group(1)
            qty      = m.group(2)
            amount   = clean_amount(m.group(3))
            mcode    = item_mcode.get(item_num) or item_map.get(item_num)
            desc     = item_desc.get(item_num, item_num)
            if mcode and re.match(r'[A-Z]-[A-Z0-9]+', str(mcode), re.I):
                rows.append(make_row(
                    source='DOT Foods BB',
                    program_num=cfg['program_num'],
                    customer_ref=cref,
                    dist_id=cfg['dist_id'],
                    bill_date=bill_date,
                    start_date=start_date,
                    end_date=end_date,
                    item=str(mcode).upper(),
                    qty=qty,
                    amount=amount,
                    trade=cfg['trade']
                ))
            else:
                rows.append({'_warning': True, 'code': item_num, 'desc': desc, 'amount': amount})

        if not rows:
            rows.append({'_error': 'DOT Foods BB: no items found'})
    except Exception as e:
        rows.append({'_error': f'DOT Foods BB: {e}'})
    return rows


def _pdfplumber_extract_first_page(filepath, timeout_s=25):
    """Extract first-page text from a PDF with a timeout to prevent hanging on Render.
    Returns (text_str, error_str) — one of them will be None."""
    import concurrent.futures
    def _extract():
        with pdfplumber.open(filepath) as pdf:
            return pdf.pages[0].extract_text() or ''
    try:
        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as exe:
            future = exe.submit(_extract)
            try:
                text = future.result(timeout=timeout_s)
                return text, None
            except concurrent.futures.TimeoutError:
                return None, 'PDF processing timed out (file may be too large or corrupt)'
            except Exception as e:
                return None, str(e)
    except Exception as e:
        return None, str(e)


def parse_supplier_billback_pdf(filepath, cfg, customer_ref):
    """Content-based dispatcher for 'Supplier Billback' PDFs — multiple distributors use this filename."""
    first_page, err = _pdfplumber_extract_first_page(filepath)
    if err is not None:
        return [{'_error': f'Could not read PDF: {err}'}]

    # Scanned / image-based PDF — no extractable text
    if len(first_page.strip()) < 50:
        return [{'_error': 'This PDF appears to be image-based (scanned). Text extraction is not supported — please enter the data manually or request a text-based PDF from the distributor.'}]

    if 'Martin Bros' in first_page:
        return parse_martin_bros(filepath, cfg, customer_ref)
    if 'Driscoll Foods' in first_page:
        return parse_trackmax(filepath, cfg, customer_ref, source_name='Driscoll Foods')
    if 'Delco Foods' in first_page:
        return parse_delco_foods(filepath, cfg, customer_ref)
    if 'Dairyland' in first_page or 'The Chefs Warehouse' in first_page or 'ChefsWhse' in first_page:
        return parse_chefs_warehouse(filepath, cfg, customer_ref)
    if re.search(r'REVENUE\s+TRACKING\s+REPORT', first_page, re.I) and 'MFG#' in first_page:
        return parse_atlas(filepath, cfg, customer_ref)
    # Kohl Wholesale — Trackmax PDF with numeric ProductIDs and M-codes in UPC slot
    if 'Kohl Wholesale' in first_page:
        return parse_sw_pdf(filepath, cfg, customer_ref, source_override='Kohl Wholesale')
    # S&W Trackmax PDF — uses numeric product IDs instead of M-codes
    if 'S&W Wholesale' in first_page or 's-wfoods' in first_page.lower():
        return parse_sw_pdf(filepath, cfg, customer_ref)
    # SGC Foodservice / other Trackmax variants with no Total Charges column
    # These use: M-code  DID(no CS)  UPC  Qty  Weight  BB To X/unit  $AmtDue
    if 'SGC Foodservice' in first_page or 'sgcfoodservice' in first_page.lower():
        return parse_sw_pdf(filepath, cfg, customer_ref)
    # Detect Trackmax format that has NO separate Total Charges column
    # (only one $ amount per line — Amount Due only)
    # Check for "BB To" pattern without a preceding $XX Total Charges
    if ('Product ID' in first_page and 'DID' in first_page and 'UPC' in first_page):
        # Choose Trackmax variant by program amount format:
        #   "BB To X.000/unit $Y"            → S&W/SGC/Y.Hata/ChristPanos style (parse_sw_pdf)
        #   "X% of FOB/Del" + "DOT FOODS"   → Dennis/Springfield/similar DOT-distributed (parse_sw_pdf)
        #   "X.00 % of Del $Y" (no DOT)     → BEK/Driscoll style (parse_trackmax)
        if re.search(r'BB\s+To\s+[\d.]+/unit', first_page, re.I):
            return parse_sw_pdf(filepath, cfg, customer_ref)
        elif (re.search(r'%\s+of\s+(?:FOB|Del)', first_page, re.I)
              and 'DOT FOODS' in first_page):
            return parse_sw_pdf(filepath, cfg, customer_ref)
        else:
            return parse_trackmax(filepath, cfg, customer_ref)
    if 'Powered byTrackmax' in first_page:
        return parse_trackmax(filepath, cfg, customer_ref)
    if 'BB Dept' in first_page and 'BB Vendor' in first_page:
        return parse_dot_foods_bb(filepath, cfg, customer_ref)
    if 'Harbor Food' in first_page:
        # Harbor sends XLS usually; PDF variant — generic fallback
        return [{'_error': 'Harbor Foodservice PDF format not yet supported — please send the XLSX version'}]

    # ── Generic Trackmax fallback ────────────────────────────────────────────────
    # All Trackmax-style PDFs share the same three invariants regardless of
    # column order, rate format, or distributor name:
    #   1. Item codes are always in the form  M-XXXXXXX  (or P-codes / numeric)
    #   2. Large digit-only numbers after the code are DID / UPC — skip them
    #   3. The last  $X.XX  on each data line is always the billback Amount Due
    #
    # This parser reads any unknown Trackmax variant without format-specific code.
    # Page-boundary dedup (same line printed at bottom of page N and top of page
    # N+1) is handled by comparing the first data line of each page to the last
    # data line of the previous page.
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            pages_text = [page.extract_text() or '' for page in pdf.pages]
        all_text = '\n'.join(pages_text)

        bill_date = start_date = end_date = ''
        m = re.search(r'generated\s+on\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: bill_date = to_yyyymmdd(m.group(1))
        m = re.search(r'between\s+(\d{1,2}/\d{1,2}/\d{4})\s+and\s+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I)
        if m: start_date, end_date = to_yyyymmdd(m.group(1)), to_yyyymmdd(m.group(2))
        if not start_date:
            m2 = re.search(r'Start\s+Date[:\s]+(\d{1,2}/\d{1,2}/\d{4}).*?Stop\s+Date[:\s]+(\d{1,2}/\d{1,2}/\d{4})', all_text, re.I | re.S)
            if m2: start_date, end_date = to_yyyymmdd(m2.group(1)), to_yyyymmdd(m2.group(2))

        inv_num = ''
        m3 = re.search(r'(?:Our\s+)?Invoice\s+Number[:\s]+(\w+)', all_text, re.I)
        if m3: inv_num = m3.group(1).strip()
        cref = customer_ref or inv_num

        # Pattern: M-code → skip digit-only IDs → first decimal = qty → last $X.XX = amount
        code_pat  = re.compile(r'(M-[A-Z][A-Z0-9]+|P\d{3,4})', re.I)
        # After the code: skip DID/UPC (pure-digit tokens 5+ chars), take first decimal as qty
        qty_pat   = re.compile(r'(?:\s+\d{5,})*\s+([\d]+\.[\d]+)', re.I)
        # Last dollar amount on the line
        amt_pat   = re.compile(r'\$([\d,]+\.[\d]{2})(?![\d])')
        skip_pat  = re.compile(r'Totals?\s+for|^Invoice|Inv\.?\s*Num|program\s+activity|'
                                r'Product\s+ID|Description|PackSize|^\s*$', re.I)

        from collections import defaultdict
        totals_qty = defaultdict(float)
        totals_amt = defaultdict(float)

        prev_last_line = None
        for page_text in pages_text:
            data_lines = []
            for line in page_text.splitlines():
                if skip_pat.search(line): continue
                cm = code_pat.search(line)
                if not cm: continue
                # Must have at least one dollar amount on the line
                if not amt_pat.search(line): continue
                data_lines.append(line)

            for i, line in enumerate(data_lines):
                # Skip first line of page if it's an exact repeat of previous page's last line
                if i == 0 and prev_last_line is not None and line == prev_last_line:
                    continue
                cm = code_pat.search(line)
                item = cm.group(1).upper()
                # Qty: first decimal after the code, skipping large digit-only IDs
                after_code = line[cm.end():]
                qm = qty_pat.match(after_code)
                qty = float(qm.group(1)) if qm else 1.0
                # Amount: last $X.XX on the line
                all_amts = amt_pat.findall(line)
                if not all_amts: continue
                amt = float(all_amts[-1].replace(',', ''))
                totals_qty[item] += qty
                totals_amt[item] += amt

            if data_lines:
                prev_last_line = data_lines[-1]

        source_name = 'Dist PDF'
        # Try to extract a distributor name from the header
        for hint in ['Martin Bros', 'Driscoll', 'Delco', 'Cheney', 'SOFO', 'BEK']:
            if hint.lower() in all_text.lower():
                source_name = hint
                break

        for item in sorted(totals_qty):
            rows.append(make_row(
                source=source_name,
                program_num=cfg['program_num'],
                customer_ref=cref,
                dist_id=cfg['dist_id'],
                bill_date=bill_date,
                start_date=start_date,
                end_date=end_date,
                item=item,
                qty=totals_qty[item],
                amount=totals_amt[item],
                trade=cfg['trade']
            ))
        if not rows:
            rows.append({'_error': 'Could not extract any M-code rows from this PDF — '
                                   'please verify it is a Trackmax-style billback.'})
    except Exception as e:
        rows.append({'_error': f'Supplier Billback PDF: {e}'})
    return rows


def parse_atlas(filepath, cfg, customer_ref):
    """Atlas Wholesale Food 'Revenue Tracking Report' PDF.
    Each product has a MFG# subtotal line:
        MFG# M-XXXXXX  <TotalQty>  <TotalCost>  <Revenue>  [0.00]
    Revenue (4th number) is the billback amount; TotalQty is the 2nd number.
    Date range comes from: FROM: MM/DD/YY TO MM/DD/YY
    """
    rows = []
    try:
        with pdfplumber.open(filepath) as pdf:
            all_text = '\n'.join(page.extract_text() or '' for page in pdf.pages)

        bill_date = start_date = end_date = ''
        dr = re.search(r'FROM:\s+(\d{1,2}/\d{1,2}/\d{2,4})\s+TO\s+(\d{1,2}/\d{1,2}/\d{2,4})', all_text, re.I)
        if dr:
            start_date = to_yyyymmdd(dr.group(1))
            end_date   = to_yyyymmdd(dr.group(2))
            bill_date  = end_date

        inv_num = ''
        m = re.search(r'DocID[_\s]*([\w-]+)', os.path.basename(filepath), re.I)
        if m: inv_num = m.group(1).strip()
        cref = customer_ref or inv_num

        # MFG# subtotal line: M-code  qty  totalcost  revenue
        mfg_pat = re.compile(
            r'MFG#\s+(M-[A-Z][A-Z0-9]+)\s+'
            r'([\d,]+)\s+'        # total qty
            r'[\d,.]+\s+'         # total cost (skip)
            r'([\d,.]+)',          # revenue = billback amount
            re.I
        )

        from collections import defaultdict
        totals_qty = defaultdict(float)
        totals_amt = defaultdict(float)

        for m in mfg_pat.finditer(all_text):
            item = m.group(1).upper()
            qty  = float(m.group(2).replace(',', ''))
            amt  = float(m.group(3).replace(',', ''))
            totals_qty[item] += qty
            totals_amt[item] += amt

        for item in sorted(totals_qty):
            rows.append(make_row(
                source='Atlas Wholesale',
                program_num=cfg['program_num'],
                customer_ref=cref,
                dist_id=cfg['dist_id'],
                bill_date=bill_date,
                start_date=start_date,
                end_date=end_date,
                item=item,
                qty=totals_qty[item],
                amount=totals_amt[item],
                trade=cfg['trade']
            ))
        if not rows:
            rows.append({'_error': 'Atlas Wholesale PDF: no MFG# product lines found'})
    except Exception as e:
        rows.append({'_error': f'Atlas Wholesale parser: {e}'})
    return rows


def parse_cheney(filepath, cfg, customer_ref):
    """Cheney Brothers XLSX billback.
    Columns (per row-8 annotations):
      W = Manufacture Part Number (M-code)   → Item Number
      Y = Quantity                            → Item Volume Qty
      AG = Net Value                          → Item Dollar Amount
      H = Document Date                       → Billback Date
      E = PO Invoice Date                     → BB Start Date AND BB End Date
      AI = CLAIM NO                           → Customer Ref
    """
    rows = []
    try:
        df = pd.read_excel(filepath, header=0)
        # Billback date from Document Date (col H); start/end from PO Invoice Date (col E)
        bill_dates = [to_yyyymmdd(v) for v in df.get('Document Date',   []) if to_yyyymmdd(v)]
        po_dates   = [to_yyyymmdd(v) for v in df.get('PO Invoice Date', []) if to_yyyymmdd(v)]
        bill_date  = bill_dates[0]  if bill_dates else ''
        start_date = po_dates[0]    if po_dates   else bill_date
        end_date   = po_dates[0]    if po_dates   else bill_date
        # Customer ref from CLAIM NO
        if not customer_ref and 'CLAIM NO' in df.columns:
            cr = df['CLAIM NO'].iloc[0]
            if pd.notna(cr):
                customer_ref = str(int(cr)) if isinstance(cr, (float, int)) else str(cr)
        # Aggregate by M-code
        from collections import defaultdict
        totals_qty = defaultdict(float)
        totals_amt = defaultdict(float)
        for _, row in df.iterrows():
            raw = str(row.get('Manufacture Part Number', '') or '').strip()
            if not raw or raw.lower() == 'nan':
                continue
            code = raw.upper()
            totals_qty[code] += float(row.get('Quantity', 0) or 0)
            totals_amt[code] += float(row.get('Net Value',  0) or 0)
        if not totals_amt:
            return [{'_error': 'Cheney: no product rows found (missing Manufacture Part Number or Net Value)'}]
        for code, amt in totals_amt.items():
            rows.append(make_row(
                source='Cheney Brothers',
                program_num=cfg['program_num'], customer_ref=customer_ref,
                dist_id=cfg['dist_id'], bill_date=bill_date,
                start_date=start_date, end_date=end_date,
                item=code, qty=totals_qty[code], amount=amt, trade=cfg['trade'],
            ))
    except Exception as e:
        rows.append({'_error': f'Cheney: {e}'})
    return rows


def parse_harbor(filepath, cfg, customer_ref):
    """Harbor Foodservice 'Supplier Billback' — XLSX only; PDFs routed by content."""
    if filepath.lower().endswith('.pdf'):
        return parse_supplier_billback_pdf(filepath, cfg, customer_ref)
    # Content-based XLSX detection — route other formats before trying Harbor columns
    try:
        df_peek = pd.read_excel(filepath, header=None, nrows=15)
        flat = ' '.join(str(v) for row in df_peek.values for v in row if str(v) != 'nan')
        if 'BB Dept' in flat:
            return parse_dot_foods_bb_xlsx(filepath, cfg, customer_ref)
        if 'Cheney Invoice No' in flat:
            return parse_cheney(filepath, cfg, customer_ref)
    except Exception:
        pass
    rows = []
    try:
        wb = __import__('openpyxl').load_workbook(filepath, data_only=True)
        ws = wb.active

        # Find invoice date (row with 'Invoice Date:' label, typically row 6)
        bill_date = ''
        period_str = ''
        for row in ws.iter_rows(max_row=10, values_only=True):
            for i, cell in enumerate(row):
                if isinstance(cell, str) and 'invoice date' in cell.lower():
                    # Date value is in the next cell
                    if i + 1 < len(row) and row[i+1] is not None:
                        bill_date = to_yyyymmdd(row[i+1])
                    break

        # Find header row (has 'Vendor Item Num' or 'Billback Amt')
        header_row_idx = None
        col_map = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_vals = [str(v).strip() if v is not None else '' for v in row]
            if any('vendor item' in c.lower() for c in row_vals):
                header_row_idx = i
                col_map = {v.strip().lower(): j for j, v in enumerate(row_vals)}
                break

        if header_row_idx is None:
            return [{'_error': 'Harbor: could not find header row'}]

        # Column indices
        def gc(*keys):
            for k in keys:
                for ck, ci in col_map.items():
                    if k.lower() in ck: return ci
            return None

        ci_mcode   = gc('vendor item num', 'vendor item')
        ci_period  = gc('invoice period', 'period')
        ci_recqty  = gc('rec qty', 'received qty')
        ci_ordqty  = gc('order qty')
        ci_bbamt   = gc('billback amt', 'billback amount')
        ci_invoice = gc('invoice #', 'invoice#')

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not any(v is not None for v in row): continue
            cells = [str(v).strip() if v is not None else '' for v in row]

            # Skip total row
            if any('total' in c.lower() for c in cells if c): continue

            # Get M-code — extract first M-XXXXX pattern from cell
            raw_mcode = cells[ci_mcode] if ci_mcode is not None else ''
            m = re.search(r'[A-Z]-[A-Z0-9]+', raw_mcode, re.I)
            if not m:
                # Check user-defined item code mapping
                item_map = cfg.get('item_map', {})
                mapped = item_map.get(raw_mcode.strip(), '')
                if mapped and re.match(r'[A-Z]-[A-Z0-9]+', mapped, re.I):
                    item = mapped.upper()
                else:
                    # Still unknown — flag it as a structured warning
                    if raw_mcode and raw_mcode not in ('', 'None'):
                        desc    = cells[10] if len(cells) > 10 else ''
                        amt_val = clean_amount(cells[ci_bbamt]) if ci_bbamt is not None else 0.0
                        rows.append({'_warning': True, 'code': raw_mcode, 'desc': desc, 'amount': amt_val})
                    continue
            else:
                item = m.group(0).upper()

            # Period for date range
            if ci_period is not None and not period_str:
                period_str = cells[ci_period]
            start_date, end_date = month_name_to_range(cells[ci_period] if ci_period is not None else period_str)

            # Qty: prefer rec qty, fall back to order qty
            qty_val = cells[ci_recqty] if ci_recqty is not None else ''
            if not qty_val and ci_ordqty is not None:
                qty_val = cells[ci_ordqty]

            # Amount
            amt_val = clean_amount(cells[ci_bbamt]) if ci_bbamt is not None else 0.0

            # Invoice # as customer ref if not provided
            inv_num = cells[ci_invoice] if ci_invoice is not None else ''
            cref = customer_ref or inv_num

            rows.append(make_row(
                source='Harbor',
                program_num=cfg['program_num'],
                customer_ref=cref,
                dist_id=cfg['dist_id'],
                bill_date=bill_date,
                start_date=start_date,
                end_date=end_date,
                item=item,
                qty=qty_val,
                amount=amt_val,
                trade=cfg['trade']
            ))
    except Exception as e:
        rows.append({'_error': f'Harbor: {e}'})
    return rows


# ─── DISPATCHER ──────────────────────────────────────────────────────────────
def detect_and_parse(filepath, user_config=None, customer_ref='', file_override=None, original_filename=None):
    # Honour user-selected supplier from the UI dropdown; fall back to auto-detection
    forced = (file_override or {}).get('supplier', '')
    if forced and forced != 'UNKNOWN':
        supplier = forced
    else:
        # Use original filename for detection — safe_name (spaces→underscores) breaks some patterns
        supplier = detect_supplier(original_filename or filepath)
    cfg = dict(DEFAULT_SUPPLIER_CONFIG.get(supplier, DEFAULT_SUPPLIER_CONFIG['UNKNOWN']))
    if user_config and supplier in user_config:
        uc = user_config[supplier]
        if uc.get('program_num'): cfg['program_num'] = uc['program_num']
        if uc.get('dist_id'):     cfg['dist_id']     = uc['dist_id']
        if uc.get('trade'):       cfg['trade']       = uc['trade']
        if uc.get('item_map'):    cfg['item_map']    = uc['item_map']  # {harbor_code: m_code}
    # Per-file overrides take highest priority (entered by user in the file row)
    if file_override:
        row_type = file_override.get('row_type', 'Program')
        if row_type == 'Operator':
            cfg['program_num']  = ''          # Operator rows have no Program #
            cfg['trade']        = 'O'
            cfg['operator_id']  = file_override.get('operator_name', '')
        else:
            if file_override.get('program_num'): cfg['program_num'] = file_override['program_num']
            cfg['trade'] = 'D'
        if file_override.get('dist_id'):     cfg['dist_id']     = file_override['dist_id']
        if file_override.get('customer_ref') and not customer_ref:
            customer_ref = file_override['customer_ref']

    fn = os.path.basename(filepath).upper()
    ext = os.path.splitext(filepath)[1].lower()

    # Capture operator override settings to apply after parsing
    _operator_override = cfg.get('operator_id', '') if cfg.get('trade') == 'O' else ''

    # Always use the dropdown display name as Source (whatever is showing in the UI)
    _source_override = (file_override or {}).get('supplier_display', '').strip()

    def _apply_type_override(rows):
        """After parsing, patch trade indicator, operator ID, and source name."""
        for row in rows:
            if '_error' in row: continue
            if _source_override:
                row['Source'] = _source_override
            if _operator_override:
                row['Trade Indicator'] = 'O'
                row['Operator ID']     = _operator_override
                row['Program #']       = ' '
        return rows

    def _ret(sup, rows):
        return sup, _apply_type_override(rows)

    if supplier == 'BEK':
        return _ret(supplier, parse_bek_or_nich(filepath, 'BEK', cfg, customer_ref))
    elif supplier == 'NICH_CO':
        return _ret(supplier, parse_bek_or_nich(filepath, 'Nich&Co', cfg, customer_ref))
    elif supplier == 'SHAMROCK':
        return _ret(supplier, parse_shamrock(filepath, cfg, customer_ref))
    elif supplier == 'DOT_CBBB':
        return _ret(supplier, parse_dot_cbbb(filepath, cfg, customer_ref))
    elif supplier in ('MCLANE','MCLANE_OR_DOT'):
        return _ret(supplier, parse_mclane(filepath, cfg, customer_ref))
    elif supplier == 'S_AND_W':
        return _ret(supplier, parse_sw(filepath, cfg, customer_ref))
    elif supplier == 'KAST':
        return _ret(supplier, parse_kast(filepath, cfg, customer_ref))
    elif supplier == 'SOFO':
        return _ret(supplier, parse_sofo(filepath, cfg, customer_ref))
    elif supplier == 'PFS' or supplier == 'PFS_STYLE':
        return _ret(supplier, parse_pfs(filepath, cfg, customer_ref))
    elif supplier == 'Y_HATA':
        return _ret(supplier, parse_yhata(filepath, cfg, customer_ref))
    elif supplier == 'TANKERSLEY':
        # Tankersley uses the same Trackmax PDF format — route through content-based dispatcher
        return _ret(supplier, parse_supplier_billback_pdf(filepath, cfg, customer_ref))
    elif supplier == 'CHRIST_PANOS':
        # Christ Panos uses Trackmax % of FOB format; cfg carries did_map for M-code resolution
        return _ret(supplier, parse_supplier_billback_pdf(filepath, cfg, customer_ref))
    elif supplier == 'HENRY_FOODS':
        return _ret(supplier, parse_henrys_foods(filepath, cfg, customer_ref))
    elif supplier == 'DELCO_FOODS':
        return _ret(supplier, parse_delco_foods(filepath, cfg, customer_ref))
    elif supplier == 'CHEFS_WH':
        return _ret(supplier, parse_chefs_warehouse(filepath, cfg, customer_ref))
    elif supplier == 'ATLAS':
        return _ret(supplier, parse_atlas(filepath, cfg, customer_ref))
    elif supplier == 'KOHL_WH':
        return _ret(supplier, parse_sw_pdf(filepath, cfg, customer_ref, source_override='Kohl Wholesale'))
    elif supplier == 'LABATT':
        return _ret(supplier, parse_labatt(filepath, cfg, customer_ref))
    elif supplier == 'HARBOR':
        result = parse_harbor(filepath, cfg, customer_ref)
        # Content-based routing may have identified a different distributor — label accordingly
        if result and isinstance(result[0], dict):
            src = result[0].get('Source', '')
            if src == 'Martin Bros':
                return _ret('MARTIN_BROS', result)
            if src == 'Driscoll Foods':
                return _ret('DRISCOLL', result)
            if src == 'Delco Foods':
                return _ret('DELCO_FOODS', result)
            if src == 'The Chefs Warehouse':
                return _ret('CHEFS_WH', result)
            if src == 'Atlas Wholesale Food':
                return _ret('ATLAS', result)
            if src == 'Kohl Wholesale':
                return _ret('KOHL_WH', result)
        return _ret(supplier, result)
    elif supplier == 'MARTIN_BROS':
        return _ret(supplier, parse_martin_bros(filepath, cfg, customer_ref))
    elif supplier == 'DRISCOLL':
        return _ret(supplier, parse_driscoll(filepath, cfg, customer_ref))
    elif supplier == 'CHENEY':
        return _ret(supplier, parse_cheney(filepath, cfg, customer_ref))
    elif supplier == 'BLAIR_CANDY':
        return supplier, [{'_error': 'Blair Candy uses scanned/image PDFs — text extraction not supported. Please enter manually.'}]
    else:
        # Try content-based detection for PDFs (S&W, Martin Bros, Driscoll, Trackmax, etc.)
        if ext == '.pdf':
            return _ret(supplier, parse_supplier_billback_pdf(filepath, cfg, customer_ref))
        return supplier, [{'_error': f'Unknown supplier format: {os.path.basename(filepath)}'}]

# ─── OUTPUT BUILDER ──────────────────────────────────────────────────────────
def build_output(all_rows):
    """Write all_rows (list of dicts) to a new workbook with exact Sheet1 format."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Header row — bold
    for col_idx, col_name in enumerate(S1_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, row_dict in enumerate(all_rows, 2):
        if '_error' in row_dict:
            ws.cell(row=row_idx, column=1, value=f"ERROR: {row_dict['_error']}")
            continue
        for col_idx, col_name in enumerate(S1_COLS, 1):
            val = row_dict.get(col_name)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width
    for col_idx, col_name in enumerate(S1_COLS, 1):
        max_len = max(len(str(col_name)), 10)
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─── HTML UI ─────────────────────────────────────────────────────────────────
HTML_PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Monin Billback Processor</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#222}
  .header{background:linear-gradient(135deg,#1a3a5c,#2e6da4);color:#fff;padding:20px 32px;display:flex;align-items:center;gap:16px}
  .header h1{font-size:1.4rem;font-weight:600}
  .header .sub{font-size:.85rem;opacity:.8}
  .container{max-width:1100px;margin:24px auto;padding:0 16px}
  .card{background:#fff;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,.08);padding:24px;margin-bottom:20px}
  .card h2{font-size:1rem;font-weight:600;color:#1a3a5c;margin-bottom:16px;border-bottom:2px solid #e8eef4;padding-bottom:8px}
  .drop-zone{border:2px dashed #2e6da4;border-radius:8px;padding:36px;text-align:center;cursor:pointer;transition:.2s;background:#f8fafc}
  .drop-zone:hover,.drop-zone.drag-over{background:#e8f0fb;border-color:#1a3a5c}
  .drop-zone svg{color:#2e6da4;margin-bottom:8px}
  .drop-zone p{color:#666;font-size:.9rem}
  .drop-zone strong{color:#2e6da4}
  #file-input{display:none}
  .file-list{margin-top:12px}
  .file-item{display:flex;flex-direction:column;gap:6px;padding:10px 14px;background:#f8fafc;border-radius:6px;margin-bottom:8px;border:1px solid #e2e8f0}
  .file-item-top{display:flex;align-items:center;gap:10px;width:100%}
  .file-item-fields{display:flex;gap:10px;flex-wrap:wrap;align-items:center;padding:6px 8px;background:#f1f5f9;border-radius:5px;border:1px dashed #cbd5e1}
  .file-item-fields label{font-size:.72rem;color:#64748b;white-space:nowrap;display:flex;align-items:center;gap:4px}
  .file-badge{font-size:.7rem;font-weight:700;padding:2px 7px;border-radius:10px;background:#dbeafe;color:#1d4ed8;white-space:nowrap}
  .file-badge.unknown{background:#fef3c7;color:#92400e}
  .file-badge.error{background:#fee2e2;color:#991b1b}
  .file-name{flex:1;font-size:.85rem;font-weight:500;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
  .cref-input,.override-input{font-size:.8rem;border:1px solid #cbd5e1;border-radius:4px;padding:4px 8px;background:#fff}
  .cref-input{width:200px}
  .override-input{width:120px}
  /* Searchable distributor combobox */
  .combo-wrap{position:relative;display:inline-block;min-width:180px;max-width:240px}
  .combo-input{font-size:.78rem;font-weight:700;border:1px solid #93c5fd;border-radius:10px;padding:3px 26px 3px 10px;background:#dbeafe;color:#1d4ed8;cursor:text;width:100%;box-sizing:border-box;outline:none}
  .combo-input::placeholder{color:#93c5fd;font-weight:400}
  .combo-input.unknown{background:#fef9c3;border-color:#f59e0b;color:#92400e}
  .combo-arrow{position:absolute;right:7px;top:50%;transform:translateY(-50%);pointer-events:none;font-size:.65rem;color:#93c5fd}
  .combo-list{position:absolute;left:0;top:calc(100% + 3px);width:240px;max-height:220px;overflow-y:auto;background:#fff;border:1px solid #93c5fd;border-radius:8px;box-shadow:0 4px 16px rgba(0,0,0,.12);z-index:9999;display:none}
  .combo-item{padding:5px 12px;font-size:.8rem;cursor:pointer;color:#1e293b}
  .combo-item:hover,.combo-item.focused{background:#dbeafe;color:#1d4ed8}
  .combo-item.no-match{color:#94a3b8;font-style:italic;cursor:default}
  .remove-btn{background:none;border:none;cursor:pointer;color:#94a3b8;font-size:1rem;padding:2px 6px}
  .remove-btn:hover{color:#dc2626}
  .config-toggle{font-size:.8rem;color:#2e6da4;cursor:pointer;text-decoration:underline;margin-bottom:10px;display:inline-block}
  .config-section{display:none}
  .config-section.open{display:block}
  .config-table{width:100%;border-collapse:collapse;font-size:.82rem}
  .config-table th{background:#f1f5f9;color:#475569;font-weight:600;padding:7px 10px;text-align:left;border-bottom:2px solid #e2e8f0}
  .config-table td{padding:6px 10px;border-bottom:1px solid #f1f5f9}
  .config-table input,.config-table select{font-size:.8rem;border:1px solid #cbd5e1;border-radius:4px;padding:3px 7px;width:100%}
  .btn{display:inline-flex;align-items:center;gap:8px;padding:11px 24px;border-radius:6px;font-weight:600;font-size:.9rem;cursor:pointer;border:none;transition:.2s}
  .btn-primary{background:#2e6da4;color:#fff}
  .btn-primary:hover{background:#1a3a5c}
  .btn-primary:disabled{background:#94a3b8;cursor:not-allowed}
  .btn-secondary{background:#f1f5f9;color:#334155;border:1px solid #e2e8f0}
  .btn-secondary:hover{background:#e2e8f0}
  .actions{display:flex;gap:12px;margin-top:8px}
  .progress{display:none;margin-top:16px}
  .progress-bar{height:6px;background:#e2e8f0;border-radius:3px;overflow:hidden;margin-bottom:8px}
  .progress-fill{height:100%;background:linear-gradient(90deg,#2e6da4,#60a5fa);width:0%;transition:width .3s;border-radius:3px}
  .progress-text{font-size:.82rem;color:#64748b}
  .results{display:none;margin-top:16px}
  .result-row{display:flex;align-items:center;justify-content:space-between;padding:8px 12px;border-radius:5px;margin-bottom:5px;font-size:.84rem}
  .result-ok{background:#f0fdf4;border-left:3px solid #22c55e}
  .result-err{background:#fff1f2;border-left:3px solid #f43f5e}
  .result-skip{background:#fffbeb;border-left:3px solid #f59e0b}
  .count-badge{font-weight:700;font-size:.8rem;padding:1px 8px;border-radius:8px;background:#dbeafe;color:#1d4ed8}
  .dl-box{background:linear-gradient(135deg,#f0fdf4,#dcfce7);border:1px solid #86efac;border-radius:8px;padding:16px 20px;text-align:center;margin-top:16px}
  .dl-box p{font-size:.9rem;color:#166534;margin-bottom:10px}
  .total-bar{background:#1a3a5c;color:#fff;border-radius:6px;padding:8px 16px;display:inline-block;font-weight:700;margin-bottom:12px;font-size:.95rem}
</style>
</head>
<body>
<div class="header">
  <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAABJ4AAAH2CAYAAAAxjEZ6AAAACXBIWXMAAC4jAAAuIwF4pT92AAAgAElEQVR4nOzdXWxcZ37n+d+pokRJlkTKLbvdbqlZNe52dzudFt3dSQfV0aqcbF56gIS0gUxnFxiIxgKLvbO8c7uAy8Bgb+bCNDY3i70QNbnZIICbGmCRnkwmLqUx1ZvBbExlMnHifqlirHZbFi2p9EZSEnn2op6SSlSRrHPqnHqe55zvBxBsS1Wn/hbJU6d+5//8nyAMQwEAAAAAAABJK9guAAAAAAAAANlE8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFIxZrsAAAAAANjNvyn960lJ01t//zdfnDz45KGxW93/Lv3b/6U+yroAADsLwjC0XQMAAAAA6N+U/nVJnXCp+6sbNk30e/yxo+NXv/X8wSf3jAX9/nhZUktSa23lJx9/evH//oGkpe/8/NL1FEoHAGyD4AkAAACAFa8cmyk9X3zuf/1M8OQXJFW1TcDUz7eeP7hafmbf/kEee/39/2f19kfvdR97UdKSpEVJdYIoAEgXwRMAAACAkXnl2My0pLmDwRP/4rlCuXAg2P/ZKM8vBFr9zRcn908eHGxqyPq15ZWVv/njozs85Lw6IdQiIRQAJI/gCQAAAECqXjk2U5I0K+mMpKkvFI61P1/43MDdTV17x4L2qa9PTAwaOm3eX2tf/k9/NLF5f23Xx479/rdV+OLnzkla+JV/9Uf1qLUBAPojeAIAAACQileOzVQlzUk6LUnjwbi+XPjiyhPBgZ06kPo6sK9w9be/cWS7eU59ffKf/y/du3l59wcWC6t7X/v93mV7FyTVCKAAYHgETwAAAAASZQKnmqRT3d/7THDkzpeKzwWBgoHmMvWKEzrdXG60b/zkLwfqqip8bWpl7Ldf7BeGEUABwJAIngAAAAAkwsxvmldP4CRJcZfWSfFCp4219tWP/9P/8eSgj9/zL19S8NSO5Z2XNPcr/+qPmAEFABEVbBcAAAAAwG+vHJuZfOXYzIKk97QldHq++MWrowydwnDzzpX/748HDp2CQ/tXdgmdJGlGUuvS//Ts/zxwIQAASQRPAAAAAIbwyrGZM5JaMnOcusZU1DfGTlz9THBk4BCoV5zQSZLa//Bnwcba4I1JxV/7yhODPG5s/ebE+K3L/+eV7xUXrnyvOBmpKADIscG2gwAAAACAHmanugVt6XCSOqHT18e+dnVce0caOq1fW165/dF7gw8uLxZWC88/O9DMqbH1m91/PS1p+sr3irNP/clGK1KBAJBDdDwBAAAAiOSVYzOzkpbkUOgUbm6sXf3bP420W16h9NkbGt8z0GPHb1++1vOfJyQtXflecTrK6wFAHhE8AQAAABjYK8dm5iV9X9Jjg5GGDZ0KgVbjhE6SdO2/fX9z8/5apOcUX/rlzw762D2r1+5s+a0JSXXCJwDYGUvtAAAAAOzqlWMzk5Lq6nT7PCaJ0Ok3X5zcHyd0uvPR0srqJ/8QqdspOHr4cnD4wMDB0/itj+/3+e1u+FR96k82lqK8PgDkBR1PAAAAAHb0yrGZaXWW1vUNnSTpheJXVuKGTpL07a8eCicPRr8vvnl/rd3+8V9ECp0kqfjS1wcOnSRpz1p7aps/6oZPDBwHgD4IngAAAABsy4ROdUnbBS96vvjFq08EByKHP11f/2dPrB07On4gznNX/uaPJ6IusQuOHr5cOD54uftufLTbQybU+TsCAGxB8AQAAACgr57Q6bF5Tl3PFp5Z+0xwJHan07Gj41e/fGz/vjjPvbncaN+7eTny84qVr0bqdtoyWHw7J658rzgfuRgAyDiCJwAAAACPGSR0Ohwc0lTheKzQSOrsYPet5w/GCq021m6s3PjJX25b27b27WkXvvi5SE85eOUftg4W385rV75XrEauCQAyjOAJAAAAwCMGCZ3GVNRXi1+OtsatRyHQavXrE7F2sJOkT//2T2It7Rs7+bW9UR5f2FjXnrXrnx/4Cfv0byMXBQAZRvAEAAAA4IFBQiepM0y8oCB2t9O3v3oofGJfMdZz4y6xC44evlz45an9UZ4zwHynR1/jM+Hx6z8MapGeBAAZRvAEAAAAQJL0yrGZSQ0QOj1beGZtmGHi5Wf2rcQdJh57iZ2i72QnSYeu/P2HMV7qzPUfBuxyBwAieAIAAACgwUOn8WBcXygc24z7OgfGCyvfev5g7NAq7hK7qDvZde1rX4r0pEJZl9X5OzwT+cUAIIMIngAAAABI0rykE7s96MuFL64ECmJ1K0nSd37pcOzQKe4SOylut9P7q0G4GWlpnvbpvvm3uaivBwBZRPAEAAAA5Nwrx2bmJJ3e7XFPF55aHWaJ3VeO729PHhyL9dxhltgVjh9didPtdPDK+ytRnxN8LuwGT1PXfxjMRn5RAMgYgicAAAAgx8ww8fndHjemosqFqXhb0KmzxO6Xy0/ECo6k+EvsJKn4O9+I/Nyx9Zsav3X5eNTnBROa6vnPuajPB4CsIXgCAAAA8m1Bu8x1kqRnC59rD7OLna0ldoWvTa0Eh6OvDDz0yX9rR31O8Pj/YTXyCwNAxhA8AQAAADn1yrGZmgaY6zQejOvZwjN74r7OMEvsNu+vtW+1GvE6pYqF1bFTX4sVeB3++OLeqM8JpsJrW35r4voPg2qc1weArCB4AgAAAHLolWMzJQ2481qp8IXLcQeK7x0L2l85fiD2Erurf/unE5v312I9t/idFwKNR8/LYg0Vl1Qoh7f6/HY1cgEAkCEETwAAAEA+zWuAJXbjwbieDCYj7wjX9c3nD+7ZMxZvNNTalQ8ur19bjvXc4ND+leK3vhhraeCRD390N9ZrPqV+3VXTcY4FAFlB8AQAAADkzCvHZqqSZgZ5bKnwhXjDlSRNPjG2cuzoeKxOqXBzY+3a3/+72IFX8Xe/GWuJ3YFrTRXur0fu0Cr8s1Aqql+XVClOHQCQFQRPAAAAQP7suoud1NnJ7kgwcTjui1R+6VDsgeLtf/xBGHeJXeH40ZXC8XgvPXnpP8cK2gq/pA+3+aNdZ2gBQJYRPAEAAAA58sqxmTkNGIY8W/hcO1AQedaRJD37mb2Xn9hXjPNU3bt15fLtj96L9boqFO4Uf+cb8bqdrv7szt7Vq7G6rIJjYeyQDQCyjOAJAAAAyJfaoA98pvD0ZtwXmX7uidjL5K79/WL8JXbffO5ecDjW6j4dbdXvxXle4YVwdZtldpKk6z8MmPMEILcIngAAAICceOXYzKykqUEe+2RwREUVj8R5na8c39+O2+1056OllXs3Y46V2renXTz5S7F20Dt05f3VOLOdJKlwIryxy0Mm4xwXALKA4AkAAADIjzODPvDpwlPbzSzaUSHQ6leOH4gV4ISbG2vtH/9F7CVrY7/37XjB0cZ6/J3sDkvBZxS7QwsAso7gCQAAAMiBV47NlCSdGuSxZqj48TivM/XZfbf3jAVxnjrcQPHnPnc57kDxiY/ea8fudjoZDtKedT3OsQEgCwieAAAAgHwYuNvpycKTq3Ff5Ktf2B8r/dlYu7ESe6B4sbA69rvfiNV1NLZ+U4c/vrgnznODw1KhHO76upMnw6U4xweALCB4AgAAAPJhbtAHPh0cXYnzAsPsZHft78/HXmJX/M4LgcZjZUc62vzLlSDcjDWNfMBuJwDINYInAAAAIOPMUPGBl5IdCg7GWmYXdye7e7cur6xfW47zVAVHD18ufuuL++I8d9+NS7f33fgoVuA1aLeTpHj/YwCQEQRPAAAAQPbNDvrAJ4NYG9lp4oli7G6nTy/+afyB4t/9Zqywq7Cxrqd/8uf3475uhG6nVtzXAIAsIHgCAAAAsm/w4Kkw+fM4L/Clz+8/HOd5a1c+uLyxFm/2dvFXvtQOnoo1E1xP/tOPVuIOFA8+Hw7a7SRJzHcCkGsETwAAAECGvXJspqoIy+wmgomxqK+xZyy4Vn5mX6zB4Nc/+PNYHUvat6dd/NXnYwVH+258pINX3o8/U+q3wigzsAieAOQawRMAAACQbdUoD96rPZGDoGNHxzeiPkeSbi432nG7ncZ+79sTcQaKd5bY/aAd60UlFSthOzioKKFVPe5rAUAWEDwBAAAA2VYd9IGHg0OxXuCrX9gfuXso3NxYvdVqxOpYKhw/ulI4Hq9haagldoelwnQYJe1anjwZtuK8FgBkBcETAAAAkG2nBn3gk8HkoAOzHzgwXliJM1T81od/fXfz/lrk56lYWB37/W/HSp2GXmL3zzdXVNCBCE9ZjPtaAJAVBE8AAABARr1ybGY6yuP3BfvvRn2N557dHzl1GqbbqfidFwJrS+yORlpiJ0kLcV8PALKC4AkAAADIrkjB037t24z6Asef2nsk6nPidjsFRw9fLn7ri/siP1HS0z/+QfwldkelwjfCqM9dnjwZMlgcQO4RPAEAAADZVYry4H3B+FSUx8ddZnf70t/ECoDGvvvNWDvgTfxiaW3fjY/iLbEbl8Ze3ozTKTUf6/UAIGMIngAAAIDsqqZ58GNPRd/Nbu3KB5fj7GRX/JUvtYOnoudVe++saPLSX0fu5Ooae3lzReOK+sJtscwOACQRPAEAAABQvB3tpp4ej9yBdP2DP4/etbRvT7v4q89HTp0KG+t6+oM/WwnCzSgDwR8o/ka4EmOukyTNT54Mo6drAJBBBE8AAABAdg28o11UhUCrkwfHIj3nbvvnH8bpdhr7rRf3xBkofrRZvzp291asJXaFF8LVwgthnOe2xTI7AHiA4AkAAABAZEcOja1Efc6Nn/7l8ajPKRw/ulL40rORO5YOXXl/9cDVnz0Z9XlSZ5h48TfC/XGeK7qdAOARBE8AAAAA9GQweTnK4z9zeM/eKI/fvL/WXr+2HK2oQuFO8Xe+EbnraO+dFX2mWY8VHAVHpbE/2FyN81xJFydPhrWYzwWATCJ4AgAAADLolWMz01EeHyhYi/L4Z5/cG2lW040f/8dIQZUkFb/53L3gcLRmp8LGup75h38XZxc6BYdN6FRU3G6nMzGfBwCZRfAEAAAAZNNkmgd/anLwmUvh5sbq6ifvRwtz9u1pF0/+UuSB4s+8f36lcH89+vZ341Lx5c2rQ4ROb0+eDOsxnwsAmUXwBAAAACCS/XsLkZblrX/60xub9yM1VGns974dOTw62qyv7L3zafSB4OPS2B9uXg0OKdZMKHWW2NHtBAB9EDwBAAAAiGTy4NjdKI+/0bwQaVle4fjRlcLxaPnRoSvvrx688r6N0KktaTbmcwEg8wieAAAAAERy6EBx4HlNm/fX2vduRmiQijFQfO+dFT3Z+qswynMkJRE6SdLc5MmwNcTzASDTCJ4AAACAbGqldeAog8VvLTc2oxw76kDx7jDxINyMNoU8mdDp1cmT4eIQzweAzCN4AgAAADLonUvnW1Eef1f3Dg/62D1jwcDHvfPx3x8Z+MExBoo/+3d/ejXyMPFkQqdzkyfDhSGeDwC5MGa7AAAAAAD23QxvDRwQTR4c7GPExtqNlY216wMvmxv7rRcH3ypP0tM/+fdXx9ZvRgqPgqPS2B9srqk4dOg0N8TzASA36HgCAAAAsuti0gcsBFod9LG3f/5fioM+Njh6+HLhS88OvFxu4hdLaweu/ixO6LSqovZFed4WhE4AEAHBEwAAAJBd15M+4IF9xU8GfWyUZXbFl74+8NyoA1d/dufIhz+KFB4VngvvmNBpf5TnbUHoBAARETwBAAAA2bU06ANvhDcTfeGN9ZuXN9YGy70Kz33ucuH4YCvy9t5Z0VM//Q+DD5mSVHghXC1+NzxA6AQAo8eMJwAAACC7WkkfcHxPMNDN6/VPfzrwMrviS788ULdTYWNdT3/wZytBuDnw3Kjib4QrhRfCgR+/jVcZJA4A8dDxBAAAAGTXwB1PknRX9y7v9pgjh/ZsDnKsOx//14HCnsJzn7scHB5stNPTP/7BytjdW4OHTt8Nrw4ZOrVF6AQAQyF4AgAAADLqnUvn61EevxFu3E/qtdevLe/+oELhzqDdTkc+/H/b+258NFiINC6NfW/zeuG5cJid65YlVQmdAGA4BE8AAABAtl0Y9IHtsJ1I8HT/9qc/H+RxxW8+d2+QbqcDV392Z+IX700M9OLj0tgfbl4NntLkQI/v74Kk6cmTYaSOMQDA4wieAAAAgGwbODy5rdWnk3jBtU9/svss2WJhtfirz+8aJkUZJh4clcb+h82V4JCG6XR6c/JkWJ08GSa+IyAA5BHBEwAAAJBti4M+8HZ4e5hd3x5YW/lg1+VzhdJnb2h8z86PeThMfNe6gqPS2B9srgYHFXemU1vSS5Mnw1rM5wMA+iB4AgAAADLMzHlqD/LY2+GdRF5zkPlOg8x2OtqsXx1kmHg3dFJRcYOz85JKkyfDesznAwC2QfAEAAAAZF990AeuhesDTAXf3v3Vq7s+f5Cd7A5deX/1wNWf7bpkrvBCuDr2LzbvxAyd2pJenzwZzrK0DgDSQfAEAAAAZN/Ay+1u6ubu85l2cPf6pV2fv1u30947K3qy9VfhbscpvBCuFn8j3K+Cdp9Q/rjuAPH5GM8FAAyI4AkAAADIvoGDpxvhrWEGc+vu9Q93fH5w9PCO3U49c512DJMehE7RdbucqpMnw1aM5wMAIiB4AgAAADLunUvnr0s6N8hj2+GNoQaM31+9uuPziy99fcdupyf/6Ucru811KrwYrsUMnbqznOhyAoARIXgCAAAA8mGgrqf1cF0b2ri23Z/fu7+541K6HQeL79vTLhzfPlM6cPVndw5eeX/H0Kn43fBq8Tvhvp0e08eypJeZ5QQAo0fwBAAAAOTAO5fOL6oTwOzqRnjr1nZ/9umN+/e3+7PN++vbBlaSNPbrL2z73MLGup766V/s+Pmk+N3wauG5MOpSwLfVmeU08HJDAEByCJ4AAACA/FgY5EGfbF45Hufgm/du39j2D4uF1cKXj31muz9+5v3zK0G4sW0nU4zQaVnSS5MnwzN0OQGAPQRPAAAAQH4MNNvoRrh9frSxsf0yt512tCt84akVje/p+2cTv1ha23vn022X2MUInbpdTvUIzwEApIDgCQAAAMiJQYeM39eGboa3Puz3Z6t3N7cdDr55f337pXTf+lLfLqqx9ZuavPTXm9s9L2LoRJcTADiG4AkAAADIl9ogD7oWXj8Y9cDr11pTff9gfM+17YaKP/3jP1sJws0D/f4sYuhElxMAOIjgCQAAAMiRdy6db2mArqeV8OqR7f7syvV7fX8/vL/W9/cLX3p2o9/vH7ry/up2S+wihE50OQGAwwieAAAAgPyp7faA9XBdd3V3pd+f3d0II71Y8de+/Fi4VNhY15OtHwZ9Hz946ESXEwA4juAJAAAAyJlBu54+3PzoiX6//8n1e8v9fn/92uO/HRzavxIcfnwl3dGfvXu53y52A4ZObdHlBABeIHgCAAAA8qm22wOubl7d3+/3r928N/DniMLzn39smd2+Gx/pwLXmY0PKCy+EqwOETucllehyAgA/EDwBAAAAOWS6nt7c6THb7W53Z21z76CvU3jh+GMB09Gf/cfHlvAVXghXi78R9g26jLakVydPhrN0OQGAPwieAAAAgPyaVyfQ2dZHmx8f3/p7q3c3HwuT+hrfcy14auKR3zp05f3Vsbu3Hpn5VHguvLNL6HRRUnXyZLgw0OsCAJxB8AQAAADk1DuXzl+XdGanx1wNr2lDG9e2/v52O9v1Kjz75K1H/rvPQPHgqFT87bDvkHHj7cmT4fTkyXBp1xcEADiH4AkAAADIsXcunV+QdGGnx3y8+cljnxs+vXn/sTBqq8IXn32ks2nio/fajwwUH5fG/mBzTUX163ZqS3p58mS4YzAGAHAbwRMAAACAuZ3+8HJ4ZWLr7620793q99hewdRTDwKlwsa6Dn988ZHZUGN/uHlVRT22s50eLq1b3O01AABuI3gCAAAAcm63QePr4bquhe1Hhoyv3Lh3cMeDju+5Fhw+8OA/O91Omw+CqOJ3w6vBIfXbwe6cOqETS+sAIAMIngAAAADonUvna+p0GvX10eYvHhkyfu9+eOT22sa2xwuePPSgI2prt1PhhXC18FzYL3R6c/JkOMeudQCQHQRPAAAAALrmtM0udzfCm7oTrl7u/b1Prt9b3e5AhWcmN7v//sTVn612u52Cw1KxGoZbHt6W9OrkybAWs24AgKMIngAAAABIkt65dH5JUm27P29uLn/2kf/+eG1lu8cGx5+a6v77kQ9/dLf778V/vrmigg70PLStztK6hRglAwAcR/AEAAAA4IF3Lp2fl3S+359t7Xq6dvP+I7vWFfdNPvj3YKKTLe29s6LC/fUJSSq8GK4FR9X7nG7oxDwnAMgogicAAAAAW81JWu73B71dT5uh9v/80wfNTBrb/3Dzu+Cpzr8fvvx3na6ocan4a48ssbsoaZrQCQCyjeAJAAAAwCPeuXT+uqTZfn+2tevp5yvrP+/+ezC277HHH7j606IkFf/78LKK6u5q1+10aiVYNgDAQQRPAAAAAB5j5j292u/PeruePvr07oN5TeNHSp0uqWJhVZLG1m+qsHH3iMalQjnsPqcbOrFzHQDkAMETAAAAgL7euXR+QdK5rb9/I7yp2+GdFUm6dz880l1ut/fwswVJCg7t/0SS9t+4tCpJha+F13qePsvyOgDID4InAAAAANt659L5OUkXtv7+P27+5MGQ8OYv1j6UpOK+w3t7H7PnzqefSFLhl8MN81tvTp4M62nVCgBwD8ETAAAAgN3MqjMM/IH1cF2fbF5ZkaTL1+4evXc/VHH80GcLY/sUXr89JUl71tsFjUvBQR2VdHHyZFgbeeUAAKsIngAAAADsqGfYeLv395c3Pzy6qXBtM9T+Syvrq5K098jUgz8fW2tvBkcfbGR3ZkTlAgAcQvAEAAAAYFfvXDrfklRVT/h0Xxtqbi6HkvS3P7t9V5L2P/Xlzi536/ckScEzuiZpmSV2AJBPBE8AAAAABmJ2uquqJ3z6ZPPK/tvhnZW798OJK9fvafzJ8pgkbX7SeUjwhG5IWrRQLgDAAQRPAAAAAAbWL3z6x82fHA0V3nnvp7cuF8cPfba4b1LhL65ek6Tg6bAgqW6lWACAdQRPAAAAACLZGj6th+v6p81Lhfbtjc9euX5PB5554Vr4809v3R8/fF/7tSmCJwDILYInAAAAAJFtDZ8+2vx43+3wzsp7P711+YnPf2sjvNLee2/f5JgkTZ4Mr1ssFQBgEcETAAAAgFi2hk//uPmTo+3b9yd+cXP8ibFg4rP3tX9C7eBjq0UCAKwieAIAAAAQmwmfpiVdXA/X9cHGTzf/y49v3T9cPnX59ifF/eFVPnMAQJ7xJgAAAABgKO9cOt9Sp/PpwqfhtQM/v/9x+MHNzx28vxxe2/h0z1N2qwMA2ETwBAAAAGBo71w6f/2dS+erkt7+6f1/Ovx3l69ubh74zhMbwZ7DtmsDANhD8AQAAAAgMe9cOn9G0stLqx+Ef/fJgeDyX5/Yd/Xtwq/brgsAYAfBEwAAAIBEvXPp/OJ93Z/6D9f+7q8+WPnv9tz+r0f/pe2aAAB2BGEY2q4BAAAAQEb9b1/6H1/79aMb//vv/uhPnrBdCwBg9AieAAAAAAAAkAqW2gEAAAAAACAVBE8AAAAAAABIBcETAAAAAAAAUkHwBAAAAAAAgFQQPAEAAAAAACAVBE8AAAAAAABIBcETAAAAAAAAUkHwBAAAAAAAgFQQPAEAAAAAACAVBE8AAAAAAABIBcETAAAAAAAAUkHwBAAAAAAAgFQQPAEAAAAAACAVY7YLAADAVZVyUJJU2uaPpyVNSmqZX/20Gs1wuz8DAAAAMi8Iw9B2DQAAjFylHFTNv279Z0nSVAovuayHAVW995+NZlh/7NEAAABABhA8AQAyrVIOptXpTiqpEy6VlE6wNKxuMFU3/1xqNMMli/UAAAAAQyN4AgBkhlkaN61OwDQt6ZTFcpJyQdKSOoHUEkv3AAAA4BOCJwCAtyrlYFKdkGnW/NPFTqakLasTQi1Kqjea4XW75QAAAADbI3jCSFXKQc12DSma5wNgNJVyMKftBzdnxfVGM5y3XUSWmKVzs+bXCcvluOCiOrM7/woAACAASURBVCHUIkvzAAAA4BqCJ4yMCRnO2q4jRW1J0yyDGUylHNSVjWVQg3iJ4dHDMWHTnDphUx66muJaVieEWiCESkalHMxLes12HUNoq/M9cYabI4OplIMzkt6yXUcP6+8hpru0JWnCZh0Jc+66rVIOFiXN2K5jCG1JC5JqnG8GY27Kv2G7jh5ll34mkB0F2wUgV+q2C0jZhKSa7SJ8UCkHs8pP6NS2/YHBV5VyMF0pB/OVctCS9J46H/4JnXY2pc7f03uVctAyf38luyV5b9J2AUOakHRaUssEuNhdyXYBW7jwPTipbIVOkpv/Py58rYcxoc57UJ3zzcBc+5rzdUMq6HjCSFXKwZKyvzTmCHd5dpazbqdzjWY4Z7sIX5i76rOSzsiNc0VbncHeLfNL2j5En9TDC7Zp89+ufJ9flDSvznI8zk8RVMpBVdK7tutISFudzqcF24W4zqHOk7cbzfCM7SIkp/5OknKh0QyrtovolbHVAW1Js9x8251Dn49eZzwE0kLwhJFysH09DW82mmHNdhGuytiHuEG8yJKn3Zk7o2fU6cywqTsvaUkJDe42/2/dnfZmZfcue3fZ1Tzfl4MzXXdZ6rbjfWoXjiwtu9hohs50H2QsFJGkV10MYSvl4Lrc7MaKy8m/Z5eYzuSm5TKcC2KRLQRPGClHTqxpa0sq0VXQX866nZYbzbBkuwiXmQ8yc7L7PXFeD4dzp/5z69C8qgvqzIJasFiDFzIw56mfc2Lu044cuFlmfbZTrwxewzk5y6ZSDhZk/yZM0uj+3oUDs56c/HlAdhA8YeQcaidNE3eT+8hht5MzSyRcYwKnmuwFL8vqDEBdsHmhZeadnZHd4G1ZnUGwCxZrcJoJC9+zXUcKLqqzFKZluxBXWex2c7L7IEPdf87eGDLvC9+3XUcKLqhzviHs7sNylyXBIFLHcHHYULddwAjM2S7AUTXbBYzYou0CXFIpB5OVclAzywjOys6Hl2V12v5LjWZYs/2Bu9EMF82Hy7I6HSg2TEk6a4aRz1mqwWlmWeKy7TpScELSEkOAd1TL2evupm67gITUbRewnUYzXFSnez5rTomh49sygZyt+Uo1S6+LHCF4gg15+DA+xQe4R5lup7wssZPYze4R5uehpU4buY27eW11OhFLLnb2NJphy9xtfFGdu8I2EEDtLKvvXRPq7II4Z7sQF5nzxahDgAsOv39kZTac6/8fddsFpOSEOuHTrO1CHGUjeDpn+yYc8oHgCSPn8MVU0mq2C3BMzXYBI5bVD6mRVMrBnFmacVb2hqWeV2fuWs3S6w+s0QyXTAfUq7J3x7s3gOLDwUN12wWk7KyZZYXHLWT89aJwPbAZlOv/H1m+hpiQ9H0zQw09TNfTqLufF0bxIqbrfbZSDuYr5aDe59e8uWacHEU9GD1mPMGKHA2YZicP5XK2k5Tzr735mtdk9+e8LWnOLFvwjrn4WpD97csvqDOE2vUPaqmrlIM8XDSdV+fnhjksxoiHarcbzdDpD15Z+DloNMPAdg07yeAg9+0wW2iLEc/4Sn3WmflerinawPzz6uy+W0++IthCxxNsqdsuYETmbBfgiJrtAiyo2y7ABnNHa0GdoNFm6HRR0rSvoZPUufPZaIazkl63XMopdZZizXMn0toyyFGaUWcpTMlyHc4wy1AujujlfDhn+T7vzPn6zfec83Um4HSlHCzx3vKQuW4Z1dc+tS5Xcz04r06AGnWXxhlJ71bKwSLfG9lB8ARb6rYLGJFTpvMjt3I420nq3EFq2S5i1HrmONneBvqcpGpWvgaNZjivzuwn28NmX5OU9+V3ddsFjEh36HjVdiEOWcjY6wyjZbuAIbVsFzCguu0CRoRNDh43qgA6ldcxX8u6OtcNXW11rs9ellRuNMOg+0vSEfP7b+vRa50Zda47qmnUidEieIIVOWudrNkuwLKa7QIsqNsuYJQq5aBkls/anOPUda7RDDO3TMgsc6vK/h3w7myOvHbE1G0XMEIT6txxnrNdiCPqI3gNXzalqNsuYEi+LBuu2y5ghKbU6bScs12II+ojeI2Ladyg6wmdTvT89pvqzNqcM7v5PvK6psN7sdEMz0gqmcd3AyjeizKC4Ak25WHJgpTjrqecdjtJObpYNMNBl+TG1/nVLM+KMOHTtEa35Gcnp9S5Qz1nu5AR8+UDa5LOmuWzuWZ+/tIOfn1YZidJvgf7vtSft/PNhDrnm5rtQmwb0ZiAetIHNDek6np4E3JZ0ouNZlgb9IagCaFqevx65yxdcX4jeIJNeXpDnbNdgCU12wVYkvnvbbN2f1HSW7Lf5SRJb+ZhmLu5cKvKjfCp+yEhNzMYzN+/C3/3o3badLnl4uu8g7rnx0+K7+9xddsFDCLHGzq8USkHC5xvUr9Bn0a4taiH14TdWZtLpjO+us2vvmGS6Yqq6tG/B96HPEbwBJvy9IZ6Om/LUnLc7ZT5i0XztW3J/m5rXefM3bFcMOHHrOzPfOqaUb7mc2T653sHp9S56M/L17mfuufHT4ovHUNZkJfVAVudFiFDPc2DJ72s13SqdZfXtdWZtdk9V8yps+lMv1/vVcpBaH49styy53qne8NnQn7MwUMfBE+wKW8X7zXbBYxYzXYBlmT6ItFcWLwrN7qcJOl8lpfXbafnTqAr4dOUOhePZ2wXMgIt2wVYdEKdD4N5HTBfT/HY3mxKkYGbKz7V71OtSTuhzmDpvIbd9RSPnei1qgkIe9//Z2PO2jylTif1g50O+9xsm8nrCBPfETzBmgxcuESVm66nPHc7KaMXiT1L696wXUuPi8rvMtbuOdS1oOetHCyRqNsuwLLugHnXvvdSZ4KhtMLeekrHxRaebT6RyWuKCCaU06HjKW80kPSxz+jhDclzu9S+rE7wtfVX7zL2E+q5vjHn3vmeP58btmCMHsETbMt0d0gfc7YLGJE52wVYlLmLxJ4dSlxZWid1Pvxlbve6qMxcq7dt17FFd4lEyXIdacncz3hMb+V06Hg9peP69n3l6/WbK12ig2rZLsAB3XmC87s+MnvSmimY9Plmruffa7s8dqHRDKt9fk1LerXncdUtz5vXw5/f0xm/wZVJBE+wrWW7gBE7k/UTpfmwedpyGTa1bBeQJNO9Vtej2+K6oJbDrsm+zPbDrg28PqGMzn0yYadvH17Tcrp3SUROpHXe4Xw2Gl79Pafc9eKb1/K0mYXh/PnGvM9Pmf+8OMyS4S2bxJza8mfX9ehA9Grc14EdBE+wrWW7gBGbkHtLY5JWG/L5bbn3IXpgWbpINK3tLs1z6rrQaIZ5vPO5kzm5F4Z0l0hkcR6QVx9eU5bZkHEbqXztPXzvaNkuIEeWbRfgkBllu6N2qzTON+2E58n1nvuH2ilvy/tIv67K3uPn5T0nMwieYFvddgEWZLbrKaFup3n5u2NOZi4OTeh01nYdfbSV76WcfZnuLxfDuO48oDnbhSSsZbsAx0wpuyHjVml8EPTxZkvLdgExtWwXEEPLdgGOyVPYncb5Juljlnr+vT7A46uVclDr82thy/NbfZ7bW3s1WpmwjeAJtrVsF2DBhDq7M2RRLYFjLCRwDFtatgtIgtm5zsXQSZLmfdn5adQazbAmdz/Ans1Y+NSyXYCDuiFjzXYhaUrp/JPGMdFfy3YBMdBh+bgJdXZSnbNdSMp8CJ6i3kw/pc5GNVt/ndbDDvu2+qwQ4frPbwRPsCrHJ5Ca7QKSllC30znPvyfqtgsYlrnj5NLOdb2WTbiC7bm8lDdL4RMfBLf3Rg52Nkx6sLaP308t2wXkiK9d4KNwNsubHKQ0U7CV8PGS7DxrSzonqZT3zWOyiOAJLnD1Dn2apjL0Aayr5sgxbPL6TdJcvLk8GL5muwDXmTkx523XsYOshE9e/6yPQHdnw6yGT0l//QmesJO67QIcd7pSDrJ8vkn6/JD08eoRH39O0kvmV2+If67RDCcbzTD3OxZnFcETXJDXk0vNdgFJodvpAR8/PEiSKuXgjNwOnZa37HaC7bnc9SRlIHzycBC0DScktTI6hyXpc31er4NsqNsuAKk4pU7YncXzTSvh46V5rTpI+NdqNMO6eR+t9fz+6d2Gxm/5+nLe9AzBE1xQt12AJVNmq/osqDlyDNu8fBM0IcBbtuvYRc12Ab4wAe4523XswvvwCQPp7mw4Z7uQhNHxhJEh6B7YCXXON1XbhSSsleTBUugmij3w23xv916v7LZJSm/wxHnTMwRPgF012wUMi26nh8zOYl5xePe6XnQ7RVezXcAAznp+dzqPy8TjmFDna12zXUiCEj3X+7ishDAEjpqQ9G7Gwu5WgsdK432r93wYZ/Okmh7OsZrZJTjsPX49xmvBIoInuKCe0HF83Mr+VAbuzNQcOYZtSQ9/TJ350L/b3SUX1GwX4BsT5Lo866nL56UR3oUFlr1RKQeLGZnDkuTXngATg+D7JJosDR1vJXisxN+3zPVG9/sz8moO8/zea9Fav8eZG90z5j/bhN/+IXhClizIz/CpZruAuBLqdjqfhW4nedbyaz781fVw61pXtel2is2HUHFCkq87oBE8RTejTthYslzHUBLubuX7aLS8eq/uwfdJdKcr5WDJ0/eXXkl+7VsJHqvXrsHRAM/v3sA9VSkH/Tqneo+7EOM1YBnBE1yQ5EVALcFjjcopjy/Cawkcw4cPx1lUl/uhk8TFRWzmbqAPYfwJ+fl19vUDrG0nJC153OmWtJbtAvLEx2WNhq912+b9+SbhoLuV4LEeMDcIu9cbp8yGNVGef12PbozyyGcD00XVvdHd3vrn8APBE6xL8iLAnPi8W/IkDwOzhLqdLmSoVbZuu4BBVcrBvDoXYz7g4mI4vvz9zWRsBhB2NiHpPc/nsCS19KmV0HFsYPnX6BB0xzelTqdlnPlDGFxvcPRW1LBvS3g11X1/MMdZ7HnofEZWSuQOwROyyJcPWr123ULUQTVHjoEIzIXXa7brGNBFLi6GtmC7gAje8GzmXct2ARlw1gThPqIDhb8D+GNC0vejduI4JKnu5XpCx3lMoxku6tEd6h7McGw0w1qjGQbmV22HY5R6Hrdgnl/Xww79izs9H24jeIIrLiR4rN51wj6p2S5gUHQ79eX8Bbj5ui3YrSKSBdsF+M50lPowZLzLp3lPLdsFZMRrlXJQ9+jrnrS67QLgBeevMTzxlqdDx1u2CxjQGT3shOx2tsYK+8yN0roehk5tSdUh64NFBE/IHPNBa8F2HTHMenThncQdo1oCx3CJD23wC/JjrlPX4u4PwQB8+nuckp/nbwznlPzb4dCHcz6yg++35GRl6LhzzGewqh5dhvuW+fseaKljpRxUK+WgLun7enjNuiyp6vGMNkgas10AkJJ5+bOcqGtCnUCnZrmOHZk36rkhD5O1bifnmTtOp2zXEQHL7JKzKOms7SIimKmUg1nTto/8OCEzh8WT9wc+APmHmVToOiGpVSkH1YSHd6clkfPNKM6tjWZ43Sybn9fD1REn1FnquKxOF1Ndj3ZxTaoTWM2qcwOq1wVJs4RO/qPjCa6oJ3kw84H13G6Pc9AZD+7AnNHwXTO+zvTwklliV7NbRWQLtgvICg+X20nSvOvnwgQv4F+XH7sPjsKEpHc9HzoeVct2ATnCB1fpRfk5jiINE+qE3XO2CxmAD+HYA41meL3RDOckvaxH39+m1Amjzkp6t+fX99VpGOgNndqSXm80QzqdMoLgCVlWs11ADBMavpsoNeaD4LDL7Jaz2Mng+B36efm1xE5i7knSfPuZm5Kf5/A4liRNi26MXmc9ncMSmeednS3bBeRIUh0vS3p8KVSeTahzvqnZLiSLGs1wsdEMS5Je1eA3wC6ax5cazZAb1RlC8ITMMhdzvt3ll5KZn5SWJLqdagnUgQGZducZ23VEtOxJ67tP6rYLiOE1z2b+xNYzF8PH96y0nHZ86DjnKIKnkUnyPZHwqa83KuXAp80tvNJohguNZjgr6YiklyS92efXy5LKjWY4bR5Pl1PGMOMJrkjr5DIv/z50T1XKwVyjGS7YLqRXgt1OCwmUg8Et2C4ghrrtArKm0QxblXJwUZ05Cz6ZV052sTEX2bOm02fYXUOz4pSkJTP3ybWghw9F8NY2c3jy7rSkaXO+adkuJgVJ7iAei3mfq4vrvFyi4wmuSOWC0ix/sn6ijaFmu4A+6HbyjJlbsHVIow/qtgvIqLrtAmI4ZT4cuSrx2UxmLsarSR/XY1MyQ8dtFwJkSc8cnrdt1+KQE+qE3a5127ZsFwAMi+AJeeDj+uApl4Yd0u20K1cHA9dsFxBT3XYBGVW3XUBMC7YL2EErjYOac+WrYghw14Q6OyK5vBQ9DpY6wbpGMzwjwu5eE5Lec+k6XARPyACCJ2SeGWTtajCwkznbBfQYdbdTfcjXGrWW7QK2Mh/QfOx2Ws5oi7sL6rYLiMmpIH5UTPhUFeFTr7cyNnSc5XpwgjnfvCTON73OVsqBjzevAScRPCEvarYLiMGJJSZ0O3nL184A1+a4ZIaZreBjCC/5eQ4fmplrxI53jzpdKQdLDAFGDqV6/jbjKaoifOr1WqUcLHK+AYZH8IRcMKGHjx+4arYLUDLdTgsJ1IEBeTzbSSJ4Spuvf79TLgTxNpgOwKr8nFeYFhfmsNCthFFrpf0CJuwuibC714w6c+ZKlusAvEbwhDxZsF1ADKdsXlgn1O3Ulp9ztnxWs13AEOq2C8g4X4Mnye/v66GYIcBVSeds1+KQ7tDxORsv7uAue0AiTHdsVdJ5y6W4pBt2V20XMgTCclhF8AQnmPbetM3Lz/Zhm0umkuh2mjcXMRgBc1Hka7eT5Hcw4oO67QKGcCrvd5zNDlRv2q7DIRPqzGGp2S4EyBITds+KsLvXhKR3PZ45yPUVrCJ4Qm6Y8MPHzpvTNj5s0e3krTnbBQyhTUiZOt8vPH2dXZaYRjOsiR2otnqjUg4WmMMCJMuE3ZxvHnU2Y5scACNB8IS8WbBdQEw1C69Jt5NnzIeu07brGILvoYjzzM+jj52fXXO2C3CBmVv4ovz+WibttDpL7wifgASZ882r4nzT63SlHHC+ASIgeEKumCGtPrYNj7TriW4nb83aLmBIBE+j4fPf80SlHPj+fZ4IM2OoKoYA9zohqWV56DiQOSZ8qorwqdcpdcJuzjfAAAiekEc12wXENDfC10pkJzu6nUbO92VIfL+MRst2AUMieDIIn/qakPSex3NYACeZ8820ON/0OqFO+FS1XQjgOoIn5I7pevJxp44zI2zpnUvgGHQ7jZDpiDthuYxh1W0XkBMt2wUMieCpR88OVD5286bpbKUc8D4EJMhcQ1dF+NSrO3Tc95t/QKoInpBXPl6MTmgEHS3mLvGwu6KdMxcnGJ2q7QLgDZ+X2kkst3uM2YFqTtLbtmtxzGuVcrDIHBYgOeZ8My3C7q3eYug4sD2CJ+RSoxnWJV2wXUcMo7ibUnPkGIjG+w/i5ucS6cvCksaq7QJc1GiGZ8QOVFvNqLMUpmS5DiBTTNj9pu06HHO6Ug6WCLuBxxE8Ic+87HpKc24F3U5eq9ouAN7wveNJykDQmhYzBPhlMQS41wlJSwwBBpLVaIY1EXZvxfkG6IPgCbnVaIaLkpZt1xFDzfFjJ3EMRGCGWg47DN425kWMSEaG/k/RwbI98/5WFeFTL4aOAykwYfeL4nzTa0qdTktukgAGwRPyrma7gBim0rhwptvJa1XbBSQgC2GIT7LwAYG7yTswO1CVRKi71VnmsADJ6tlh08cbummZkPT9Sjmo2S4EcAHBE3LN3KXx8U0yjVlPNUeOgej4AI6osrDcrmq7ANf17Hjn406uaTpdKQd15rAAyTHh07QIu7d6o1IOFjjfIO8IngBpwXYBMZwwy6sSkVC30wW6nazJQvBUt10AvJOF7/vUmR2oZsUOVFudUmcpDN9HQEJ6wm7ON486rc75hvAJuUXwBHSGjPu47KTm2LGSOAYiMhcxw4aGyJ8sLG08ZbsAn5gdqF63XYdjTqjzYbBquxAgK0zYPSfCp61OSGoRdiOvCJ6Qe+bujI873J1K4mI5wW6n+rC1IBYuYBBHFpbaiQHj0TSa4bw6O1D5eLMlLROS3q2UgzSWsAO5ZcIndrx71IQ6Yfec7UKAUSN4AjoWbBcQUxIXyjVHjoF4shI8ZaEDB6NXsl2Ab8xsw6oIn7Z6i6HjQLLM+eZlcb7pNaHOJgc124UAo0TwBEgys4l8bAmeGeaOP91OmZCVeQGZ6MDByGUleB2pnh2oGAL8qNOVcrDEHBYgOY1muCjC7n7eqJSDRc43yAuCJ+Chmu0CYqpZem6Sx0B8fPBGnnHBHhPh07ZOSFpiDguQHHO+KYnzzVYz6iy9K1muA0gdwRNgmK4nH7ecPh3nDYtup8zggzfiyEqHGeHAEMwQ4Gn52fGbpil1PgzO2i4EyIqeHe8uWC7FNYTdyAWCJ+BRPg4Zl+LNeqol8LoLCRwDwyF4QhxZmanF938CzBDgN23X4ZgJSd9n6DiQHBN2V0XYvdWEpPcYOo4sI3gCepjuHR/vxMxFWSOeULfTshkaCbtO2C4AgP8azbAmdqDq561KOVhgDguQHBN2v267DgedrZQDX2+CAzsieAIe5+MJf0LRup5c2Q0PAOAIczPhJTEEeKvT6iy9I3wCEtJohvPqhN2cbx71WqUccL5B5hA8AVuY3TeWbdcRw5lB3qQq5aCq4btk6HYC4IJTtgvIGtP5W5Wf74NpOiGpxRwWIDnmWrIqwqetTqkTdnO+QWYQPAH91WwXEMOEpLkBHldL4LWSOAYAwEFmB6ppsQPVVhPqfBics10IkBXssLmtE+qcb6q2CwGSQPAE9GHuwPh4t3fHJXTmzWvYDgG6nZCGrOyyBmRCzw5UDAF+1IQ6c1hqtgsBsoLwaVsTkt7VYDeWAacRPAHbW7BdQAxTu9yJrSXwGkkcA9iKWQaAY8wOVHMifOrnDdsFAFlizjfT4nzTz2nbBQDDIngCtjcvP9ec1/r9ZkLdTm26nZCSku0CAPRnwid2vAOQOnO+edt2HQCSRfAEbMMsM/Bxh7upSjmY7fP7tQSO7ePfB4DsYlnGiJibDi/LzxsyADzSaIZnRNgNZArBE7CzBdsFxPTIrKekup00uuCpPqLXAeC367YLyBOz62tVhE8AUmbC7pfE+QbIBIInYAeNZtiSn2vNT23ZBaOWwDHnTRcY3ELHB4CRYcc7AKPSaIZ1dcJuHzf8AdCD4AnYXc12ATHVJC+7nRANYSDyrGW7gDwyN2Wqki7YrQRA1hF2A9lA8ATswvOup2nR7ZR1LdsFwEtV2wUkpGW7gLwyO1BV5ef7IwCPmGvQqqTzlksBEBPBEzCYBdsFxLQgup2yrmW7AMCilu0C8s7sQPW67ToAZJsJu2dF2A14ieAJGIBZY+7jkoITCRxjkW4npy3ZLiAhJdsFwEst2wVAajTDebEDFYARMGE35xvAMwRPwOBqtguwpGa7AOyI4Am5ZW4KwAFmB6oXxQ5UAFJmzjcvi/MN4A2CJ2BA5gNO3nbVOGdmXMFR5uvDhReimrZdQAIYNOsYMwS4Kr42AFLWaIaL6pxvuAYCPEDwBERTs13AiNVsF4CBZKXrCaMzabuABPB97yDCJwCjwo53gD8InoAITGtvXrqe6HbyR912AQmo2i4A3qnbLgD9mSHA02IIMICUmWvVqvycxQrkBsETEF1ednir2S4AA1u0XQC8k4WldnXbBWBnZgjw27brAJBtJuyuirAbcBbBExDdgrK/npxuJ4+YVnPfvydLtgvImQnbBQzpIucoPzSa4RmxAxWAETBh95u26wDwOIInIKJGM7yu7Hc9Zf3/L4t873qasl1AXlTKQRbmO9VtF4DBmWXqL8n/gByA4xrNsCbCbsA5BE9APFkOZi6YDhr4xffgKSuBiA+ysMxuwXYBiMbsDFsV4ROAlJmw+0VxvgGcQfAExGC6nrK6jrxmuwBEZ7YV9v0CKwuBiA9KtgsY0jLhuJ/M160kdqACkDJ22ATcQvAExFezXUAKLpi70vCT711PJdsF5ETJdgFD8v37PNfMjZuqpPOWSwGQcYRPgDsInoCYzGDbrHU91WwXgKEs2C5gSCXbBeSE751lWV7qnAtmB6pZZe89FIBjesJuzjeARQRPwHAWbBeQILqdPGe+fsu26xhC1XYBOeHzLK0L7GaXHWYHKoYAA0iVCbvnJL1tuxYgrwiegCGYD/oXbNeRkJrtApAIn7tBSrYLyIlTtgsYwoLtApAsMwT4Vfk/ow6A4xrN8IwIuwErCJ6A4dVsF5CAi3Q7ZcaC/P0AN2W7gKyrlAOfl9ktm5ACGWO+rlX5e+4C4AlzvnlZnG+AkSJ4AoaUgeVNkt9dMuhhZhl4+/WslIOq7RoyrmS7gCF4+32N3ZkhwNNiCDCAlJmdgKsifAJGhuAJWVKy+No1i689LLoIsmfBdgFD8Lkjxwe+/v225ff3NQZg5ndVRfgEIGUm7C6J8w0wEgRPcEJCyz9KCRwjFhPc+Nr1VLNdAJJlPrz5OkDT12DEF1XbBcQ0b7r5kHFmCPC02IEKSFylHPi8uUTiena8O2+5FCDzCJ7giiy8Efq4DIRup+yqyc8W8qrtAjLOx8Hiy/Lz/IohmB2o3rRdB+CQUgLH4ObOFibsnhVhN5AqgicgOQvy74N+zXYBSIfHs56mKuWgZLuILPJ4flaNbqd8ajTDmtiBCuhiA44UmbD7ddt1AFlF8AQkxMMP+nQ7JcfVjr15+bkEtGq7gIyq2i4ghouOn6eS+Nl39fzhBPP1f1H+3djJA753kSmNZjivTtjN+QZIGMETssSFCyCfgieXa63aLiCiE7YL6MeEoWds1xFD1XYBGVW1XUAMrn//JvGzz9KXXZghwFX5GaRnGd+7yBwTdldF+AQkiuAJrqgmcAzrH/7NB30f1oizQ1ROmC2Dr20suAAAE1dJREFUL9iuI6Kq7QKyxgyU9W2+07lGM6zbLgJuMOHTtNiBCjmU4FBwwsIBcL4BkkfwBCSvZruAAbBDVL7Mya87d1MJ7XSJh6q2C4ioLfe7nTBiPTtQ+XCDB+7x+X0lqdpdWB3gBbNDcFWET0AiCJ6QKS5sE2veqFy+KG7L7WV2XnI5KDHfkzXLZUQ1Z7uAjJm1XUBEc66H4wkOa0/qOLlgdqCak/S27VqQyC5rozRhuwD4xZxvpuX2dT3gBYInuKKa0HFc+fC/YLuAHfjQ7WQ9QIzB6ZrNwEyfltz5FpS4zqe/z/NmiSiwrUYzPCN2vLONXdZGp5TQcVy5TvaKCbvftF0H4DOCJyAFZi6Jix/yfel24sIoHXPyZ8kdy+0SUikHs/LnTn9b/nS7lRw7Tu6YIcAvy5/zGhBXKaHjOH2TzGWNZlgTYTcQG8ETXJHUB8xqQsdJQs12AX340O3kq6rtAnZjltzNWS4jijnbBWSET91Osx6do0oJHYeukSGY7riqCJ9GqlIOSrZriMPjGxpJBUalhI6TSybsfkmcb4DICJ7gCl/uxg/MdD25tvWzD91OSJH5kObLbBSfAhMnmbl3vvw9vunZLnalpA7k64d4V5gdqEpiCPAolWwXEJOvHT9JBWYE3UMy71NVuXeNDziN4AnWJTigVXKv66Rmu4Ae5zzqJPBt23fJve+9bZnZKD58QJsyy8QQny/L7M6bZQw+KTl6rFzq2fHOxWXuWeRrgOOrUlIH8rjryxkm7J6WH9dSgBMInuCCJC9eSgkea2imJdeVOyI12wVknG8X4VX50Sp+xnYBnqvZLmAAF+XnssokP7zxQTABZgeqqtiBahR8/Z71te4kO5V8u15xUk/YzfkGGADBE1yQ5EWAiy3ELixvO2fm+zjP4yUnJ2wXEEXPBZPr4dMpj78nrDLdpC6eE3u1Jc151I0p6cESxiQ7yUoJHiv3zA5Ur9uuI+N8DS+8qzvhlQGSRx3arjNh95wIn4BdETzBBdUkD5bCG/SwFmT/w33N8utHUbJdQFy+BSSmVXzOdh0DqNkuwFM12wXsoi2par4PfZN014SvXRjOajTDeXV2oLL9/ptVvn7PlmwXEEMp4eP5+rVzlgmf2PEO2AHBE1xQSvh4Tr2hmjv5NruevOl2Mry7G9nDqe+9QZhh465fLJ32LdSzzQTwrs9KO+Np6CQl3zHg3bnDB2a5e1WET2ko2S4gppLtAmJI+vxQSvh40IPzzcvifAP0RfAEq8xyhaSXgrh4AW8zeKpZfO04XPz6DcrL2s3FkuvhU812AZ6p2S5gF6+a7ztfJf2zPkG4mg4TblbFEOCkub6Mdzs+3txK+nzj1WgAn5ibeVURPgGPIXiCbdUUjunch3/T9WRj/bdv3U6SnxeFXc597w3KhABv265jB3Q9DciDbiffQycpnZ/1agrHhAifkubgSIMofAxdEj+fe/41dJo535TE+QZ4BMETbKumcMwTppPKNTULr7lg4TWH5W14I79rV6MZnpHbAzIXbBfgiZrtAnbgfehkAtA0uj28Pn+4zgwBnpbb5zhflGwXMAyfbmJUykFa54VqSseFHtnA5YLlUgBnEDzBtqpnx43NdB6N8oL3QqMZ1kf4ekkp2S5gCFM+XdD2YwZkvmm7jm2c4i7tzirlYE7udjt5HzoZVc+Oix6On+N84XtIWrJdQARVz44Lw4TdVRF2A5IInmCR6UpKq+W5mtJxhzXKWU+1Eb5WknydG9FVtV3AsBrNsCZ3Zz4tONrRaJ35e7E5T24nWQmdpPR+xk/4Hlz7wvFznA98D558qr+a0nFdvUGROSbsft12HYBtBE+wqZrisWdTPHZsZt33KNpuvex2ykg3S9V2AUlweOD4lPwNVdO2IGnCdhFbtJWt0ElK92c8zWOjh/mefEkMAY7D99DCp+BpJq0DZ+SaywuNZjgvN6+pgJEheIJNaYZDLi95qmXkNdJQsl1AAqq2C0iKwx/MXuOC+VGVcjCrFD+gxNSWVM1S6GTmraTZlenkTZOsMjdoqpKW7VbijxRnDo1SyXYBgzDn9TRxvhkh8174oty7pgJGguAJNuXyDdVc6Ka504WX3U5GFi5oXQ49I3P4g9kiS+46zPfbgt0qHnNRndBpyXYhCat6fnxsYb5Hp8UOVIPKwvu0Lx1bubxOzjJ22ESeETzBCnMXJ+0lIXMpH38Yac5hqaV47LRl4YJWytjFXM8HM5d2Z5mQtGi7CNtM+LYot5bYnVc2Qycp/feViRF0OWCLnh2ozlsuxQdV2wUkwZPOrWrKx5/y5O8hUwifkFcET7BlFBfWzg5qNe22aXSQLHvc7ST5cxdyN1XbBSStZ3eWt23X0uNUpRws2C7Csnmlt0lDHG82muGs+SCfKeb9ZBR/1wRPFphz3KzYgWo3VdsFJMTpwMUE0KPYbGVuBK+BLcz5Zlqcb5AjBE8YOXOHflQX1nMjep04ap4ccyQyNrNnJqvLwBrN8Iykl+XOjILTlXIwZ7sIGyrlYF7Sadt1GG1JL5vdwrLqzIhe53RWzx8+MDtQMQS4DxO++r7zbFfVdgG7GNV1MkG3ReZ849INPSA1BE+wYRTL7LrmRvQ6cSwq2Q/vy54P8a3aLiBhmb2YazTDRbk1E+Vs3sIn8//7mu06jAuSSub7IstG+TOd2fOHD8x7qUsBuyuqtgtIUNV2AdsZ8Q3aKZb32mVu6BF2I/MInmDDqO4aSw6/oZqlKEnOeqoleCwbqrYLSNgov89HrtEMW6ZN/E3btRi5CZ/M/+dZ23UYbzaaYTWLS+t6mb/zUXZ61Eb4WujDBKlVET71cvJ6KiaXNwKZ02jn9mX6esUHDu8iDCSG4AkjZZZTjXoeictvqPNK5k3G924nKTvznbqcnTGWJLO06kW50f2U+fCpUg7OyI3Q6aKkFzO+tK7X3Ihfbypjy4+9xI53j6naLiBhVdsFbGPU162n8nC94rqeXYQJn5BJBE8YtZqF1zzl6gW86RJIYnlKLYFjWONqV1oCXA49E9NohksOdT+dzerAcfP/9ZbtOtTpcprO6K51jzHvHzaC8ZqF18QWjWbYUufDoEu7eo6c+TlwaffMJDh37WGhu7KrZuE1sYV5Xy2JsBsZRPCEkTFbttrqaqlZet1B1IZ8fha6nZy7+EvIXJ6GBPd0P9n+gHa6Ug7qWfm7r5SDyUo5WJL9QeIXJJVz1OXUVbP0us7eNMmbnl0987wDVRbfp2dsF9BHzdLrnqbryQ3mpnRV0nnLpQCJInjCKCU5zygqZy/gzd3UYS5mF5KpxKosXtBKnbvDueh66jLdT1V1BmXabBc/Janl6s/9oEw3YEujX6Lca1mdHeuq5nyVGxa7nbpqFl8bW5gdqFzo7LQhk+/TLi3Pttjt1FWz+NroYcLuWeU77EbGEDxhJBy4eJfsBl+7iVtbe4jnOsF8sM5a+36vM1npvInCdOGV1PmQZiuAmpD0bqUczPv2NTBdTvOSvi97Px9tdb5+0znYsW47NcuvfyrDS5G9ZDr+crUDlbmGsxmIpMmln6+a5dc/7fvNmqwxYXeuzjfILoInjMqC7QLUGfbsZPeJWdMdZ3nSfAZ2k5qzXUDKctf11GXu2NXUGc5r867da+p0P81ZrGFgps6WOnXbck5SqdEMaxk4x8Ri3i9s3zCRJO+C06wzwfqLys8Q4DnbBaRoxoWfr0o5qMmNcM/rm5lZZM43trvIgaERPCF1Dr2ZSlLN4TXstYiPz0K3U0luzlhIWi67nroazbBl7tqVZS+AmlBn8PiSq3d0K+VgrlIOWursWmery+mcOnOc5vIaOEmdjjPZ7z7ompI7tcAwN4yqyvgQYPOz4FJXUBqs3hwyM1DfsFlDjxPmuh0OMeFTVYRP8BjBE1Ll2Jup1Pkwt2C7iH7MNqpRLmDpdvLHhPjg6EoAdUKd5XdOdECZJXW9gZOtkL43cGpZqsEli3JrCfBrrgameZaT8Cnry+El+9ciC5Zff6s3zPU7HGLON9PK9vkGGUbwhNSYu2QuzgU55fDdnEE7mLLQ7TSpfC1B44OjsSWAelt27uBNqdMBdb1SDhZGfZFdKQfTlXKwoM6SOluBU1udv38Cpx4OLbHbatHhjt3c6tmBKqtDgPPwPj1l60aEmeVnc/OI7SzmuVPbVeZ9uir7uwcDkQVhGNquARlVKQeLcnsZ1aumddUppvNhtw+hb/q+pbkJ/1zqhhuFZXUGNfveqZYoc3E7p84HHJvLctvqhOV1SfUkgxgTGFTNL9sdBMvqBNcLfC8+ygzy/r7tOnZwUVKVr5ubTIgQZTbbBbMLqJPMzZJ3bdcxIiN/fzZh19lRvV4MTn9/5p25eXU6wlO8/+wAvxE8IRUxToa2vGhaV50x4IVI2efuBPMhvGm5DFvOmW4f9GE+6MzJjfPHsqQl86tlfl3f6ZxhOqcm1dnRr6RO0FSSG3PuzqkTNtVtF+Ii87Wry/1lRYRPDosYJjj9wb5SDupys/svLSP7YO5B6NTFNYvDIt7EJXiCVQRPSJy5eH/Pdh0Duthohk6tYzfdHy1t/+HH+4uAGHeFs8a5wNM1PV1Qc3JzGYIvLqrT3bRIULEzzz5kO9mxiw4ToA8yJ8z14Om63A9ik9RuNMORLC/z7O/2JW5YuCtCiEnwBKvGbBeA7Gk0w6VKOTivzl1/1zk3g6rRDK+bYGa7Oxi1EZaTlnl1BiTmUYvQaXcmJJlXZyv5kjrL0+ZECDWIi+oMq130uTPSAp/m5jn33oWHGs2wbsKnuvwJF/qZU77eq0f53lyTH7sFXid0clujGS5UysGS/D/fIOPoeAIctEPXk/fdTsAwekKoWfnTnTIKF9QJIwibAEeY9/K6tg/M3240wzwM7waQMrPiZFHbL+1/vdEMfbrBgowheAIc1Wc5WludwZctOxUBbjEf6qo9v/LUDXVRZgi6OoPQWUYHOMicpxb0+GYrvKcDSNQOYXdbUolrBdhE8AQ4zOywNK9O+/ccbxjA9rYEUdPKVkfUBXXOA3URNAHeMXNYZtUJoC6q857OsmsAiTLXQrPq7BR8QtJ5STXON7CN4AkAkFmm9XzrL5dnILT1cCe9JUlLXCwCAADAZwRPAIBcMXcDpyWVtvya1GiW612UdF2dOW69v5boZAIAAEDWEDwBALBFTzjVtfW/d7OkTrj04L8JlQAAAJBHBE8AAAAAAABIRcF2AQAAAAAAAMgmgicAAAAAAACkguAJAAAAAAAAqSB4AgAAAAAAQCoIngAAAAAAAJAKgicAAAAAAACkguAJAAAAAAAAqSB4AgAAAAAAQCoIngAAAAAAAJAKgicAAADg/2/HjgUAAAAABvlbj2JfYQQALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMAiEzn2uAfq614AAAAASUVORK5CYII=" alt="Monin" style="height:48px;object-fit:contain;flex-shrink:0;filter:brightness(0) invert(1)">
  <div class="sub" style="color:rgba(255,255,255,.75);font-size:.88rem;margin-left:14px">Supplier Billback Processor</div>
</div>

<div class="container">

  <!-- CONFIG PANEL -->
  <div class="card">
    <h2>⚙️ Supplier Configuration</h2>
    <div style="display:flex;align-items:center;gap:16px;margin-bottom:10px">
      <span class="config-toggle" onclick="toggleConfig()" style="margin-bottom:0">▶ Show / Edit Supplier Settings</span>
      <button class="btn btn-secondary" onclick="saveConfig()" style="padding:5px 14px;font-size:.8rem">💾 Save Settings</button>
    </div>
    <div style="font-size:.78rem;color:#94a3b8;margin-bottom:8px">Settings are saved in your browser and restored automatically next time.</div>
    <div class="config-section" id="config-section">
      <table class="config-table">
        <thead><tr><th>Supplier</th><th>Program #</th><th>Distributor ID</th><th>Trade</th></tr></thead>
        <tbody id="config-tbody"></tbody>
      </table>
    </div>
  </div>

  <!-- FILE DROP -->
  <div class="card">
    <h2>📂 Upload Billback Files</h2>
    <div class="drop-zone" id="drop-zone" onclick="document.getElementById('file-input').click()">
      <svg width="40" height="40" fill="none" stroke="currentColor" stroke-width="1.5" viewBox="0 0 24 24"><path d="M7 16a4 4 0 01-.88-7.903A5 5 0 1115.9 6L16 6a5 5 0 011 9.9M15 13l-3-3m0 0l-3 3m3-3v12"/></svg>
      <p><strong>Click to browse</strong> or drag &amp; drop files here</p>
      <p style="margin-top:4px;font-size:.8rem">Supports: .xlsx, .xls, .txt, .tab, .pdf</p>
    </div>
    <input type="file" id="file-input" multiple accept=".xlsx,.xls,.txt,.tab,.pdf,.csv">
    <div class="file-list" id="file-list"></div>
    <div class="actions">
      <button class="btn btn-primary" id="process-btn" onclick="processFiles()" disabled>
        <span>▶ Process Files</span>
      </button>
      <button class="btn btn-secondary" onclick="clearAll()">✕ Clear All</button>
    </div>
    <div class="progress" id="progress">
      <div class="progress-bar"><div class="progress-fill" id="progress-fill"></div></div>
      <div class="progress-text" id="progress-text">Processing...</div>
    </div>
    <div class="results" id="results"></div>
  </div>

</div>

<script>
const SUPPLIERS = ['KAST','SOFO','PFS','LABATT','Y_HATA','BEK','NICH_CO','SHAMROCK','DOT_CBBB','MCLANE','S_AND_W','CHENEY','HARBOR','MARTIN_BROS','DOT_FOODS_BB','DRISCOLL','TANKERSLEY','CHRIST_PANOS','HENRY_FOODS','DELCO_FOODS','CHEFS_WH'];
const DEFAULT_CFG = {
  KAST:    {program_num:'1004089', dist_id:'134810000', trade:'D'},
  SOFO:    {program_num:'', dist_id:'', trade:'D'},
  PFS:     {program_num:'', dist_id:'', trade:'D'},
  LABATT:  {program_num:'', dist_id:'', trade:'D'},
  Y_HATA:  {program_num:'', dist_id:'', trade:'D'},
  BEK:     {program_num:'', dist_id:'', trade:'D'},
  NICH_CO: {program_num:'', dist_id:'', trade:'D'},
  SHAMROCK:{program_num:'', dist_id:'', trade:'D'},
  DOT_CBBB:{program_num:'', dist_id:'', trade:'D'},
  MCLANE:  {program_num:'', dist_id:'', trade:'D'},
  S_AND_W: {program_num:'', dist_id:'', trade:'D'},
  CHENEY:  {program_num:'', dist_id:'', trade:'D'},
  HARBOR:      {program_num:'', dist_id:'', trade:'D'},
  MARTIN_BROS:  {program_num:'', dist_id:'', trade:'D'},
  DOT_FOODS_BB: {program_num:'', dist_id:'', trade:'D'},
  DRISCOLL:     {program_num:'', dist_id:'', trade:'D'},
  TANKERSLEY:   {program_num:'', dist_id:'', trade:'D'},
  CHRIST_PANOS: {program_num:'', dist_id:'', trade:'D'},
  HENRY_FOODS:  {program_num:'', dist_id:'', trade:'D'},
  DELCO_FOODS:  {program_num:'', dist_id:'', trade:'D'},
  CHEFS_WH:     {program_num:'', dist_id:'', trade:'D'},
};

// ── Config persistence (localStorage) ────────────────────────────────────────
const SAVE_KEY = 'monin_billback_cfg_v2';

function saveConfig() {
  try { localStorage.setItem(SAVE_KEY, JSON.stringify(getConfigRaw())); } catch(e){}
  showToast('Settings saved ✓');
}

function loadSavedConfig() {
  try {
    const saved = JSON.parse(localStorage.getItem(SAVE_KEY) || 'null');
    if (!saved) return;
    SUPPLIERS.forEach(s => {
      if (!saved[s]) return;
      const d = saved[s];
      const pe = document.getElementById(`cfg_${s}_prog`);
      const de = document.getElementById(`cfg_${s}_dist`);
      const te = document.getElementById(`cfg_${s}_trade`);
      if (pe && d.program_num !== undefined) pe.value = d.program_num;
      if (de && d.dist_id !== undefined)     de.value = d.dist_id;
      if (te && d.trade !== undefined)       te.value = d.trade;
    });
    // Restore Harbor item mappings
    if (saved.HARBOR && saved.HARBOR.item_map) {
      Object.entries(saved.HARBOR.item_map).forEach(([code, mcode]) => addHarborMapping(code, mcode));
    }
  } catch(e) {}
}

function showToast(msg) {
  let t = document.getElementById('toast');
  if (!t) {
    t = document.createElement('div');
    t.id = 'toast';
    t.style.cssText = 'position:fixed;bottom:24px;right:24px;background:#1a3a5c;color:#fff;padding:10px 18px;border-radius:6px;font-size:.85rem;z-index:999;opacity:0;transition:.3s';
    document.body.appendChild(t);
  }
  t.textContent = msg;
  t.style.opacity = '1';
  setTimeout(() => { t.style.opacity = '0'; }, 2000);
}

// ── Harbor item mapping ───────────────────────────────────────────────────────
let harborMappings = []; // [{code, mcode}]

function addHarborMapping(code='', mcode='') {
  harborMappings.push({code, mcode});
  renderHarborMappings();
}

function removeHarborMapping(idx) {
  harborMappings.splice(idx, 1);
  renderHarborMappings();
}

function renderHarborMappings() {
  const container = document.getElementById('harbor-mappings');
  if (!container) return;
  container.innerHTML = harborMappings.map((m, i) => `
    <div style="display:flex;gap:6px;margin-bottom:5px;align-items:center">
      <input value="${m.code}" onchange="harborMappings[${i}].code=this.value"
        placeholder="Harbor code (e.g. 505018)" style="flex:1;font-size:.8rem;border:1px solid #cbd5e1;border-radius:4px;padding:3px 7px">
      <span style="color:#94a3b8;font-size:.9rem">→</span>
      <input value="${m.mcode}" onchange="harborMappings[${i}].mcode=this.value"
        placeholder="M-code (e.g. M-FR109F)" style="flex:1;font-size:.8rem;border:1px solid #cbd5e1;border-radius:4px;padding:3px 7px">
      <button onclick="removeHarborMapping(${i})" style="background:none;border:none;cursor:pointer;color:#94a3b8;padding:2px 5px">✕</button>
    </div>`).join('');
}

// ── Config table ─────────────────────────────────────────────────────────────
const tbody = document.getElementById('config-tbody');
SUPPLIERS.forEach(s => {
  const d = DEFAULT_CFG[s] || {program_num:'',dist_id:'',trade:'D'};
  let extra = '';
  if (s === 'HARBOR') {
    extra = `<tr id="harbor-map-row" style="display:none"><td colspan="4" style="padding:8px 10px 12px">
      <div style="font-size:.8rem;font-weight:600;color:#475569;margin-bottom:6px">
        🔗 Item Code Mapping <span style="font-weight:400;color:#94a3b8">(map Harbor codes with no M-code to Monin M-codes)</span>
      </div>
      <div id="harbor-mappings"></div>
      <button onclick="addHarborMapping()" style="font-size:.78rem;color:#2e6da4;background:none;border:1px dashed #93c5fd;border-radius:4px;padding:3px 10px;cursor:pointer;margin-top:4px">+ Add mapping</button>
    </td></tr>`;
  }
  tbody.innerHTML += `<tr id="row_${s}">
    <td><strong>${s.replace(/_/g,' ')}</strong>${s==='HARBOR'?' <span style="font-size:.7rem;color:#2e6da4;cursor:pointer;text-decoration:underline" onclick="toggleHarborMap()">item codes</span>':''}</td>
    <td><input id="cfg_${s}_prog" value="${d.program_num}" placeholder="e.g. 1004089"></td>
    <td><input id="cfg_${s}_dist" value="${d.dist_id}" placeholder="e.g. 134810000"></td>
    <td><select id="cfg_${s}_trade"><option value="D" ${d.trade==='D'?'selected':''}>D</option><option value="O" ${d.trade==='O'?'selected':''}>O</option></select></td>
  </tr>${extra}`;
});

function toggleHarborMap() {
  const row = document.getElementById('harbor-map-row');
  row.style.display = row.style.display === 'none' ? '' : 'none';
  renderHarborMappings();
}

function toggleConfig() {
  const sec = document.getElementById('config-section');
  const tog = document.querySelector('.config-toggle');
  sec.classList.toggle('open');
  tog.textContent = sec.classList.contains('open') ? '▼ Hide Supplier Settings' : '▶ Show / Edit Supplier Settings';
}

function getConfigRaw() {
  const cfg = {};
  SUPPLIERS.forEach(s => {
    cfg[s] = {
      program_num: document.getElementById(`cfg_${s}_prog`).value.trim(),
      dist_id:     document.getElementById(`cfg_${s}_dist`).value.trim(),
      trade:       document.getElementById(`cfg_${s}_trade`).value,
    };
  });
  // Add Harbor item mappings
  if (harborMappings.length) {
    const map = {};
    harborMappings.forEach(m => { if (m.code && m.mcode) map[m.code.trim()] = m.mcode.trim(); });
    cfg['HARBOR'].item_map = map;
  }
  return cfg;
}

function getConfig() { return getConfigRaw(); }

// Load saved config on startup
window.addEventListener('DOMContentLoaded', loadSavedConfig);

let selectedFiles = [];

function detectSupplier(filename) {
  const fn = filename.toUpperCase();
  if (fn.includes('KAST'))    return 'KAST';
  if (fn.includes('SOFO'))    return 'SOFO';
  if (fn.includes('PFS') || /BIDRRPT|BIDREPT|BIDBILL/.test(fn)) return 'PFS';
  if (fn.includes('BLAIR'))   return 'BLAIR_CANDY';
  if (fn.includes('LABATT'))  return 'LABATT';
  if (fn.includes('SHAMROCK')) return 'SHAMROCK';
  if (/S\s*AND\s*W|S\s*&\s*W/.test(fn)) return 'S_AND_W';
  if (fn.includes('BEK'))     return 'BEK';
  if (fn.includes('NICH'))    return 'NICH_CO';
  if (fn.includes('CBBB'))    return 'DOT_CBBB';
  if (fn.includes('TANKERSLEY')) return 'TANKERSLEY';
  if (/CHRIST.*PANOS|PANOS.*CHRIST/.test(fn)) return 'CHRIST_PANOS';
  if (/HENRY.{0,4}FOOD|PURCHASE.DETAIL/.test(fn)) return 'HENRY_FOODS';
  if (fn.includes('DELCO')) return 'DELCO_FOODS';
  if (/CHEFS.{0,6}WH|CHEFSWAREHOUSE|DAIRYLAND/.test(fn)) return 'CHEFS_WH';
  if (/ATLAS[\s_]?WHOLESALE|ATLAS[\s_]?FOOD/.test(fn)) return 'ATLAS';
  if (/Y[\s.]?HATA|Y_HATA/.test(fn)) return 'Y_HATA';
  if (fn.includes('DRISCOLL')) return 'DRISCOLL';
  if (fn.includes('KOHL')) return 'KOHL_WH';
  if (fn.includes('DELCO')) return 'DELCO_FOODS';
  if (/CHEFS.{0,6}WH|CHEFSWAREHOUSE|DAIRYLAND/.test(fn)) return 'CHEFS_WH';
  if (fn.includes('HARBOR') || fn.includes('SUPPLIER BILLBACK') || fn.includes('SUPPLIER_BILLBACK')) return 'HARBOR';
  return 'UNKNOWN';
}

// Full distributor list from Tellus (169 unique names)
const DIST_LIST = ["6 Degrees Grp","Accardi","Affiliated Foods","Alamode","Aloha","Atlas","Avalon FS","BEK","Bakemark","Balford","BiRite","Blair Candy","Brown Foodservice","CA Curtze","CASH-WA","CBCSpecialty","CCappuccino Connection","CORA","Cadillac","Cappuccino Connection","Carisam","ChefsWhse","Cheney","Cheney Bros","Cheung","Christ Panos","City Line","Clark","Clark Assoc","Country Club","Custom Food Svc","DDDI","DOT","Darden","Delco Foods","Dennis","DiCarlo","Dillanos","Dora's Natural","Dough Works","Driscoll Foods","EK Beverage","EVCO","EdwardDon","Espresso","FSA","Feesers","Fellers Foodservice","Ferraro","Fischer Foods","Fontana","Food Pro","Fred Hutch","GFS","Greco & Sons","HFM","HPC","HT Hackney","Harbor","Hardies","Harold Levinson","Henry's Foods","Hillcrest","Hillcrest Foodservice","Houstons","I Supply","IDF FD Serv Dist","IFD","Imperial Bag","Imperial Dade","Incredible","Indianhead FS","Individual FS","Iowa Des Moines","Iowa Des Moines Supply","Ital","JMC","JPoelp","Jacmar","Jakes","John Gross","John Gross & Co","Julius Silvert","KINEXO","Kaleel","Kast","Katsiroubas Bros.","Kohl Wholesale","Kuna","Labatt FS","Lineage","Lucky Goat","Lund","M&K","MAINES","MBM","Martin Bros","Martin Brower","Maximum Quality","McLane FS","McLane Grocery","Merchants","Merlino Foods","Nicholas&Co","Northern Lights","Odeko","PFD","PFG","PFS","Palmer","Paulines","Pippin","Pocono","Pon Food","Reinhart","Reliant","S&W","SF Supply","SHAHEEN BROS","SHAMROCK","SOFO","SOTF","SOTO","SPECS","SSA","SW Traders","SYSCO","SYSOCSYGMA","Saladino","Saval","Scavuzzos","Schenck","Schiffs","Sheetz","Shoreline","Six Degrees Grp","Smart&Final","Snowcap","Southeastern Food","Southern Star","Springfield Grocer","Sunbelt","Sunrise","Sutherland","Sweetwaters","Sygma","Sysco","Tankersley","Tapia","Testa","The Chefs Warehouse","Tidewater","Two Valleys","Two Valleys Dist","USCI","USF Culinary","USFS","Upper Lake","Van Eerden","Velmar","Vitco","WB Mason","Wasserstrom","West Coast Ship","White Castle","Willow Run","Win Depot","Y Hata","Youngs"];

// Map display name → internal parser key
function distNameToKey(name) {
  if (!name) return 'UNKNOWN';
  const u = name.toUpperCase().replace(/[^A-Z0-9]/g,'');
  if (u === 'KAST' || u.includes('KAST')) return 'KAST';
  if (u === 'SOFO') return 'SOFO';
  if (u === 'PFS' || u === 'PFSROMA') return 'PFS';
  if (u.includes('BLAIR')) return 'BLAIR_CANDY';
  if (u.includes('LABATT')) return 'LABATT';
  if (u.includes('SHAMROCK')) return 'SHAMROCK';
  if (u === 'SW' || u === 'SANDW' || u.includes('SANW')) return 'S_AND_W';
  if (u === 'BEK') return 'BEK';
  if (u.includes('NICHOL') || u === 'NICHCO') return 'NICH_CO';
  if (u.includes('CBBB')) return 'DOT_CBBB';
  if (u === 'YHATA' || u.includes('HATA')) return 'Y_HATA';
  if (u.includes('MCLANE')) return 'MCLANE';
  if (u.includes('MARTINBROS')) return 'MARTIN_BROS';
  if (u.includes('DRISCOLL')) return 'DRISCOLL';
  if (u.includes('HARBOR')) return 'HARBOR';
  if (u.includes('CHENEY')) return 'CHENEY';
  if (u.includes('DELCO')) return 'DELCO_FOODS';
  if (u.includes('CHEFSWAREHOUSE') || u.includes('CHEFSWHSE') || u.includes('DAIRYLAND') || u.includes('CHEFSW')) return 'CHEFS_WH';
  if (u.includes('ATLAS')) return 'ATLAS';
  if (u.includes('KOHL')) return 'KOHL_WH';
  // Everything else: route through content-based PDF dispatcher (don't force Harbor)
  return 'UNKNOWN';
}

// Map parser key → display name for auto-detect
const KEY_TO_DISPLAY = {
  KAST:'Kast', SOFO:'SOFO', PFS:'PFS', LABATT:'Labatt FS',
  Y_HATA:'Y Hata', BEK:'BEK', NICH_CO:'Nicholas&Co', SHAMROCK:'SHAMROCK',
  DOT_CBBB:'DOT', MCLANE:'McLane FS', S_AND_W:'S&W', CHENEY:'Cheney Brothers',
  HARBOR:'Harbor', MARTIN_BROS:'Martin Bros',
  DOT_FOODS_BB:'DOT', DRISCOLL:'Driscoll Foods', TANKERSLEY:'Tankersley', CHRIST_PANOS:'Christ Panos',
  HENRY_FOODS:"Henry's Foods",
  DELCO_FOODS:'Delco Foods',
  CHEFS_WH:'The Chefs Warehouse',
  ATLAS:'Atlas Wholesale Food',
  KOHL_WH:'Kohl Wholesale',
  BLAIR_CANDY:'Blair Candy', UNKNOWN:''
};

// Track currently open combobox index
let openComboIdx = null;

function supplierCombo(idx, detectedKey) {
  const displayName = KEY_TO_DISPLAY[detectedKey] || '';
  const isUnknown = !displayName || detectedKey === 'UNKNOWN';
  return `<div class="combo-wrap" id="combo_${idx}">
    <input type="text" class="combo-input${isUnknown?' unknown':''}" id="supplier_text_${idx}"
      value="${displayName.replace(/"/g,'&quot;')}"
      placeholder="Search distributor…"
      autocomplete="off"
      oninput="filterDist(${idx})"
      onkeydown="comboKey(event,${idx})"
      onfocus="showDist(${idx})">
    <span class="combo-arrow">▼</span>
    <input type="hidden" id="supplier_${idx}" value="${detectedKey}">
    <div class="combo-list" id="dist_list_${idx}"></div>
  </div>`;
}

function buildDistItems(idx, filter) {
  const q = (filter||'').toLowerCase();
  const matches = q ? DIST_LIST.filter(n => n.toLowerCase().includes(q)) : DIST_LIST;
  if (!matches.length) return `<div class="combo-item no-match">No match — will use generic PDF parser</div>`;
  return matches.map(n =>
    `<div class="combo-item" onmousedown="selectDist(${idx},'${n.replace(/\\/g,'\\\\').replace(/'/g,"\\'")}')">${n}</div>`
  ).join('');
}

function showDist(idx) {
  if (openComboIdx !== null && openComboIdx !== idx) hideDist(openComboIdx);
  openComboIdx = idx;
  const list = document.getElementById('dist_list_'+idx);
  const inp  = document.getElementById('supplier_text_'+idx);
  list.innerHTML = buildDistItems(idx, inp.value);
  list.style.display = 'block';
}

function hideDist(idx) {
  const list = document.getElementById('dist_list_'+idx);
  if (list) list.style.display = 'none';
  if (openComboIdx === idx) openComboIdx = null;
}

function filterDist(idx) {
  const inp  = document.getElementById('supplier_text_'+idx);
  const list = document.getElementById('dist_list_'+idx);
  list.innerHTML = buildDistItems(idx, inp.value);
  list.style.display = 'block';
  openComboIdx = idx;
  // Update hidden key as user types (in case they clear and type something new)
  if (!inp.value.trim()) {
    document.getElementById('supplier_'+idx).value = 'UNKNOWN';
    inp.className = 'combo-input unknown';
  }
}

function selectDist(idx, name) {
  const key  = distNameToKey(name);
  document.getElementById('supplier_text_'+idx).value = name;
  const hiddenEl = document.getElementById('supplier_'+idx);
  hiddenEl.value = key;
  const inp  = document.getElementById('supplier_text_'+idx);
  inp.className = 'combo-input';
  hideDist(idx);
  // Update placeholder hints from config
  const cfg = getConfig()[key] || {};
  const pe = document.getElementById('prog_'+idx);
  const de = document.getElementById('dist_'+idx);
  if (pe) pe.placeholder = cfg.program_num ? `default: ${cfg.program_num}` : 'Program # (optional)';
  if (de) de.placeholder = cfg.dist_id     ? `default: ${cfg.dist_id}`     : 'Distributor ID (optional)';
}

function comboKey(e, idx) {
  const list = document.getElementById('dist_list_'+idx);
  const items = list.querySelectorAll('.combo-item:not(.no-match)');
  const focused = list.querySelector('.combo-item.focused');
  if (e.key === 'Escape') { hideDist(idx); return; }
  if (e.key === 'ArrowDown') {
    e.preventDefault();
    const next = focused ? focused.nextElementSibling : items[0];
    if (focused) focused.classList.remove('focused');
    if (next && !next.classList.contains('no-match')) next.classList.add('focused');
    if (next) next.scrollIntoView({block:'nearest'});
    return;
  }
  if (e.key === 'ArrowUp') {
    e.preventDefault();
    const prev = focused ? focused.previousElementSibling : items[items.length-1];
    if (focused) focused.classList.remove('focused');
    if (prev && !prev.classList.contains('no-match')) prev.classList.add('focused');
    if (prev) prev.scrollIntoView({block:'nearest'});
    return;
  }
  if (e.key === 'Enter' && focused) {
    e.preventDefault();
    focused.onmousedown();
  }
}

// Close combo when clicking elsewhere
document.addEventListener('click', function(e) {
  if (openComboIdx !== null) {
    const wrap = document.getElementById('combo_'+openComboIdx);
    if (wrap && !wrap.contains(e.target)) hideDist(openComboIdx);
  }
});

function onSupplierChange(idx) {
  // kept for backward compat — actual updates now in selectDist
}

function renderFiles() {
  const list = document.getElementById('file-list');
  if (!selectedFiles.length) { list.innerHTML=''; return; }
  list.innerHTML = selectedFiles.map((f,i) => {
    const sup = detectSupplier(f.name);
    const cfg = getConfig()[sup] || {};
    const progPlaceholder = cfg.program_num ? `default: ${cfg.program_num}` : 'Program # (optional)';
    const distPlaceholder = cfg.dist_id     ? `default: ${cfg.dist_id}`     : 'Distributor ID (optional)';
    const unknownStyle = sup === 'UNKNOWN' ? 'background:#fef9c3;' : '';
    return `<div class="file-item">
      <div class="file-item-top">
        ${supplierCombo(i, sup)}
        <span class="file-name" title="${f.name}">${f.name}</span>
        <button class="remove-btn" onclick="removeFile(${i})" title="Remove">✕</button>
      </div>
      <div class="file-item-fields">
        <label title="Override Program # for this file only" id="prog_label_${i}">
          Prog #
          <input class="override-input" id="prog_${i}" placeholder="${progPlaceholder}">
        </label>
        <label title="Override Distributor ID for this file only">
          Dist ID
          <input class="override-input" id="dist_${i}" placeholder="${distPlaceholder}">
        </label>
        <label title="Customer Reference — leave blank to auto-detect from file">
          Customer Ref
          <input class="cref-input" id="cref_${i}" placeholder="auto-detect from file">
        </label>
        <label title="Program: fill Program # / Trade=D  |  Operator: fill Operator ID / Trade=O">
          Type
          <select class="override-input" id="type_${i}" onchange="onTypeChange(${i})" style="width:90px;cursor:pointer;">
            <option value="Program">Program</option>
            <option value="Operator">Operator</option>
          </select>
        </label>
        <label title="Operator name — used as Operator ID when Type = Operator" id="opname_label_${i}" style="display:none;">
          Operator Name
          <input class="override-input" id="opname_${i}" placeholder="e.g. Starbucks">
        </label>
      </div>
    </div>`;
  }).join('');
  document.getElementById('process-btn').disabled = false;
}

function onTypeChange(i) {
  const type = document.getElementById(`type_${i}`)?.value;
  const progLabel  = document.getElementById(`prog_label_${i}`);
  const opLabel    = document.getElementById(`opname_label_${i}`);
  const progInput  = document.getElementById(`prog_${i}`);
  if (type === 'Operator') {
    if (progLabel)  progLabel.style.display  = 'none';
    if (opLabel)    opLabel.style.display     = '';
    if (progInput)  progInput.value           = '';   // clear program # for operator rows
  } else {
    if (progLabel)  progLabel.style.display   = '';
    if (opLabel)    opLabel.style.display      = 'none';
  }
}

function removeFile(idx) {
  selectedFiles.splice(idx, 1);
  renderFiles();
  if (!selectedFiles.length) document.getElementById('process-btn').disabled = true;
}

function clearAll() {
  selectedFiles = [];
  renderFiles();
  document.getElementById('results').style.display = 'none';
  document.getElementById('progress').style.display = 'none';
}

document.getElementById('file-input').onchange = function(e) {
  Array.from(e.target.files).forEach(f => selectedFiles.push(f));
  renderFiles();
  this.value = '';
};

const dz = document.getElementById('drop-zone');
dz.ondragover = e => { e.preventDefault(); dz.classList.add('drag-over'); };
dz.ondragleave = () => dz.classList.remove('drag-over');
dz.ondrop = e => {
  e.preventDefault(); dz.classList.remove('drag-over');
  Array.from(e.dataTransfer.files).forEach(f => selectedFiles.push(f));
  renderFiles();
};

// Persists the last-used file overrides so retryFile() can re-use them
let _lastOverrides = {};

function friendlyError(msg) {
  if (!msg) return 'Unknown error';
  if (/image.based|scanned/i.test(msg))
    return 'Scanned PDF — text cannot be extracted automatically. Please enter this one manually.';
  if (/timed out/i.test(msg))
    return 'File took too long to process (may be too large or corrupt).';
  if (/no product rows|no monin|no M-code rows/i.test(msg))
    return 'No Monin items found in file.';
  if (/could not find header|header row/i.test(msg))
    return 'File format not recognised — try selecting the correct distributor below.';
  if (/nothing to repeat|invalid pattern|regex/i.test(msg))
    return 'Parser error — try selecting the correct distributor below.';
  if (/not a pdf|\/Root|xref/i.test(msg))
    return 'File does not appear to be a valid PDF.';
  if (/unknown supplier|unknown distributor/i.test(msg))
    return 'Distributor not recognised — select the correct one below.';
  return msg;
}

function _retrySelectHtml(filename, rIdx, currentKey) {
  const opts = SUPPLIERS.map(k =>
    `<option value="${k}" ${k===currentKey?'selected':''}>${KEY_TO_DISPLAY[k]||k}</option>`
  ).join('');
  const esc = filename.replace(/\\/g,'\\\\').replace(/'/g,"\\'");
  return `
    <div style="margin-top:10px;padding:10px 12px;background:#fef2f2;border-radius:6px;
                border-left:3px solid #f87171;display:flex;flex-wrap:wrap;gap:8px;align-items:center">
      <span style="font-size:.78rem;color:#7f1d1d;font-weight:600">↩ Try a different distributor:</span>
      <select id="retry-sel-${rIdx}"
              style="font-size:.8rem;padding:4px 6px;border:1px solid #fca5a5;border-radius:4px;background:#fff">
        <option value="">— select —</option>
        ${opts}
      </select>
      <button onclick="retryFile('${esc}',${rIdx})"
              style="font-size:.78rem;padding:4px 12px;background:#dc2626;color:#fff;
                     border:none;border-radius:4px;cursor:pointer;font-weight:600">
        Retry
      </button>
    </div>`;
}

function retryFile(filename, rIdx) {
  const sel = document.getElementById(`retry-sel-${rIdx}`);
  if (!sel || !sel.value) { alert('Please select a distributor first.'); return; }
  const key     = sel.value;
  const display = KEY_TO_DISPLAY[key] || key;

  // Update the upload-area dropdown for this file so processFiles() picks it up
  const fileIdx = selectedFiles.findIndex(f => f.name === filename);
  if (fileIdx === -1) { alert('File no longer available — please re-upload.'); return; }
  const textEl   = document.getElementById(`supplier_text_${fileIdx}`);
  const hiddenEl = document.getElementById(`supplier_${fileIdx}`);
  if (textEl)   { textEl.value = display; textEl.className = 'combo-input'; }
  if (hiddenEl) hiddenEl.value = key;

  // Re-run everything with the updated distributor
  processFiles();
}

async function processFiles() {
  if (!selectedFiles.length) return;
  const btn = document.getElementById('process-btn');
  btn.disabled = true;
  const prog = document.getElementById('progress');
  const fill = document.getElementById('progress-fill');
  const ptxt = document.getElementById('progress-text');
  const results = document.getElementById('results');
  prog.style.display = 'block';
  results.style.display = 'none';
  results.innerHTML = '';

  const config = getConfig();
  const fileOverrides = {};
  selectedFiles.forEach((f,i) => {
    const rowType    = document.getElementById(`type_${i}`)?.value || 'Program';
    const opName     = document.getElementById(`opname_${i}`)?.value.trim() || '';
    const progVal    = rowType === 'Operator' ? '' : (document.getElementById(`prog_${i}`)?.value.trim() || '');
    const supplierEl = document.getElementById(`supplier_${i}`);
    fileOverrides[f.name] = {
      supplier:         supplierEl?.value || '',
      supplier_display: document.getElementById(`supplier_text_${i}`)?.value.trim() || '',
      program_num:       progVal,
      dist_id:           document.getElementById(`dist_${i}`)?.value.trim() || '',
      customer_ref:      document.getElementById(`cref_${i}`)?.value.trim() || '',
      row_type:          rowType,
      operator_name:     opName,
    };
  });
  _lastOverrides = fileOverrides;

  const form = new FormData();
  selectedFiles.forEach(f => form.append('files', f));
  form.append('config', JSON.stringify(config));
  form.append('file_overrides', JSON.stringify(fileOverrides));

  fill.style.width = '30%';
  ptxt.textContent = 'Uploading files...';

  try {
    const resp = await fetch('/process', {method:'POST', body: form});
    fill.style.width = '70%';
    ptxt.textContent = 'Generating Tellus file...';
    const data = await resp.json();
    fill.style.width = '100%';

    let html = '';
    let totalRows = 0;
    let totalAmount = 0;
    let totalQty = 0;
    let hasErrors = false;
    let rIdx = 0;

    data.results.forEach(r => {
      const ri = rIdx++;
      const currentKey = (fileOverrides[r.file] || {}).supplier || '';
      if (r.error) {
        hasErrors = true;
        const canRetry = !/scanned|image.based|not a pdf|timed out/i.test(r.error);
        html += `<div class="result-row result-err" style="flex-direction:column;align-items:flex-start">
          <span>❌ <strong>${r.file}</strong></span>
          <span style="font-size:.82rem;color:#7f1d1d;margin-top:2px">${friendlyError(r.error)}</span>
          ${canRetry ? _retrySelectHtml(r.file, ri, currentKey) : ''}
        </div>`;
      } else if (r.rows === 0) {
        html += `<div class="result-row result-skip" style="flex-direction:column;align-items:flex-start">
          <div style="display:flex;justify-content:space-between;width:100%;flex-wrap:wrap;gap:4px">
            <span>⚠️ <strong>${r.file}</strong> <em>(${r.supplier})</em> — No Monin items found</span>
            <span class="count-badge" style="background:#fef3c7;color:#92400e">0 rows</span>
          </div>
          ${_retrySelectHtml(r.file, ri, currentKey)}
        </div>`;
      } else {
        totalRows += r.rows;
        totalAmount += r.total_amount || 0;
        totalQty += r.total_qty || 0;
        const hasWarns = r.warnings && r.warnings.length > 0;
        const warnHtml = hasWarns
          ? `<div style="margin-top:8px;padding:8px 10px;background:#fffbeb;border-radius:5px;border-left:3px solid #f59e0b">
               <div style="font-size:.78rem;font-weight:600;color:#92400e;margin-bottom:4px">
                 ⚠️ ${r.warnings.length} item${r.warnings.length>1?'s':''} missing M-code — $${r.warn_total.toFixed(2)} not included in this file
               </div>
               <div style="font-size:.75rem;color:#b45309;line-height:1.6">
                 ${r.warnings.map(w => `• ${w}`).join('<br>')}
               </div>
               <div style="font-size:.75rem;color:#92400e;margin-top:5px;font-style:italic">
                 These items need to be entered manually in Tellus. Ask your supplier to include the Monin M-code in future files, or add the mapping in Supplier Settings above.
               </div>
             </div>` : '';
        const amtStr = (r.total_amount || 0).toLocaleString('en-US',{style:'currency',currency:'USD'});
        const qtyStr = (r.total_qty || 0).toLocaleString();
        html += `<div class="result-row result-ok" style="flex-direction:column;align-items:flex-start">
          <div style="display:flex;justify-content:space-between;width:100%;flex-wrap:wrap;gap:4px">
            <span>✅ <strong>${r.file}</strong> <em>(${r.supplier})</em></span>
            <span style="display:flex;gap:6px;align-items:center;flex-wrap:wrap">
              <span class="count-badge">${r.rows} rows</span>
              <span class="count-badge" style="background:#dcfce7;color:#166534">Qty: ${qtyStr}</span>
              <span class="count-badge" style="background:#eff6ff;color:#1e40af">${amtStr}</span>
              ${hasWarns?`<span class="count-badge" style="background:#fef3c7;color:#92400e">⚠️ ${r.warnings.length} manual</span>`:''}
            </span>
          </div>${warnHtml}
        </div>`;
      }
    });

    const grandAmt = totalAmount.toLocaleString('en-US',{style:'currency',currency:'USD'});
    html = `<div class="total-bar">📋 ${totalRows} rows · ${totalQty.toLocaleString()} cases · ${grandAmt} total</div>` + html;

    if (data.download_id) {
      html += `<div class="dl-box">
        <p>✅ Tellus upload file is ready!</p>
        <a href="/download/${data.download_id}" class="btn btn-primary" style="text-decoration:none;display:inline-flex">
          ⬇️ Download Tellus_Upload_${new Date().toLocaleDateString('en-CA').replace(/-/g,'')}.xlsx
        </a>
      </div>`;
    }

    results.innerHTML = html;
    results.style.display = 'block';
  } catch(err) {
    results.innerHTML = `<div class="result-row result-err">❌ Server error: ${err.message}</div>`;
    results.style.display = 'block';
  }
  prog.style.display = 'none';
  btn.disabled = false;
}
</script>
</body>
</html>
"""

# ─── HTTP SERVER ──────────────────────────────────────────────────────────────
_downloads = {}   # id → bytes


# ── Simple session-based auth (only active when APP_PASSWORD env var is set) ──
import hashlib, secrets as _secrets
_active_sessions = set()  # set of valid session tokens

APP_PASSWORD = os.environ.get('APP_PASSWORD', '')  # set this on Render

LOGIN_PAGE = """<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Monin Billback — Login</title>
<style>
  body{margin:0;min-height:100vh;display:flex;align-items:center;justify-content:center;
       background:#f0f4ff;font-family:'Segoe UI',sans-serif}
  .card{background:#fff;border-radius:16px;padding:48px 40px;box-shadow:0 4px 24px rgba(0,0,0,.1);
        text-align:center;width:320px}
  h2{margin:0 0 8px;color:#1e3a5f;font-size:1.4rem}
  p{color:#64748b;margin:0 0 28px;font-size:.9rem}
  input{width:100%;box-sizing:border-box;padding:10px 14px;border:1px solid #cbd5e1;
        border-radius:8px;font-size:1rem;margin-bottom:14px;outline:none}
  input:focus{border-color:#3b82f6}
  button{width:100%;padding:11px;background:#1d4ed8;color:#fff;border:none;border-radius:8px;
         font-size:1rem;font-weight:600;cursor:pointer}
  button:hover{background:#1e40af}
  .err{color:#dc2626;font-size:.85rem;margin-bottom:12px}
</style></head>
<body><div class="card">
  <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAABJ4AAAH2CAYAAAAxjEZ6AAAACXBIWXMAAC4jAAAuIwF4pT92AAAgAElEQVR4nOzdXWxcZ37n+d+pokRJlkTKLbvdbqlZNe52dzudFt3dSQfV0aqcbF56gIS0gUxnFxiIxgKLvbO8c7uAy8Bgb+bCNDY3i70QNbnZIICbGmCRnkwmLqUx1ZvBbExlMnHifqlirHZbFi2p9EZSEnn2op6SSlSRrHPqnHqe55zvBxBsS1Wn/hbJU6d+5//8nyAMQwEAAAAAAABJK9guAAAAAAAAANlE8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFJB8AQAAAAAAIBUEDwBAAAAAAAgFQRPAAAAAAAASAXBEwAAAAAAAFIxZrsAAAAAANjNvyn960lJ01t//zdfnDz45KGxW93/Lv3b/6U+yroAADsLwjC0XQMAAAAA6N+U/nVJnXCp+6sbNk30e/yxo+NXv/X8wSf3jAX9/nhZUktSa23lJx9/evH//oGkpe/8/NL1FEoHAGyD4AkAAACAFa8cmyk9X3zuf/1M8OQXJFW1TcDUz7eeP7hafmbf/kEee/39/2f19kfvdR97UdKSpEVJdYIoAEgXwRMAAACAkXnl2My0pLmDwRP/4rlCuXAg2P/ZKM8vBFr9zRcn908eHGxqyPq15ZWVv/njozs85Lw6IdQiIRQAJI/gCQAAAECqXjk2U5I0K+mMpKkvFI61P1/43MDdTV17x4L2qa9PTAwaOm3eX2tf/k9/NLF5f23Xx479/rdV+OLnzkla+JV/9Uf1qLUBAPojeAIAAACQileOzVQlzUk6LUnjwbi+XPjiyhPBgZ06kPo6sK9w9be/cWS7eU59ffKf/y/du3l59wcWC6t7X/v93mV7FyTVCKAAYHgETwAAAAASZQKnmqRT3d/7THDkzpeKzwWBgoHmMvWKEzrdXG60b/zkLwfqqip8bWpl7Ldf7BeGEUABwJAIngAAAAAkwsxvmldP4CRJcZfWSfFCp4219tWP/9P/8eSgj9/zL19S8NSO5Z2XNPcr/+qPmAEFABEVbBcAAAAAwG+vHJuZfOXYzIKk97QldHq++MWrowydwnDzzpX/748HDp2CQ/tXdgmdJGlGUuvS//Ts/zxwIQAASQRPAAAAAIbwyrGZM5JaMnOcusZU1DfGTlz9THBk4BCoV5zQSZLa//Bnwcba4I1JxV/7yhODPG5s/ebE+K3L/+eV7xUXrnyvOBmpKADIscG2gwAAAACAHmanugVt6XCSOqHT18e+dnVce0caOq1fW165/dF7gw8uLxZWC88/O9DMqbH1m91/PS1p+sr3irNP/clGK1KBAJBDdDwBAAAAiOSVYzOzkpbkUOgUbm6sXf3bP420W16h9NkbGt8z0GPHb1++1vOfJyQtXflecTrK6wFAHhE8AQAAABjYK8dm5iV9X9Jjg5GGDZ0KgVbjhE6SdO2/fX9z8/5apOcUX/rlzw762D2r1+5s+a0JSXXCJwDYGUvtAAAAAOzqlWMzk5Lq6nT7PCaJ0Ok3X5zcHyd0uvPR0srqJ/8QqdspOHr4cnD4wMDB0/itj+/3+e1u+FR96k82lqK8PgDkBR1PAAAAAHb0yrGZaXWW1vUNnSTpheJXVuKGTpL07a8eCicPRr8vvnl/rd3+8V9ECp0kqfjS1wcOnSRpz1p7aps/6oZPDBwHgD4IngAAAABsy4ROdUnbBS96vvjFq08EByKHP11f/2dPrB07On4gznNX/uaPJ6IusQuOHr5cOD54uftufLTbQybU+TsCAGxB8AQAAACgr57Q6bF5Tl3PFp5Z+0xwJHan07Gj41e/fGz/vjjPvbncaN+7eTny84qVr0bqdtoyWHw7J658rzgfuRgAyDiCJwAAAACPGSR0Ohwc0lTheKzQSOrsYPet5w/GCq021m6s3PjJX25b27b27WkXvvi5SE85eOUftg4W385rV75XrEauCQAyjOAJAAAAwCMGCZ3GVNRXi1+OtsatRyHQavXrE7F2sJOkT//2T2It7Rs7+bW9UR5f2FjXnrXrnx/4Cfv0byMXBQAZRvAEAAAA4IFBQiepM0y8oCB2t9O3v3oofGJfMdZz4y6xC44evlz45an9UZ4zwHynR1/jM+Hx6z8MapGeBAAZRvAEAAAAQJL0yrGZSQ0QOj1beGZtmGHi5Wf2rcQdJh57iZ2i72QnSYeu/P2HMV7qzPUfBuxyBwAieAIAAACgwUOn8WBcXygc24z7OgfGCyvfev5g7NAq7hK7qDvZde1rX4r0pEJZl9X5OzwT+cUAIIMIngAAAABI0rykE7s96MuFL64ECmJ1K0nSd37pcOzQKe4SOylut9P7q0G4GWlpnvbpvvm3uaivBwBZRPAEAAAA5Nwrx2bmJJ3e7XFPF55aHWaJ3VeO729PHhyL9dxhltgVjh9didPtdPDK+ytRnxN8LuwGT1PXfxjMRn5RAMgYgicAAAAgx8ww8fndHjemosqFqXhb0KmzxO6Xy0/ECo6k+EvsJKn4O9+I/Nyx9Zsav3X5eNTnBROa6vnPuajPB4CsIXgCAAAA8m1Bu8x1kqRnC59rD7OLna0ldoWvTa0Eh6OvDDz0yX9rR31O8Pj/YTXyCwNAxhA8AQAAADn1yrGZmgaY6zQejOvZwjN74r7OMEvsNu+vtW+1GvE6pYqF1bFTX4sVeB3++OLeqM8JpsJrW35r4voPg2qc1weArCB4AgAAAHLolWMzJQ2481qp8IXLcQeK7x0L2l85fiD2Erurf/unE5v312I9t/idFwKNR8/LYg0Vl1Qoh7f6/HY1cgEAkCEETwAAAEA+zWuAJXbjwbieDCYj7wjX9c3nD+7ZMxZvNNTalQ8ur19bjvXc4ND+leK3vhhraeCRD390N9ZrPqV+3VXTcY4FAFlB8AQAAADkzCvHZqqSZgZ5bKnwhXjDlSRNPjG2cuzoeKxOqXBzY+3a3/+72IFX8Xe/GWuJ3YFrTRXur0fu0Cr8s1Aqql+XVClOHQCQFQRPAAAAQP7suoud1NnJ7kgwcTjui1R+6VDsgeLtf/xBGHeJXeH40ZXC8XgvPXnpP8cK2gq/pA+3+aNdZ2gBQJYRPAEAAAA58sqxmTkNGIY8W/hcO1AQedaRJD37mb2Xn9hXjPNU3bt15fLtj96L9boqFO4Uf+cb8bqdrv7szt7Vq7G6rIJjYeyQDQCyjOAJAAAAyJfaoA98pvD0ZtwXmX7uidjL5K79/WL8JXbffO5ecDjW6j4dbdXvxXle4YVwdZtldpKk6z8MmPMEILcIngAAAICceOXYzKykqUEe+2RwREUVj8R5na8c39+O2+1056OllXs3Y46V2renXTz5S7F20Dt05f3VOLOdJKlwIryxy0Mm4xwXALKA4AkAAADIjzODPvDpwlPbzSzaUSHQ6leOH4gV4ISbG2vtH/9F7CVrY7/37XjB0cZ6/J3sDkvBZxS7QwsAso7gCQAAAMiBV47NlCSdGuSxZqj48TivM/XZfbf3jAVxnjrcQPHnPnc57kDxiY/ea8fudjoZDtKedT3OsQEgCwieAAAAgHwYuNvpycKTq3Ff5Ktf2B8r/dlYu7ESe6B4sbA69rvfiNV1NLZ+U4c/vrgnznODw1KhHO76upMnw6U4xweALCB4AgAAAPJhbtAHPh0cXYnzAsPsZHft78/HXmJX/M4LgcZjZUc62vzLlSDcjDWNfMBuJwDINYInAAAAIOPMUPGBl5IdCg7GWmYXdye7e7cur6xfW47zVAVHD18ufuuL++I8d9+NS7f33fgoVuA1aLeTpHj/YwCQEQRPAAAAQPbNDvrAJ4NYG9lp4oli7G6nTy/+afyB4t/9Zqywq7Cxrqd/8uf3475uhG6nVtzXAIAsIHgCAAAAsm/w4Kkw+fM4L/Clz+8/HOd5a1c+uLyxFm/2dvFXvtQOnoo1E1xP/tOPVuIOFA8+Hw7a7SRJzHcCkGsETwAAAECGvXJspqoIy+wmgomxqK+xZyy4Vn5mX6zB4Nc/+PNYHUvat6dd/NXnYwVH+258pINX3o8/U+q3wigzsAieAOQawRMAAACQbdUoD96rPZGDoGNHxzeiPkeSbi432nG7ncZ+79sTcQaKd5bY/aAd60UlFSthOzioKKFVPe5rAUAWEDwBAAAA2VYd9IGHg0OxXuCrX9gfuXso3NxYvdVqxOpYKhw/ulI4Hq9haagldoelwnQYJe1anjwZtuK8FgBkBcETAAAAkG2nBn3gk8HkoAOzHzgwXliJM1T81od/fXfz/lrk56lYWB37/W/HSp2GXmL3zzdXVNCBCE9ZjPtaAJAVBE8AAABARr1ybGY6yuP3BfvvRn2N557dHzl1GqbbqfidFwJrS+yORlpiJ0kLcV8PALKC4AkAAADIrkjB037t24z6Asef2nsk6nPidjsFRw9fLn7ri/siP1HS0z/+QfwldkelwjfCqM9dnjwZMlgcQO4RPAEAAADZVYry4H3B+FSUx8ddZnf70t/ECoDGvvvNWDvgTfxiaW3fjY/iLbEbl8Ze3ozTKTUf6/UAIGMIngAAAIDsqqZ58GNPRd/Nbu3KB5fj7GRX/JUvtYOnoudVe++saPLSX0fu5Ooae3lzReOK+sJtscwOACQRPAEAAABQvB3tpp4ej9yBdP2DP4/etbRvT7v4q89HTp0KG+t6+oM/WwnCzSgDwR8o/ka4EmOukyTNT54Mo6drAJBBBE8AAABAdg28o11UhUCrkwfHIj3nbvvnH8bpdhr7rRf3xBkofrRZvzp291asJXaFF8LVwgthnOe2xTI7AHiA4AkAAABAZEcOja1Efc6Nn/7l8ajPKRw/ulL40rORO5YOXXl/9cDVnz0Z9XlSZ5h48TfC/XGeK7qdAOARBE8AAAAA9GQweTnK4z9zeM/eKI/fvL/WXr+2HK2oQuFO8Xe+EbnraO+dFX2mWY8VHAVHpbE/2FyN81xJFydPhrWYzwWATCJ4AgAAADLolWMz01EeHyhYi/L4Z5/cG2lW040f/8dIQZUkFb/53L3gcLRmp8LGup75h38XZxc6BYdN6FRU3G6nMzGfBwCZRfAEAAAAZNNkmgd/anLwmUvh5sbq6ifvRwtz9u1pF0/+UuSB4s+8f36lcH89+vZ341Lx5c2rQ4ROb0+eDOsxnwsAmUXwBAAAACCS/XsLkZblrX/60xub9yM1VGns974dOTw62qyv7L3zafSB4OPS2B9uXg0OKdZMKHWW2NHtBAB9EDwBAAAAiGTy4NjdKI+/0bwQaVle4fjRlcLxaPnRoSvvrx688r6N0KktaTbmcwEg8wieAAAAAERy6EBx4HlNm/fX2vduRmiQijFQfO+dFT3Z+qswynMkJRE6SdLc5MmwNcTzASDTCJ4AAACAbGqldeAog8VvLTc2oxw76kDx7jDxINyMNoU8mdDp1cmT4eIQzweAzCN4AgAAADLonUvnW1Eef1f3Dg/62D1jwcDHvfPx3x8Z+MExBoo/+3d/ejXyMPFkQqdzkyfDhSGeDwC5MGa7AAAAAAD23QxvDRwQTR4c7GPExtqNlY216wMvmxv7rRcH3ypP0tM/+fdXx9ZvRgqPgqPS2B9srqk4dOg0N8TzASA36HgCAAAAsuti0gcsBFod9LG3f/5fioM+Njh6+HLhS88OvFxu4hdLaweu/ixO6LSqovZFed4WhE4AEAHBEwAAAJBd15M+4IF9xU8GfWyUZXbFl74+8NyoA1d/dufIhz+KFB4VngvvmNBpf5TnbUHoBAARETwBAAAA2bU06ANvhDcTfeGN9ZuXN9YGy70Kz33ucuH4YCvy9t5Z0VM//Q+DD5mSVHghXC1+NzxA6AQAo8eMJwAAACC7WkkfcHxPMNDN6/VPfzrwMrviS788ULdTYWNdT3/wZytBuDnw3Kjib4QrhRfCgR+/jVcZJA4A8dDxBAAAAGTXwB1PknRX9y7v9pgjh/ZsDnKsOx//14HCnsJzn7scHB5stNPTP/7BytjdW4OHTt8Nrw4ZOrVF6AQAQyF4AgAAADLqnUvn61EevxFu3E/qtdevLe/+oELhzqDdTkc+/H/b+258NFiINC6NfW/zeuG5cJid65YlVQmdAGA4BE8AAABAtl0Y9IHtsJ1I8HT/9qc/H+RxxW8+d2+QbqcDV392Z+IX700M9OLj0tgfbl4NntLkQI/v74Kk6cmTYaSOMQDA4wieAAAAgGwbODy5rdWnk3jBtU9/svss2WJhtfirz+8aJkUZJh4clcb+h82V4JCG6XR6c/JkWJ08GSa+IyAA5BHBEwAAAJBti4M+8HZ4e5hd3x5YW/lg1+VzhdJnb2h8z86PeThMfNe6gqPS2B9srgYHFXemU1vSS5Mnw1rM5wMA+iB4AgAAADLMzHlqD/LY2+GdRF5zkPlOg8x2OtqsXx1kmHg3dFJRcYOz85JKkyfDesznAwC2QfAEAAAAZF990AeuhesDTAXf3v3Vq7s+f5Cd7A5deX/1wNWf7bpkrvBCuDr2LzbvxAyd2pJenzwZzrK0DgDSQfAEAAAAZN/Ay+1u6ubu85l2cPf6pV2fv1u30947K3qy9VfhbscpvBCuFn8j3K+Cdp9Q/rjuAPH5GM8FAAyI4AkAAADIvoGDpxvhrWEGc+vu9Q93fH5w9PCO3U49c512DJMehE7RdbucqpMnw1aM5wMAIiB4AgAAADLunUvnr0s6N8hj2+GNoQaM31+9uuPziy99fcdupyf/6Ucru811KrwYrsUMnbqznOhyAoARIXgCAAAA8mGgrqf1cF0b2ri23Z/fu7+541K6HQeL79vTLhzfPlM6cPVndw5eeX/H0Kn43fBq8Tvhvp0e08eypJeZ5QQAo0fwBAAAAOTAO5fOL6oTwOzqRnjr1nZ/9umN+/e3+7PN++vbBlaSNPbrL2z73MLGup766V/s+Pmk+N3wauG5MOpSwLfVmeU08HJDAEByCJ4AAACA/FgY5EGfbF45Hufgm/du39j2D4uF1cKXj31muz9+5v3zK0G4sW0nU4zQaVnSS5MnwzN0OQGAPQRPAAAAQH4MNNvoRrh9frSxsf0yt512tCt84akVje/p+2cTv1ha23vn022X2MUInbpdTvUIzwEApIDgCQAAAMiJQYeM39eGboa3Puz3Z6t3N7cdDr55f337pXTf+lLfLqqx9ZuavPTXm9s9L2LoRJcTADiG4AkAAADIl9ogD7oWXj8Y9cDr11pTff9gfM+17YaKP/3jP1sJws0D/f4sYuhElxMAOIjgCQAAAMiRdy6db2mArqeV8OqR7f7syvV7fX8/vL/W9/cLX3p2o9/vH7ry/up2S+wihE50OQGAwwieAAAAgPyp7faA9XBdd3V3pd+f3d0II71Y8de+/Fi4VNhY15OtHwZ9Hz946ESXEwA4juAJAAAAyJlBu54+3PzoiX6//8n1e8v9fn/92uO/HRzavxIcfnwl3dGfvXu53y52A4ZObdHlBABeIHgCAAAA8qm22wOubl7d3+/3r928N/DniMLzn39smd2+Gx/pwLXmY0PKCy+EqwOETucllehyAgA/EDwBAAAAOWS6nt7c6THb7W53Z21z76CvU3jh+GMB09Gf/cfHlvAVXghXi78R9g26jLakVydPhrN0OQGAPwieAAAAgPyaVyfQ2dZHmx8f3/p7q3c3HwuT+hrfcy14auKR3zp05f3Vsbu3Hpn5VHguvLNL6HRRUnXyZLgw0OsCAJxB8AQAAADk1DuXzl+XdGanx1wNr2lDG9e2/v52O9v1Kjz75K1H/rvPQPHgqFT87bDvkHHj7cmT4fTkyXBp1xcEADiH4AkAAADIsXcunV+QdGGnx3y8+cljnxs+vXn/sTBqq8IXn32ks2nio/fajwwUH5fG/mBzTUX163ZqS3p58mS4YzAGAHAbwRMAAACAuZ3+8HJ4ZWLr7620793q99hewdRTDwKlwsa6Dn988ZHZUGN/uHlVRT22s50eLq1b3O01AABuI3gCAAAAcm63QePr4bquhe1Hhoyv3Lh3cMeDju+5Fhw+8OA/O91Omw+CqOJ3w6vBIfXbwe6cOqETS+sAIAMIngAAAADonUvna+p0GvX10eYvHhkyfu9+eOT22sa2xwuePPSgI2prt1PhhXC18FzYL3R6c/JkOMeudQCQHQRPAAAAALrmtM0udzfCm7oTrl7u/b1Prt9b3e5AhWcmN7v//sTVn612u52Cw1KxGoZbHt6W9OrkybAWs24AgKMIngAAAABIkt65dH5JUm27P29uLn/2kf/+eG1lu8cGx5+a6v77kQ9/dLf778V/vrmigg70PLStztK6hRglAwAcR/AEAAAA4IF3Lp2fl3S+359t7Xq6dvP+I7vWFfdNPvj3YKKTLe29s6LC/fUJSSq8GK4FR9X7nG7oxDwnAMgogicAAAAAW81JWu73B71dT5uh9v/80wfNTBrb/3Dzu+Cpzr8fvvx3na6ocan4a48ssbsoaZrQCQCyjeAJAAAAwCPeuXT+uqTZfn+2tevp5yvrP+/+ezC277HHH7j606IkFf/78LKK6u5q1+10aiVYNgDAQQRPAAAAAB5j5j292u/PeruePvr07oN5TeNHSp0uqWJhVZLG1m+qsHH3iMalQjnsPqcbOrFzHQDkAMETAAAAgL7euXR+QdK5rb9/I7yp2+GdFUm6dz880l1ut/fwswVJCg7t/0SS9t+4tCpJha+F13qePsvyOgDID4InAAAAANt659L5OUkXtv7+P27+5MGQ8OYv1j6UpOK+w3t7H7PnzqefSFLhl8MN81tvTp4M62nVCgBwD8ETAAAAgN3MqjMM/IH1cF2fbF5ZkaTL1+4evXc/VHH80GcLY/sUXr89JUl71tsFjUvBQR2VdHHyZFgbeeUAAKsIngAAAADsqGfYeLv395c3Pzy6qXBtM9T+Syvrq5K098jUgz8fW2tvBkcfbGR3ZkTlAgAcQvAEAAAAYFfvXDrfklRVT/h0Xxtqbi6HkvS3P7t9V5L2P/Xlzi536/ckScEzuiZpmSV2AJBPBE8AAAAABmJ2uquqJ3z6ZPPK/tvhnZW798OJK9fvafzJ8pgkbX7SeUjwhG5IWrRQLgDAAQRPAAAAAAbWL3z6x82fHA0V3nnvp7cuF8cPfba4b1LhL65ek6Tg6bAgqW6lWACAdQRPAAAAACLZGj6th+v6p81Lhfbtjc9euX5PB5554Vr4809v3R8/fF/7tSmCJwDILYInAAAAAJFtDZ8+2vx43+3wzsp7P711+YnPf2sjvNLee2/f5JgkTZ4Mr1ssFQBgEcETAAAAgFi2hk//uPmTo+3b9yd+cXP8ibFg4rP3tX9C7eBjq0UCAKwieAIAAAAQmwmfpiVdXA/X9cHGTzf/y49v3T9cPnX59ifF/eFVPnMAQJ7xJgAAAABgKO9cOt9Sp/PpwqfhtQM/v/9x+MHNzx28vxxe2/h0z1N2qwMA2ETwBAAAAGBo71w6f/2dS+erkt7+6f1/Ovx3l69ubh74zhMbwZ7DtmsDANhD8AQAAAAgMe9cOn9G0stLqx+Ef/fJgeDyX5/Yd/Xtwq/brgsAYAfBEwAAAIBEvXPp/OJ93Z/6D9f+7q8+WPnv9tz+r0f/pe2aAAB2BGEY2q4BAAAAQEb9b1/6H1/79aMb//vv/uhPnrBdCwBg9AieAAAAAAAAkAqW2gEAAAAAACAVBE8AAAAAAABIBcETAAAAAAAAUkHwBAAAAAAAgFQQPAEAAAAAACAVBE8AAAAAAABIBcETAAAAAAAAUkHwBAAAAAAAgFQQPAEAAAAAACAVBE8AAAAAAABIBcETAAAAAAAAUkHwBAAAAAAAgFQQPAEAAAAAACAVY7YLAADAVZVyUJJU2uaPpyVNSmqZX/20Gs1wuz8DAAAAMi8Iw9B2DQAAjFylHFTNv279Z0nSVAovuayHAVW995+NZlh/7NEAAABABhA8AQAyrVIOptXpTiqpEy6VlE6wNKxuMFU3/1xqNMMli/UAAAAAQyN4AgBkhlkaN61OwDQt6ZTFcpJyQdKSOoHUEkv3AAAA4BOCJwCAtyrlYFKdkGnW/NPFTqakLasTQi1Kqjea4XW75QAAAADbI3jCSFXKQc12DSma5wNgNJVyMKftBzdnxfVGM5y3XUSWmKVzs+bXCcvluOCiOrM7/woAACAASURBVCHUIkvzAAAA4BqCJ4yMCRnO2q4jRW1J0yyDGUylHNSVjWVQg3iJ4dHDMWHTnDphUx66muJaVieEWiCESkalHMxLes12HUNoq/M9cYabI4OplIMzkt6yXUcP6+8hpru0JWnCZh0Jc+66rVIOFiXN2K5jCG1JC5JqnG8GY27Kv2G7jh5ll34mkB0F2wUgV+q2C0jZhKSa7SJ8UCkHs8pP6NS2/YHBV5VyMF0pB/OVctCS9J46H/4JnXY2pc7f03uVctAyf38luyV5b9J2AUOakHRaUssEuNhdyXYBW7jwPTipbIVOkpv/Py58rYcxoc57UJ3zzcBc+5rzdUMq6HjCSFXKwZKyvzTmCHd5dpazbqdzjWY4Z7sIX5i76rOSzsiNc0VbncHeLfNL2j5En9TDC7Zp89+ufJ9flDSvznI8zk8RVMpBVdK7tutISFudzqcF24W4zqHOk7cbzfCM7SIkp/5OknKh0QyrtovolbHVAW1Js9x8251Dn49eZzwE0kLwhJFysH09DW82mmHNdhGuytiHuEG8yJKn3Zk7o2fU6cywqTsvaUkJDe42/2/dnfZmZfcue3fZ1Tzfl4MzXXdZ6rbjfWoXjiwtu9hohs50H2QsFJGkV10MYSvl4Lrc7MaKy8m/Z5eYzuSm5TKcC2KRLQRPGClHTqxpa0sq0VXQX866nZYbzbBkuwiXmQ8yc7L7PXFeD4dzp/5z69C8qgvqzIJasFiDFzIw56mfc2Lu044cuFlmfbZTrwxewzk5y6ZSDhZk/yZM0uj+3oUDs56c/HlAdhA8YeQcaidNE3eT+8hht5MzSyRcYwKnmuwFL8vqDEBdsHmhZeadnZHd4G1ZnUGwCxZrcJoJC9+zXUcKLqqzFKZluxBXWex2c7L7IEPdf87eGDLvC9+3XUcKLqhzviHs7sNylyXBIFLHcHHYULddwAjM2S7AUTXbBYzYou0CXFIpB5OVclAzywjOys6Hl2V12v5LjWZYs/2Bu9EMF82Hy7I6HSg2TEk6a4aRz1mqwWlmWeKy7TpScELSEkOAd1TL2evupm67gITUbRewnUYzXFSnez5rTomh49sygZyt+Uo1S6+LHCF4gg15+DA+xQe4R5lup7wssZPYze4R5uehpU4buY27eW11OhFLLnb2NJphy9xtfFGdu8I2EEDtLKvvXRPq7II4Z7sQF5nzxahDgAsOv39kZTac6/8fddsFpOSEOuHTrO1CHGUjeDpn+yYc8oHgCSPn8MVU0mq2C3BMzXYBI5bVD6mRVMrBnFmacVb2hqWeV2fuWs3S6w+s0QyXTAfUq7J3x7s3gOLDwUN12wWk7KyZZYXHLWT89aJwPbAZlOv/H1m+hpiQ9H0zQw09TNfTqLufF0bxIqbrfbZSDuYr5aDe59e8uWacHEU9GD1mPMGKHA2YZicP5XK2k5Tzr735mtdk9+e8LWnOLFvwjrn4WpD97csvqDOE2vUPaqmrlIM8XDSdV+fnhjksxoiHarcbzdDpD15Z+DloNMPAdg07yeAg9+0wW2iLEc/4Sn3WmflerinawPzz6uy+W0++IthCxxNsqdsuYETmbBfgiJrtAiyo2y7ABnNHa0GdoNFm6HRR0rSvoZPUufPZaIazkl63XMopdZZizXMn0toyyFGaUWcpTMlyHc4wy1AujujlfDhn+T7vzPn6zfec83Um4HSlHCzx3vKQuW4Z1dc+tS5Xcz04r06AGnWXxhlJ71bKwSLfG9lB8ARb6rYLGJFTpvMjt3I420nq3EFq2S5i1HrmONneBvqcpGpWvgaNZjivzuwn28NmX5OU9+V3ddsFjEh36HjVdiEOWcjY6wyjZbuAIbVsFzCguu0CRoRNDh43qgA6ldcxX8u6OtcNXW11rs9ellRuNMOg+0vSEfP7b+vRa50Zda47qmnUidEieIIVOWudrNkuwLKa7QIsqNsuYJQq5aBkls/anOPUda7RDDO3TMgsc6vK/h3w7myOvHbE1G0XMEIT6txxnrNdiCPqI3gNXzalqNsuYEi+LBuu2y5ghKbU6bScs12II+ojeI2Ladyg6wmdTvT89pvqzNqcM7v5PvK6psN7sdEMz0gqmcd3AyjeizKC4Ak25WHJgpTjrqecdjtJObpYNMNBl+TG1/nVLM+KMOHTtEa35Gcnp9S5Qz1nu5AR8+UDa5LOmuWzuWZ+/tIOfn1YZidJvgf7vtSft/PNhDrnm5rtQmwb0ZiAetIHNDek6np4E3JZ0ouNZlgb9IagCaFqevx65yxdcX4jeIJNeXpDnbNdgCU12wVYkvnvbbN2f1HSW7Lf5SRJb+ZhmLu5cKvKjfCp+yEhNzMYzN+/C3/3o3badLnl4uu8g7rnx0+K7+9xddsFDCLHGzq8USkHC5xvUr9Bn0a4taiH14TdWZtLpjO+us2vvmGS6Yqq6tG/B96HPEbwBJvy9IZ6Om/LUnLc7ZT5i0XztW3J/m5rXefM3bFcMOHHrOzPfOqaUb7mc2T653sHp9S56M/L17mfuufHT4ovHUNZkJfVAVudFiFDPc2DJ72s13SqdZfXtdWZtdk9V8yps+lMv1/vVcpBaH49styy53qne8NnQn7MwUMfBE+wKW8X7zXbBYxYzXYBlmT6ItFcWLwrN7qcJOl8lpfXbafnTqAr4dOUOhePZ2wXMgIt2wVYdEKdD4N5HTBfT/HY3mxKkYGbKz7V71OtSTuhzmDpvIbd9RSPnei1qgkIe9//Z2PO2jylTif1g50O+9xsm8nrCBPfETzBmgxcuESVm66nPHc7KaMXiT1L696wXUuPi8rvMtbuOdS1oOetHCyRqNsuwLLugHnXvvdSZ4KhtMLeekrHxRaebT6RyWuKCCaU06HjKW80kPSxz+jhDclzu9S+rE7wtfVX7zL2E+q5vjHn3vmeP58btmCMHsETbMt0d0gfc7YLGJE52wVYlLmLxJ4dSlxZWid1Pvxlbve6qMxcq7dt17FFd4lEyXIdacncz3hMb+V06Hg9peP69n3l6/WbK12ig2rZLsAB3XmC87s+MnvSmimY9Plmruffa7s8dqHRDKt9fk1LerXncdUtz5vXw5/f0xm/wZVJBE+wrWW7gBE7k/UTpfmwedpyGTa1bBeQJNO9Vtej2+K6oJbDrsm+zPbDrg28PqGMzn0yYadvH17Tcrp3SUROpHXe4Xw2Gl79Pafc9eKb1/K0mYXh/PnGvM9Pmf+8OMyS4S2bxJza8mfX9ehA9Grc14EdBE+wrWW7gBGbkHtLY5JWG/L5bbn3IXpgWbpINK3tLs1z6rrQaIZ5vPO5kzm5F4Z0l0hkcR6QVx9eU5bZkHEbqXztPXzvaNkuIEeWbRfgkBllu6N2qzTON+2E58n1nvuH2ilvy/tIv67K3uPn5T0nMwieYFvddgEWZLbrKaFup3n5u2NOZi4OTeh01nYdfbSV76WcfZnuLxfDuO48oDnbhSSsZbsAx0wpuyHjVml8EPTxZkvLdgExtWwXEEPLdgGOyVPYncb5Juljlnr+vT7A46uVclDr82thy/NbfZ7bW3s1WpmwjeAJtrVsF2DBhDq7M2RRLYFjLCRwDFtatgtIgtm5zsXQSZLmfdn5adQazbAmdz/Ans1Y+NSyXYCDuiFjzXYhaUrp/JPGMdFfy3YBMdBh+bgJdXZSnbNdSMp8CJ6i3kw/pc5GNVt/ndbDDvu2+qwQ4frPbwRPsCrHJ5Ca7QKSllC30znPvyfqtgsYlrnj5NLOdb2WTbiC7bm8lDdL4RMfBLf3Rg52Nkx6sLaP308t2wXkiK9d4KNwNsubHKQ0U7CV8PGS7DxrSzonqZT3zWOyiOAJLnD1Dn2apjL0Aayr5sgxbPL6TdJcvLk8GL5muwDXmTkx523XsYOshE9e/6yPQHdnw6yGT0l//QmesJO67QIcd7pSDrJ8vkn6/JD08eoRH39O0kvmV2+If67RDCcbzTD3OxZnFcETXJDXk0vNdgFJodvpAR8/PEiSKuXgjNwOnZa37HaC7bnc9SRlIHzycBC0DScktTI6hyXpc31er4NsqNsuAKk4pU7YncXzTSvh46V5rTpI+NdqNMO6eR+t9fz+6d2Gxm/5+nLe9AzBE1xQt12AJVNmq/osqDlyDNu8fBM0IcBbtuvYRc12Ab4wAe4523XswvvwCQPp7mw4Z7uQhNHxhJEh6B7YCXXON1XbhSSsleTBUugmij3w23xv916v7LZJSm/wxHnTMwRPgF012wUMi26nh8zOYl5xePe6XnQ7RVezXcAAznp+dzqPy8TjmFDna12zXUiCEj3X+7ishDAEjpqQ9G7Gwu5WgsdK432r93wYZ/Okmh7OsZrZJTjsPX49xmvBIoInuKCe0HF83Mr+VAbuzNQcOYZtSQ9/TJ350L/b3SUX1GwX4BsT5Lo866nL56UR3oUFlr1RKQeLGZnDkuTXngATg+D7JJosDR1vJXisxN+3zPVG9/sz8moO8/zea9Fav8eZG90z5j/bhN/+IXhClizIz/CpZruAuBLqdjqfhW4nedbyaz781fVw61pXtel2is2HUHFCkq87oBE8RTejTthYslzHUBLubuX7aLS8eq/uwfdJdKcr5WDJ0/eXXkl+7VsJHqvXrsHRAM/v3sA9VSkH/Tqneo+7EOM1YBnBE1yQ5EVALcFjjcopjy/Cawkcw4cPx1lUl/uhk8TFRWzmbqAPYfwJ+fl19vUDrG0nJC153OmWtJbtAvLEx2WNhq912+b9+SbhoLuV4LEeMDcIu9cbp8yGNVGef12PbozyyGcD00XVvdHd3vrn8APBE6xL8iLAnPi8W/IkDwOzhLqdLmSoVbZuu4BBVcrBvDoXYz7g4mI4vvz9zWRsBhB2NiHpPc/nsCS19KmV0HFsYPnX6BB0xzelTqdlnPlDGFxvcPRW1LBvS3g11X1/MMdZ7HnofEZWSuQOwROyyJcPWr123ULUQTVHjoEIzIXXa7brGNBFLi6GtmC7gAje8GzmXct2ARlw1gThPqIDhb8D+GNC0vejduI4JKnu5XpCx3lMoxku6tEd6h7McGw0w1qjGQbmV22HY5R6Hrdgnl/Xww79izs9H24jeIIrLiR4rN51wj6p2S5gUHQ79eX8Bbj5ui3YrSKSBdsF+M50lPowZLzLp3lPLdsFZMRrlXJQ9+jrnrS67QLgBeevMTzxlqdDx1u2CxjQGT3shOx2tsYK+8yN0roehk5tSdUh64NFBE/IHPNBa8F2HTHMenThncQdo1oCx3CJD23wC/JjrlPX4u4PwQB8+nuckp/nbwznlPzb4dCHcz6yg++35GRl6LhzzGewqh5dhvuW+fseaKljpRxUK+WgLun7enjNuiyp6vGMNkgas10AkJJ5+bOcqGtCnUCnZrmOHZk36rkhD5O1bifnmTtOp2zXEQHL7JKzKOms7SIimKmUg1nTto/8OCEzh8WT9wc+APmHmVToOiGpVSkH1YSHd6clkfPNKM6tjWZ43Sybn9fD1REn1FnquKxOF1Ndj3ZxTaoTWM2qcwOq1wVJs4RO/qPjCa6oJ3kw84H13G6Pc9AZD+7AnNHwXTO+zvTwklliV7NbRWQLtgvICg+X20nSvOvnwgQv4F+XH7sPjsKEpHc9HzoeVct2ATnCB1fpRfk5jiINE+qE3XO2CxmAD+HYA41meL3RDOckvaxH39+m1Amjzkp6t+fX99VpGOgNndqSXm80QzqdMoLgCVlWs11ADBMavpsoNeaD4LDL7Jaz2Mng+B36efm1xE5i7knSfPuZm5Kf5/A4liRNi26MXmc9ncMSmeednS3bBeRIUh0vS3p8KVSeTahzvqnZLiSLGs1wsdEMS5Je1eA3wC6ax5cazZAb1RlC8ITMMhdzvt3ll5KZn5SWJLqdagnUgQGZducZ23VEtOxJ67tP6rYLiOE1z2b+xNYzF8PH96y0nHZ86DjnKIKnkUnyPZHwqa83KuXAp80tvNJohguNZjgr6YiklyS92efXy5LKjWY4bR5Pl1PGMOMJrkjr5DIv/z50T1XKwVyjGS7YLqRXgt1OCwmUg8Et2C4ghrrtArKm0QxblXJwUZ05Cz6ZV052sTEX2bOm02fYXUOz4pSkJTP3ybWghw9F8NY2c3jy7rSkaXO+adkuJgVJ7iAei3mfq4vrvFyi4wmuSOWC0ix/sn6ijaFmu4A+6HbyjJlbsHVIow/qtgvIqLrtAmI4ZT4cuSrx2UxmLsarSR/XY1MyQ8dtFwJkSc8cnrdt1+KQE+qE3a5127ZsFwAMi+AJeeDj+uApl4Yd0u20K1cHA9dsFxBT3XYBGVW3XUBMC7YL2EErjYOac+WrYghw14Q6OyK5vBQ9DpY6wbpGMzwjwu5eE5Lec+k6XARPyACCJ2SeGWTtajCwkznbBfQYdbdTfcjXGrWW7QK2Mh/QfOx2Ws5oi7sL6rYLiMmpIH5UTPhUFeFTr7cyNnSc5XpwgjnfvCTON73OVsqBjzevAScRPCEvarYLiMGJJSZ0O3nL184A1+a4ZIaZreBjCC/5eQ4fmplrxI53jzpdKQdLDAFGDqV6/jbjKaoifOr1WqUcLHK+AYZH8IRcMKGHjx+4arYLUDLdTgsJ1IEBeTzbSSJ4Spuvf79TLgTxNpgOwKr8nFeYFhfmsNCthFFrpf0CJuwuibC714w6c+ZKlusAvEbwhDxZsF1ADKdsXlgn1O3Ulp9ztnxWs13AEOq2C8g4X4Mnye/v66GYIcBVSeds1+KQ7tDxORsv7uAue0AiTHdsVdJ5y6W4pBt2V20XMgTCclhF8AQnmPbetM3Lz/Zhm0umkuh2mjcXMRgBc1Hka7eT5Hcw4oO67QKGcCrvd5zNDlRv2q7DIRPqzGGp2S4EyBITds+KsLvXhKR3PZ45yPUVrCJ4Qm6Y8MPHzpvTNj5s0e3krTnbBQyhTUiZOt8vPH2dXZaYRjOsiR2otnqjUg4WmMMCJMuE3ZxvHnU2Y5scACNB8IS8WbBdQEw1C69Jt5NnzIeu07brGILvoYjzzM+jj52fXXO2C3CBmVv4ovz+WibttDpL7wifgASZ882r4nzT63SlHHC+ASIgeEKumCGtPrYNj7TriW4nb83aLmBIBE+j4fPf80SlHPj+fZ4IM2OoKoYA9zohqWV56DiQOSZ8qorwqdcpdcJuzjfAAAiekEc12wXENDfC10pkJzu6nUbO92VIfL+MRst2AUMieDIIn/qakPSex3NYACeZ8820ON/0OqFO+FS1XQjgOoIn5I7pevJxp44zI2zpnUvgGHQ7jZDpiDthuYxh1W0XkBMt2wUMieCpR88OVD5286bpbKUc8D4EJMhcQ1dF+NSrO3Tc95t/QKoInpBXPl6MTmgEHS3mLvGwu6KdMxcnGJ2q7QLgDZ+X2kkst3uM2YFqTtLbtmtxzGuVcrDIHBYgOeZ8My3C7q3eYug4sD2CJ+RSoxnWJV2wXUcMo7ibUnPkGIjG+w/i5ucS6cvCksaq7QJc1GiGZ8QOVFvNqLMUpmS5DiBTTNj9pu06HHO6Ug6WCLuBxxE8Ic+87HpKc24F3U5eq9ouAN7wveNJykDQmhYzBPhlMQS41wlJSwwBBpLVaIY1EXZvxfkG6IPgCbnVaIaLkpZt1xFDzfFjJ3EMRGCGWg47DN425kWMSEaG/k/RwbI98/5WFeFTL4aOAykwYfeL4nzTa0qdTktukgAGwRPyrma7gBim0rhwptvJa1XbBSQgC2GIT7LwAYG7yTswO1CVRKi71VnmsADJ6tlh08cbummZkPT9Sjmo2S4EcAHBE3LN3KXx8U0yjVlPNUeOgej4AI6osrDcrmq7ANf17Hjn406uaTpdKQd15rAAyTHh07QIu7d6o1IOFjjfIO8IngBpwXYBMZwwy6sSkVC30wW6nazJQvBUt10AvJOF7/vUmR2oZsUOVFudUmcpDN9HQEJ6wm7ON486rc75hvAJuUXwBHSGjPu47KTm2LGSOAYiMhcxw4aGyJ8sLG08ZbsAn5gdqF63XYdjTqjzYbBquxAgK0zYPSfCp61OSGoRdiOvCJ6Qe+bujI873J1K4mI5wW6n+rC1IBYuYBBHFpbaiQHj0TSa4bw6O1D5eLMlLROS3q2UgzSWsAO5ZcIndrx71IQ6Yfec7UKAUSN4AjoWbBcQUxIXyjVHjoF4shI8ZaEDB6NXsl2Ab8xsw6oIn7Z6i6HjQLLM+eZlcb7pNaHOJgc124UAo0TwBEgys4l8bAmeGeaOP91OmZCVeQGZ6MDByGUleB2pnh2oGAL8qNOVcrDEHBYgOY1muCjC7n7eqJSDRc43yAuCJ+Chmu0CYqpZem6Sx0B8fPBGnnHBHhPh07ZOSFpiDguQHHO+KYnzzVYz6iy9K1muA0gdwRNgmK4nH7ecPh3nDYtup8zggzfiyEqHGeHAEMwQ4Gn52fGbpil1PgzO2i4EyIqeHe8uWC7FNYTdyAWCJ+BRPg4Zl+LNeqol8LoLCRwDwyF4QhxZmanF938CzBDgN23X4ZgJSd9n6DiQHBN2V0XYvdWEpPcYOo4sI3gCepjuHR/vxMxFWSOeULfTshkaCbtO2C4AgP8azbAmdqDq561KOVhgDguQHBN2v267DgedrZQDX2+CAzsieAIe5+MJf0LRup5c2Q0PAOAIczPhJTEEeKvT6iy9I3wCEtJohvPqhN2cbx71WqUccL5B5hA8AVuY3TeWbdcRw5lB3qQq5aCq4btk6HYC4IJTtgvIGtP5W5Wf74NpOiGpxRwWIDnmWrIqwqetTqkTdnO+QWYQPAH91WwXEMOEpLkBHldL4LWSOAYAwEFmB6ppsQPVVhPqfBics10IkBXssLmtE+qcb6q2CwGSQPAE9GHuwPh4t3fHJXTmzWvYDgG6nZCGrOyyBmRCzw5UDAF+1IQ6c1hqtgsBsoLwaVsTkt7VYDeWAacRPAHbW7BdQAxTu9yJrSXwGkkcA9iKWQaAY8wOVHMifOrnDdsFAFlizjfT4nzTz2nbBQDDIngCtjcvP9ec1/r9ZkLdTm26nZCSku0CAPRnwid2vAOQOnO+edt2HQCSRfAEbMMsM/Bxh7upSjmY7fP7tQSO7ePfB4DsYlnGiJibDi/LzxsyADzSaIZnRNgNZArBE7CzBdsFxPTIrKekup00uuCpPqLXAeC367YLyBOz62tVhE8AUmbC7pfE+QbIBIInYAeNZtiSn2vNT23ZBaOWwDHnTRcY3ELHB4CRYcc7AKPSaIZ1dcJuHzf8AdCD4AnYXc12ATHVJC+7nRANYSDyrGW7gDwyN2Wqki7YrQRA1hF2A9lA8ATswvOup2nR7ZR1LdsFwEtV2wUkpGW7gLwyO1BV5ef7IwCPmGvQqqTzlksBEBPBEzCYBdsFxLQgup2yrmW7AMCilu0C8s7sQPW67ToAZJsJu2dF2A14ieAJGIBZY+7jkoITCRxjkW4npy3ZLiAhJdsFwEst2wVAajTDebEDFYARMGE35xvAMwRPwOBqtguwpGa7AOyI4Am5ZW4KwAFmB6oXxQ5UAFJmzjcvi/MN4A2CJ2BA5gNO3nbVOGdmXMFR5uvDhReimrZdQAIYNOsYMwS4Kr42AFLWaIaL6pxvuAYCPEDwBERTs13AiNVsF4CBZKXrCaMzabuABPB97yDCJwCjwo53gD8InoAITGtvXrqe6HbyR912AQmo2i4A3qnbLgD9mSHA02IIMICUmWvVqvycxQrkBsETEF1ednir2S4AA1u0XQC8k4WldnXbBWBnZgjw27brAJBtJuyuirAbcBbBExDdgrK/npxuJ4+YVnPfvydLtgvImQnbBQzpIucoPzSa4RmxAxWAETBh95u26wDwOIInIKJGM7yu7Hc9Zf3/L4t873qasl1AXlTKQRbmO9VtF4DBmWXqL8n/gByA4xrNsCbCbsA5BE9APFkOZi6YDhr4xffgKSuBiA+ysMxuwXYBiMbsDFsV4ROAlJmw+0VxvgGcQfAExGC6nrK6jrxmuwBEZ7YV9v0CKwuBiA9KtgsY0jLhuJ/M160kdqACkDJ22ATcQvAExFezXUAKLpi70vCT711PJdsF5ETJdgFD8v37PNfMjZuqpPOWSwGQcYRPgDsInoCYzGDbrHU91WwXgKEs2C5gSCXbBeSE751lWV7qnAtmB6pZZe89FIBjesJuzjeARQRPwHAWbBeQILqdPGe+fsu26xhC1XYBOeHzLK0L7GaXHWYHKoYAA0iVCbvnJL1tuxYgrwiegCGYD/oXbNeRkJrtApAIn7tBSrYLyIlTtgsYwoLtApAsMwT4Vfk/ow6A4xrN8IwIuwErCJ6A4dVsF5CAi3Q7ZcaC/P0AN2W7gKyrlAOfl9ktm5ACGWO+rlX5e+4C4AlzvnlZnG+AkSJ4AoaUgeVNkt9dMuhhZhl4+/WslIOq7RoyrmS7gCF4+32N3ZkhwNNiCDCAlJmdgKsifAJGhuAJWVKy+No1i689LLoIsmfBdgFD8Lkjxwe+/v225ff3NQZg5ndVRfgEIGUm7C6J8w0wEgRPcEJCyz9KCRwjFhPc+Nr1VLNdAJJlPrz5OkDT12DEF1XbBcQ0b7r5kHFmCPC02IEKSFylHPi8uUTiena8O2+5FCDzCJ7giiy8Efq4DIRup+yqyc8W8qrtAjLOx8Hiy/Lz/IohmB2o3rRdB+CQUgLH4ObOFibsnhVhN5AqgicgOQvy74N+zXYBSIfHs56mKuWgZLuILPJ4flaNbqd8ajTDmtiBCuhiA44UmbD7ddt1AFlF8AQkxMMP+nQ7JcfVjr15+bkEtGq7gIyq2i4ghouOn6eS+Nl39fzhBPP1f1H+3djJA753kSmNZjivTtjN+QZIGMETssSFCyCfgieXa63aLiCiE7YL6MeEoWds1xFD1XYBGVW1XUAMrn//JvGzz9KXXZghwFX5GaRnGd+7yBwTdldF+AQkiuAJrqgmcAzrH/7NB30f1oizQ1ROmC2Dr20suAAAE1dJREFUL9iuI6Kq7QKyxgyU9W2+07lGM6zbLgJuMOHTtNiBCjmU4FBwwsIBcL4BkkfwBCSvZruAAbBDVL7Mya87d1MJ7XSJh6q2C4ioLfe7nTBiPTtQ+XCDB+7x+X0lqdpdWB3gBbNDcFWET0AiCJ6QKS5sE2veqFy+KG7L7WV2XnI5KDHfkzXLZUQ1Z7uAjJm1XUBEc66H4wkOa0/qOLlgdqCak/S27VqQyC5rozRhuwD4xZxvpuX2dT3gBYInuKKa0HFc+fC/YLuAHfjQ7WQ9QIzB6ZrNwEyfltz5FpS4zqe/z/NmiSiwrUYzPCN2vLONXdZGp5TQcVy5TvaKCbvftF0H4DOCJyAFZi6Jix/yfel24sIoHXPyZ8kdy+0SUikHs/LnTn9b/nS7lRw7Tu6YIcAvy5/zGhBXKaHjOH2TzGWNZlgTYTcQG8ETXJHUB8xqQsdJQs12AX340O3kq6rtAnZjltzNWS4jijnbBWSET91Osx6do0oJHYeukSGY7riqCJ9GqlIOSrZriMPjGxpJBUalhI6TSybsfkmcb4DICJ7gCl/uxg/MdD25tvWzD91OSJH5kObLbBSfAhMnmbl3vvw9vunZLnalpA7k64d4V5gdqEpiCPAolWwXEJOvHT9JBWYE3UMy71NVuXeNDziN4AnWJTigVXKv66Rmu4Ae5zzqJPBt23fJve+9bZnZKD58QJsyy8QQny/L7M6bZQw+KTl6rFzq2fHOxWXuWeRrgOOrUlIH8rjryxkm7J6WH9dSgBMInuCCJC9eSgkea2imJdeVOyI12wVknG8X4VX50Sp+xnYBnqvZLmAAF+XnssokP7zxQTABZgeqqtiBahR8/Z71te4kO5V8u15xUk/YzfkGGADBE1yQ5EWAiy3ELixvO2fm+zjP4yUnJ2wXEEXPBZPr4dMpj78nrDLdpC6eE3u1Jc151I0p6cESxiQ7yUoJHiv3zA5Ur9uuI+N8DS+8qzvhlQGSRx3arjNh95wIn4BdETzBBdUkD5bCG/SwFmT/w33N8utHUbJdQFy+BSSmVXzOdh0DqNkuwFM12wXsoi2par4PfZN014SvXRjOajTDeXV2oLL9/ptVvn7PlmwXEEMp4eP5+rVzlgmf2PEO2AHBE1xQSvh4Tr2hmjv5NruevOl2Mry7G9nDqe+9QZhh465fLJ32LdSzzQTwrs9KO+Np6CQl3zHg3bnDB2a5e1WET2ko2S4gppLtAmJI+vxQSvh40IPzzcvifAP0RfAEq8xyhaSXgrh4AW8zeKpZfO04XPz6DcrL2s3FkuvhU812AZ6p2S5gF6+a7ztfJf2zPkG4mg4TblbFEOCkub6Mdzs+3txK+nzj1WgAn5ibeVURPgGPIXiCbdUUjunch3/T9WRj/bdv3U6SnxeFXc597w3KhABv265jB3Q9DciDbiffQycpnZ/1agrHhAifkubgSIMofAxdEj+fe/41dJo535TE+QZ4BMETbKumcMwTppPKNTULr7lg4TWH5W14I79rV6MZnpHbAzIXbBfgiZrtAnbgfehkAtA0uj28Pn+4zgwBnpbb5zhflGwXMAyfbmJUykFa54VqSseFHtnA5YLlUgBnEDzBtqpnx43NdB6N8oL3QqMZ1kf4ekkp2S5gCFM+XdD2YwZkvmm7jm2c4i7tzirlYE7udjt5HzoZVc+Oix6On+N84XtIWrJdQARVz44Lw4TdVRF2A5IInmCR6UpKq+W5mtJxhzXKWU+1Eb5WknydG9FVtV3AsBrNsCZ3Zz4tONrRaJ35e7E5T24nWQmdpPR+xk/4Hlz7wvFznA98D558qr+a0nFdvUGROSbsft12HYBtBE+wqZrisWdTPHZsZt33KNpuvex2ykg3S9V2AUlweOD4lPwNVdO2IGnCdhFbtJWt0ElK92c8zWOjh/mefEkMAY7D99DCp+BpJq0DZ+SaywuNZjgvN6+pgJEheIJNaYZDLi95qmXkNdJQsl1AAqq2C0iKwx/MXuOC+VGVcjCrFD+gxNSWVM1S6GTmraTZlenkTZOsMjdoqpKW7VbijxRnDo1SyXYBgzDn9TRxvhkh8174oty7pgJGguAJNuXyDdVc6Ka504WX3U5GFi5oXQ49I3P4g9kiS+46zPfbgt0qHnNRndBpyXYhCat6fnxsYb5Hp8UOVIPKwvu0Lx1bubxOzjJ22ESeETzBCnMXJ+0lIXMpH38Yac5hqaV47LRl4YJWytjFXM8HM5d2Z5mQtGi7CNtM+LYot5bYnVc2Qycp/feViRF0OWCLnh2ozlsuxQdV2wUkwZPOrWrKx5/y5O8hUwifkFcET7BlFBfWzg5qNe22aXSQLHvc7ST5cxdyN1XbBSStZ3eWt23X0uNUpRws2C7Csnmlt0lDHG82muGs+SCfKeb9ZBR/1wRPFphz3KzYgWo3VdsFJMTpwMUE0KPYbGVuBK+BLcz5Zlqcb5AjBE8YOXOHflQX1nMjep04ap4ccyQyNrNnJqvLwBrN8Iykl+XOjILTlXIwZ7sIGyrlYF7Sadt1GG1JL5vdwrLqzIhe53RWzx8+MDtQMQS4DxO++r7zbFfVdgG7GNV1MkG3ReZ849INPSA1BE+wYRTL7LrmRvQ6cSwq2Q/vy54P8a3aLiBhmb2YazTDRbk1E+Vs3sIn8//7mu06jAuSSub7IstG+TOd2fOHD8x7qUsBuyuqtgtIUNV2AdsZ8Q3aKZb32mVu6BF2I/MInmDDqO4aSw6/oZqlKEnOeqoleCwbqrYLSNgov89HrtEMW6ZN/E3btRi5CZ/M/+dZ23UYbzaaYTWLS+t6mb/zUXZ61Eb4WujDBKlVET71cvJ6KiaXNwKZ02jn9mX6esUHDu8iDCSG4AkjZZZTjXoeictvqPNK5k3G924nKTvznbqcnTGWJLO06kW50f2U+fCpUg7OyI3Q6aKkFzO+tK7X3Ihfbypjy4+9xI53j6naLiBhVdsFbGPU162n8nC94rqeXYQJn5BJBE8YtZqF1zzl6gW86RJIYnlKLYFjWONqV1oCXA49E9NohksOdT+dzerAcfP/9ZbtOtTpcprO6K51jzHvHzaC8ZqF18QWjWbYUufDoEu7eo6c+TlwaffMJDh37WGhu7KrZuE1sYV5Xy2JsBsZRPCEkTFbttrqaqlZet1B1IZ8fha6nZy7+EvIXJ6GBPd0P9n+gHa6Ug7qWfm7r5SDyUo5WJL9QeIXJJVz1OXUVbP0us7eNMmbnl0987wDVRbfp2dsF9BHzdLrnqbryQ3mpnRV0nnLpQCJInjCKCU5zygqZy/gzd3UYS5mF5KpxKosXtBKnbvDueh66jLdT1V1BmXabBc/Janl6s/9oEw3YEujX6Lca1mdHeuq5nyVGxa7nbpqFl8bW5gdqFzo7LQhk+/TLi3Pttjt1FWz+NroYcLuWeU77EbGEDxhJBy4eJfsBl+7iVtbe4jnOsF8sM5a+36vM1npvInCdOGV1PmQZiuAmpD0bqUczPv2NTBdTvOSvi97Px9tdb5+0znYsW47NcuvfyrDS5G9ZDr+crUDlbmGsxmIpMmln6+a5dc/7fvNmqwxYXeuzjfILoInjMqC7QLUGfbsZPeJWdMdZ3nSfAZ2k5qzXUDKctf11GXu2NXUGc5r867da+p0P81ZrGFgps6WOnXbck5SqdEMaxk4x8Ri3i9s3zCRJO+C06wzwfqLys8Q4DnbBaRoxoWfr0o5qMmNcM/rm5lZZM43trvIgaERPCF1Dr2ZSlLN4TXstYiPz0K3U0luzlhIWi67nroazbBl7tqVZS+AmlBn8PiSq3d0K+VgrlIOWursWmery+mcOnOc5vIaOEmdjjPZ7z7ompI7tcAwN4yqyvgQYPOz4FJXUBqs3hwyM1DfsFlDjxPmuh0OMeFTVYRP8BjBE1Ll2Jup1Pkwt2C7iH7MNqpRLmDpdvLHhPjg6EoAdUKd5XdOdECZJXW9gZOtkL43cGpZqsEli3JrCfBrrgameZaT8Cnry+El+9ciC5Zff6s3zPU7HGLON9PK9vkGGUbwhNSYu2QuzgU55fDdnEE7mLLQ7TSpfC1B44OjsSWAelt27uBNqdMBdb1SDhZGfZFdKQfTlXKwoM6SOluBU1udv38Cpx4OLbHbatHhjt3c6tmBKqtDgPPwPj1l60aEmeVnc/OI7SzmuVPbVeZ9uir7uwcDkQVhGNquARlVKQeLcnsZ1aumddUppvNhtw+hb/q+pbkJ/1zqhhuFZXUGNfveqZYoc3E7p84HHJvLctvqhOV1SfUkgxgTGFTNL9sdBMvqBNcLfC8+ygzy/r7tOnZwUVKVr5ubTIgQZTbbBbMLqJPMzZJ3bdcxIiN/fzZh19lRvV4MTn9/5p25eXU6wlO8/+wAvxE8IRUxToa2vGhaV50x4IVI2efuBPMhvGm5DFvOmW4f9GE+6MzJjfPHsqQl86tlfl3f6ZxhOqcm1dnRr6RO0FSSG3PuzqkTNtVtF+Ii87Wry/1lRYRPDosYJjj9wb5SDupys/svLSP7YO5B6NTFNYvDIt7EJXiCVQRPSJy5eH/Pdh0Duthohk6tYzfdHy1t/+HH+4uAGHeFs8a5wNM1PV1Qc3JzGYIvLqrT3bRIULEzzz5kO9mxiw4ToA8yJ8z14Om63A9ik9RuNMORLC/z7O/2JW5YuCtCiEnwBKvGbBeA7Gk0w6VKOTivzl1/1zk3g6rRDK+bYGa7Oxi1EZaTlnl1BiTmUYvQaXcmJJlXZyv5kjrL0+ZECDWIi+oMq130uTPSAp/m5jn33oWHGs2wbsKnuvwJF/qZU77eq0f53lyTH7sFXid0clujGS5UysGS/D/fIOPoeAIctEPXk/fdTsAwekKoWfnTnTIKF9QJIwibAEeY9/K6tg/M3240wzwM7waQMrPiZFHbL+1/vdEMfbrBgowheAIc1Wc5WludwZctOxUBbjEf6qo9v/LUDXVRZgi6OoPQWUYHOMicpxb0+GYrvKcDSNQOYXdbUolrBdhE8AQ4zOywNK9O+/ccbxjA9rYEUdPKVkfUBXXOA3URNAHeMXNYZtUJoC6q857OsmsAiTLXQrPq7BR8QtJ5STXON7CN4AkAkFmm9XzrL5dnILT1cCe9JUlLXCwCAADAZwRPAIBcMXcDpyWVtvya1GiW612UdF2dOW69v5boZAIAAEDWEDwBALBFTzjVtfW/d7OkTrj04L8JlQAAAJBHBE8AAAAAAABIRcF2AQAAAAAAAMgmgicAAAAAAACkguAJAAAAAAAAqSB4AgAAAAAAQCoIngAAAAAAAJAKgicAAAAAAACkguAJAAAAAAAAqSB4AgAAAAAAQCoIngAAAAAAAJAKgicAAADg/2/HjgUAAAAABvlbj2JfYQQALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMBCPAEAAACwEE8AAAAALMQTAAAAAAvxBAAAAMAiEzn2uAfq614AAAAASUVORK5CYII=" alt="Monin" style="height:64px;object-fit:contain;margin-bottom:12px">
  <h2>Billback Processor</h2>
  <p>Enter the team password to continue.</p>
  {error}
  <form method="post" action="/login">
    <input type="password" name="pw" placeholder="Password" autofocus>
    <button type="submit">Sign In</button>
  </form>
</div></body></html>"""

def _get_session(headers):
    """Return session token from Cookie header, or None."""
    cookie = headers.get('Cookie', '')
    for part in cookie.split(';'):
        k, _, v = part.strip().partition('=')
        if k.strip() == 'bb_session':
            return v.strip()
    return None

def _is_authed(headers):
    """Return True if auth is disabled or valid session cookie present."""
    if not APP_PASSWORD:
        return True  # running locally without password
    tok = _get_session(headers)
    return tok in _active_sessions


class BillbackHandler(BaseHTTPRequestHandler):
    """Clean multipart handler using email library for proper filename extraction."""
    def log_message(self, format, *args): pass

    def do_GET(self):
        # ── Auth check ──
        if not _is_authed(self.headers):
            self._serve_login()
            return

        if self.path == '/':
            self.send_response(200)
            self.send_header('Content-Type','text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(HTML_PAGE.encode())
            return
        if self.path.startswith('/download/'):
            did = self.path.split('/')[-1]
            data = _downloads.get(did)
            if data:
                self.send_response(200)
                self.send_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                fname = f'Tellus_Upload_{TODAY}.xlsx'
                self.send_header('Content-Disposition',f'attachment; filename="{fname}"')
                self.send_header('Content-Length', str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return
        self.send_response(404); self.end_headers()

    def _serve_login(self, error=''):
        err_html = f'<p class="err">{error}</p>' if error else ''
        page = LOGIN_PAGE.replace('{error}', err_html).encode()
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.end_headers()
        self.wfile.write(page)

    def do_POST(self):
        # ── Login form submission ──
        if self.path == '/login':
            length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(length).decode()
            from urllib.parse import parse_qs
            params = parse_qs(body)
            pw = params.get('pw', [''])[0]
            if APP_PASSWORD and pw == APP_PASSWORD:
                tok = _secrets.token_hex(24)
                _active_sessions.add(tok)
                self.send_response(302)
                self.send_header('Location', '/')
                self.send_header('Set-Cookie', f'bb_session={tok}; Path=/; HttpOnly; SameSite=Lax')
                self.end_headers()
            else:
                self._serve_login(error='Incorrect password — please try again.')
            return

        if not _is_authed(self.headers):
            self.send_response(403); self.end_headers(); return

        if self.path != '/process':
            self.send_response(404); self.end_headers(); return
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            content_type = self.headers.get('Content-Type', '')

            # Parse multipart using email package
            msg_bytes = f'Content-Type: {content_type}\r\n\r\n'.encode() + body
            msg = email.message_from_bytes(msg_bytes)

            user_config = {}
            file_overrides = {}   # {filename: {program_num, dist_id, customer_ref}}
            file_parts = []

            for part in msg.walk():
                cd = part.get('Content-Disposition', '')
                if not cd: continue
                name_m = re.search(r'name="([^"]+)"', cd)
                fname_m = re.search(r'filename="([^"]+)"', cd)
                field_name = name_m.group(1) if name_m else ''
                filename   = fname_m.group(1) if fname_m else ''
                payload = part.get_payload(decode=True)
                if payload is None: continue

                if field_name == 'config':
                    try: user_config = json.loads(payload.decode())
                    except: pass
                elif field_name in ('file_overrides', 'customer_refs'):
                    # Support both new file_overrides and legacy customer_refs field
                    try:
                        parsed = json.loads(payload.decode())
                        if field_name == 'file_overrides':
                            file_overrides = parsed
                        else:
                            # Legacy: customer_refs was {filename: ref_str}
                            for fn, ref in parsed.items():
                                file_overrides.setdefault(fn, {})['customer_ref'] = ref
                    except: pass
                elif field_name == 'files' and filename:
                    file_parts.append((filename, payload))

            tmpdir = tempfile.mkdtemp()
            results = []
            all_rows = []

            for filename, content in file_parts:
                # Save with original filename for supplier detection
                safe_name = re.sub(r'[^\w.\-]', '_', filename)
                tmp_path = os.path.join(tmpdir, safe_name)
                with open(tmp_path, 'wb') as f:
                    f.write(content)

                fo   = file_overrides.get(filename, {})
                cref = fo.get('customer_ref', '')
                try:
                    supplier, rows = detect_and_parse(tmp_path, user_config, cref, file_override=fo, original_filename=filename)
                    fatal_errs = [r['_error'] for r in rows if '_error' in r]
                    warnings   = [r for r in rows if '_warning' in r]
                    ok         = [r for r in rows if '_error' not in r and '_warning' not in r]
                    all_rows.extend(ok)
                    warn_total = round(sum(w['amount'] for w in warnings), 4)
                    warn_list  = [f"{w['code']} — {w['desc']} (${w['amount']})" for w in warnings]
                    if fatal_errs and not ok:
                        results.append({'file': filename, 'supplier': supplier,
                                        'error': fatal_errs[0], 'rows': 0,
                                        'warnings': [], 'warn_total': 0,
                                        'total_amount': 0.0, 'total_qty': 0})
                    else:
                        total_amount = round(sum(
                            float(r.get('Item Dollar Amount') or 0) for r in ok), 2)
                        total_qty = int(round(sum(
                            float(r.get('Item Volume Qty') or 0) for r in ok)))
                        results.append({'file': filename, 'supplier': supplier,
                                        'rows': len(ok), 'error': None,
                                        'warnings': warn_list, 'warn_total': warn_total,
                                        'total_amount': total_amount, 'total_qty': total_qty})
                except Exception as ex:
                    results.append({'file': filename, 'supplier': 'ERROR',
                                    'error': str(ex), 'rows': 0,
                                    'warnings': [], 'warn_total': 0})

            shutil.rmtree(tmpdir, ignore_errors=True)

            download_id = None
            if all_rows:
                xlsx_bytes = build_output(all_rows)
                download_id = str(uuid.uuid4())[:8]
                _downloads[download_id] = xlsx_bytes

            resp = {'results': results, 'download_id': download_id,
                    'total_rows': len(all_rows)}
            self._respond_json(resp)

        except Exception as e:
            try:
                self._respond_json({'error': str(e), 'trace': traceback.format_exc(),
                                    'results': [], 'download_id': None, 'total_rows': 0})
            except Exception:
                # Last resort — connection may be broken; send a bare minimal body
                try:
                    body = b'{"error":"Internal server error","results":[],"download_id":null,"total_rows":0}'
                    self.send_response(500)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Content-Length', str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                except Exception:
                    pass  # Socket is dead; nothing we can do

    def _respond_json(self, data):
        body = json.dumps(data).encode()
        self.send_response(200)
        self.send_header('Content-Type','application/json')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

def run():
    # When deployed online (Render, Railway, etc.) use the PORT env var.
    # When running locally as .exe, fall back to 8765-8775.
    env_port = os.environ.get('PORT')
    is_web = bool(env_port)

    if is_web:
        port = int(env_port)
        host = '0.0.0.0'
        server = HTTPServer((host, port), BillbackHandler)
        print(f'Monin Billback running on port {port}')
        try:
            server.serve_forever()
        except KeyboardInterrupt:
            pass
    else:
        port = 8765
        host = 'localhost'
        for p in range(8765, 8776):
            try:
                server = HTTPServer((host, p), BillbackHandler)
                port = p
                break
            except OSError:
                continue
        else:
            print('ERROR: Could not find a free port (8765-8775). Close other apps and try again.')
            input('Press Enter to exit.')
            return
        url = f'http://{host}:{port}'
        threading.Timer(1.5, lambda: webbrowser.open(url)).start()
        try:
            server.serve_forever()
        except KeyboardInterrupt:
            pass

if __name__ == '__main__':
    run()
